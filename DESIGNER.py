# -*- coding: utf-8 -*-
"""
Designer de Campos — Fluxos [v3.0] - INTEGRADO COM IMPORTADOR BPMN
- Adicionado: Modais escuros consistentes que substituem as caixas de diálogo brancas do sistema.
- Aprimorado (Importador BPMN): Layout redesenhado com filtros, ordenação e reordenação manual antes de importar.
- Aprimorado (Menu): Contexto redesenhado com visual moderno e destaques de ação.

--- Histórico de Versões Anteriores ---
- Corrigido (UI/Validador): Espaçamento e alinhamento de botões nos avisos do Validador foram corrigidos.
- Corrigido (UI): Botão "Cancelar" em diálogos agora tem um estilo visual de secundário, não de desativado.
- Implementado: Opção contextual "Definir/Editar Tipos de Documento..." re-adicionada ao menu para campos "Anexo".
- Corrigido (Menu de Contexto): Clique com o botão direito agora é detetado em toda a área da linha.
- Corrigido (Menu de Contexto): Menu agora desaparece corretamente ao clicar fora dele.
"""

from __future__ import annotations
import os, re, uuid, json, datetime, sys, traceback, tkinter as tk, unicodedata, importlib
# Imports adicionados para o importador de BPMN
import io
import zipfile
import xml.etree.ElementTree as ET
from collections import defaultdict
from dataclasses import dataclass, field as dc_field
from typing import Dict, List, Optional, Tuple, Set, Any, Callable
from tkinter import filedialog, messagebox, ttk
import customtkinter as ctk
from urllib.parse import quote, unquote
import time


def _ensure_source_integrity() -> None:
    """Interrompe a execução se o arquivo fonte estiver corrompido."""
    try:
        path = os.path.abspath(__file__)
        with open(path, "rb") as fh:
            snippet = fh.read(256).lstrip()
    except (OSError, TypeError):
        return

    marker = b"429: Too Many Requests"
    if snippet.startswith(marker):
        sys.stderr.write(
            "\nO arquivo 'Designer_Atualizado_Validador_BPMN 3.0.py' foi baixado de forma incompleta e contém uma resposta de erro (429).\n"
            "Faça o download novamente do arquivo completo antes de tentar abrir o programa.\n\n"
        )
        raise SystemExit(1)


_ensure_source_integrity()

try:
    import ctypes
    from ctypes import wintypes
except ImportError:
    ctypes = None
    wintypes = None

# Visão HTML
try:
    from tkinterweb import HtmlFrame
except Exception:
    HtmlFrame = None  # checado ao abrir a visão HTML

APP_VERSION = "v3.0"
CONFIG_PATH = os.path.join(os.path.expanduser("~"), ".designer_campos_config.json")
TEMPLATES_DB_PATH = os.path.join(os.path.expanduser("~"), ".designer_campos_templates.json")

DARK_BG = "#111111"
DARK_BG2 = "#161616"
DARK_BG3 = "#1a1a1a"
HILIGHT_ORIGIN_BG = "#171a1f"


@dataclass(frozen=True)
class ThemePalette:
    """Cores reutilizáveis para superfícies, botões e etiquetas."""

    background: str = DARK_BG
    background_alt: str = DARK_BG2
    background_raised: str = DARK_BG3
    surface: str = "#1b1f24"
    surface_hover: str = "#232a34"
    surface_border: str = "#2b3138"
    surface_low: str = "#14171c"
    field_surface: str = "#1d232b"
    field_hover: str = "#242c35"
    field_border: str = "#2a3036"
    badge_bg: str = "#223044"
    badge_text: str = "#dbe2ec"
    chip_bg: str = "#223e63"
    chip_text: str = "#82b7ff"
    text_muted: str = "#9aa4b2"
    text_subtle: str = "#9aa5b1"
    menu_border: str = "#333333"
    menu_separator: str = "#333333"
    menu_hover: str = "#2a2d2e"
    danger_hover: str = "#472427"
    danger_text: str = "#fda4af"
    button_secondary_bg: str = "#1f2833"
    button_secondary_hover: str = "#24303f"
    button_secondary_text: str = "#dbe2ec"
    button_secondary_border: str = "#334155"
    menu_text: str = "#dce4ee"


@dataclass(frozen=True)
class ButtonVariantConfig:
    fg_color: Optional[str] = None
    hover_color: Optional[str] = None
    text_color: Optional[str] = None
    border_color: Optional[str] = None
    border_width: Optional[int] = None


class ThemeCatalog:
    """Mantém tokens reutilizáveis de estilo."""

    def __init__(self):
        self.palette = ThemePalette()
        self.button_variants = {
            "secondary": ButtonVariantConfig(
                fg_color=self.palette.button_secondary_bg,
                hover_color=self.palette.button_secondary_hover,
                text_color=self.palette.button_secondary_text,
                border_color=self.palette.button_secondary_border,
                border_width=1,
            ),
            "ghost": ButtonVariantConfig(
                fg_color="transparent",
                hover_color=self.palette.surface_hover,
                border_color=self.palette.surface_border,
                border_width=1,
            ),
            "menu": ButtonVariantConfig(
                fg_color="transparent",
                hover_color=self.palette.menu_hover,
                text_color=self.palette.menu_text,
                border_width=0,
            ),
        }

    def color(self, key: str) -> str:
        return getattr(self.palette, key)

    def apply_button(self, widget: ctk.CTkButton, variant: str) -> None:
        config = self.button_variants.get(variant)
        if not config:
            return
        options = {k: v for k, v in config.__dict__.items() if v is not None}
        if not options:
            return
        for key, value in options.items():
            try:
                widget.configure(**{key: value})
            except (tk.TclError, AttributeError):
                continue


THEME = ThemeCatalog()


def _safe_configure(widget: tk.Misc, **kwargs) -> None:
    """Aplica opções a widgets CustomTkinter ignorando chaves desconhecidas."""
    for key, value in kwargs.items():
        try:
            widget.configure(**{key: value})
        except (tk.TclError, AttributeError, TypeError):
            continue

SHORTCUT_SECTIONS = [
    (
        "Fluxo e arquivos",
        [
            ("Ctrl+Shift+N", "Criar um fluxo em branco"),
            ("Ctrl+O", "Abrir um fluxo salvo"),
            ("Ctrl+S", "Salvar as alterações do fluxo atual"),
            ("Ctrl+E", "Exportar o fluxo para Excel"),
        ],
    ),
    (
        "Templates",
        [
            ("Ctrl+T", "Abrir a biblioteca de templates"),
            ("Ctrl+Shift+T", "Salvar o fluxo como um novo template"),
        ],
    ),
    (
        "Edição",
        [
            ("Ctrl+Z", "Desfazer a última alteração"),
            ("Ctrl+Y", "Refazer o que foi desfeito"),
            ("Ctrl+Shift+Z", "Refazer o que foi desfeito"),
        ],
    ),
    (
        "Navegação",
        [
            ("F1", "Mostrar esta janela de atalhos"),
            ("Shift + rolagem do mouse", "Deslocar horizontalmente a grade de campos"),
        ],
    ),
]

# --- Funções de Parsing do BPMN (Copiado do Simulador) ---
def strip_ns(tag: str) -> str:
    """Remove o namespace do XML de uma tag (ex: {http://...}Task -> Task)."""
    return tag.split('}')[-1] if '}' in tag else tag

def normalize_label(s: str) -> str:
    """Limpa e normaliza os nomes/rótulos extraídos do XML."""
    if not s: return ""
    return " ".join(s.replace("\r", " ").replace("\n", " ").split())

def parse_bizagi_group_by_diagram(bpm_path: str):
    """
    Processa um ficheiro .bpm e extrai todos os nós, transições e nomes de diagrama,
    usando a lógica de Pool para identificar o nome do diagrama.
    """
    nodes_by_diag = {}
    transitions_by_diag = {}
    diagrams_labels = []

    with zipfile.ZipFile(bpm_path) as z:
        diags = [n for n in z.namelist() if n.lower().endswith(".diag")]
        if not diags:
            raise RuntimeError("Nenhum .diag encontrado dentro do .bpm")
        for d in diags:
            try:
                bio = io.BytesIO(z.read(d))
                if not zipfile.is_zipfile(bio): continue
                with zipfile.ZipFile(bio) as inner:
                    if "Diagram.xml" not in inner.namelist(): continue
                    root = ET.fromstring(inner.read("Diagram.xml"))
            except Exception:
                diagrams_labels.append((d, d)) # Fallback em caso de erro de leitura
                continue

            ns = {'xpdl': root.tag.split('}')[0].strip('{')}
            
            pool_names = [
                normalize_label(p.get("Name", "") or "")
                for p in root.findall(".//xpdl:Pool", ns)
            ]
            label = d 
            if pool_names:
                preferred = [p for p in pool_names if p and p.lower() != "processo principal"]
                label = preferred[0] if preferred else (pool_names[0] or d)

            diagrams_labels.append((d, label))

            nodes = {}
            transitions = []
            for act in root.findall('.//xpdl:Activity', ns):
                aid = act.get('Id', '')
                name = normalize_label(act.get('Name', '') or '')
                route = act.find('xpdl:Route', ns)
                implementation = act.find('xpdl:Implementation', ns)
                typ = 'Route' if route is not None else ('Task' if implementation is not None else 'Activity')
                nodes[aid] = {"id": aid, "name": name, "type": typ, "has_implementation": implementation is not None}

            for tr in root.findall('.//xpdl:Transition', ns):
                transitions.append({
                    "from": tr.get('From', ''), "to": tr.get('To', ''),
                    "name": normalize_label(tr.get('Name', ''))
                })
            nodes_by_diag[d] = nodes
            transitions_by_diag[d] = transitions

    diagrams_labels.sort(key=lambda x: x[1])
    return diagrams_labels, nodes_by_diag, transitions_by_diag

def build_task_fields_for_diagram(nodes, transitions):
    """Filtra tarefas e extrai campos dos gateways."""
    out_by, in_by = {}, {}
    for t in transitions:
        out_by.setdefault(t["from"], []).append(t)
        in_by.setdefault(t["to"], []).append(t)

    tasks = [n for n in nodes.values() if n["type"] == "Task" and n.get("has_implementation")]
    task_fields = {t["id"]: [] for t in tasks}

    for gw in [n for n in nodes.values() if n["type"] == "Route"]:
        incoming, outgoing = in_by.get(gw["id"], []), out_by.get(gw["id"], [])
        opts = [t["name"] for t in outgoing if t.get("name")]
        if not opts and outgoing:
            opts = [nodes.get(t["to"], {}).get("name", "") for t in outgoing]
            opts = [o for o in opts if o]
        if not opts and len(outgoing) == 2:
            opts = ["Sim", "Não"]
        
        tipo = "Lista"
        if set(o.lower() for o in opts) in [{"sim", "não"}, {"sim", "nao"}]:
            tipo = "Lista (Sim/Não)"
        
        campo_nome = normalize_label(gw.get("name", ""))
        for inc in incoming:
            src = nodes.get(inc["from"])
            if src and src["type"] == "Task" and src.get("has_implementation"):
                field_data = {"id": f"{src['id']}_{gw['id']}", "campo": campo_nome, "tipo": tipo, "opcoes": opts}
                task_fields[src["id"]].append(field_data)

    return tasks, task_fields
    

def _uid() -> str:
    return uuid.uuid4().hex[:8]

def _now_iso() -> str:
    try:
        return datetime.datetime.utcnow().replace(microsecond=0).isoformat() + "Z"
    except Exception:
        return ""

def _best_desktop_dir() -> str:
    home = os.path.expanduser("~")
    for path in [
        os.path.join(home, "Desktop"),
        os.path.join(home, "OneDrive", "Desktop"),
        os.path.join(home, "OneDrive", "Área de Trabalho"),
        os.path.join(home, "OneDrive", "Area de Trabalho"),
    ]:
        try:
            os.makedirs(path, exist_ok=True)
            return path
        except Exception:
            pass
    return home


def _solid_color() -> str:
    try:
        fg = ctk.ThemeManager.theme["CTkButton"]["fg_color"]
        if isinstance(fg, (list, tuple)) and len(fg) >= 2:
            return fg[1]
        return fg
    except Exception:
        return "#1f6aa5"

def _apply_secondary_style(btn: ctk.CTkButton) -> None:
    """Aplica um estilo secundário consistente para botões de cancelar."""
    try:
        THEME.apply_button(btn, "secondary")
    except Exception:
        pass


def _animate_fade_in(win: tk.Misc, *, duration: int = 140, steps: int = 6) -> None:
    # Versão otimizada: remove o loop de animação que bloqueia a thread
    try:
        win.attributes("-alpha", 1.0)
        win.deiconify()
    except Exception:
        pass

_APP_ROOT: Optional["App"] = None
_MSGBOX_FALLBACK_ROOT: Optional[ctk.CTk] = None

if sys.platform == "win32" and ctypes is not None and wintypes is not None:

    class MONITORINFO(ctypes.Structure):
        _fields_ = [
            ("cbSize", wintypes.DWORD),
            ("rcMonitor", wintypes.RECT),
            ("rcWork", wintypes.RECT),
            ("dwFlags", wintypes.DWORD),
        ]


def _get_monitor_bounds_for_window(win: tk.Misc) -> Tuple[int, int, int, int]:
    if sys.platform == "win32" and ctypes is not None and wintypes is not None:
        hwnd = None
        try:
            hwnd = int(win.winfo_id())
        except Exception:
            pass
        if hwnd:
            try:
                monitor = ctypes.windll.user32.MonitorFromWindow(hwnd, 2)
            except Exception:
                monitor = None
            if monitor:
                info = MONITORINFO()
                info.cbSize = ctypes.sizeof(MONITORINFO)
                if ctypes.windll.user32.GetMonitorInfoW(monitor, ctypes.byref(info)):
                    rect = info.rcWork
                    return rect.left, rect.top, rect.right, rect.bottom

    try:
        screen_w = win.winfo_screenwidth()
        screen_h = win.winfo_screenheight()
    except Exception:
        screen_w = screen_h = 0
    return 0, 0, screen_w, screen_h


def _get_monitor_bounds_for_point(x: int, y: int, fallback: tk.Misc) -> Tuple[int, int, int, int]:
    if sys.platform == "win32" and ctypes is not None and wintypes is not None:
        try:
            point = wintypes.POINT(x, y)
            monitor = ctypes.windll.user32.MonitorFromPoint(point, 2)
        except Exception:
            monitor = None
        if monitor:
            info = MONITORINFO()
            info.cbSize = ctypes.sizeof(MONITORINFO)
            if ctypes.windll.user32.GetMonitorInfoW(monitor, ctypes.byref(info)):
                rect = info.rcWork
                return rect.left, rect.top, rect.right, rect.bottom

    return _get_monitor_bounds_for_window(fallback)


def _clamp_to_bounds(x: int, y: int, width: int, height: int, bounds: Tuple[int, int, int, int]) -> Tuple[int, int]:
    left, top, right, bottom = bounds
    max_x = max(left, right - width)
    max_y = max(top, bottom - height)
    clamped_x = min(max(x, left), max_x)
    clamped_y = min(max(y, top), max_y)
    return clamped_x, clamped_y


def _center_within(master: Optional[tk.Misc], width: int, height: int) -> Optional[Tuple[int, int]]:
    """Calcula a posição central relativa a uma janela mestre.

    A função é extremamente defensiva porque alguns builds antigos do
    CustomTkinter/Tk reportam dimensões inconsistentes antes do
    mainloop. Se não conseguirmos dados confiáveis, retornamos ``None``
    e deixamos o chamador tratar o fallback padrão (centro da tela).
    """

    if master is None:
        return None

    try:
        target = master.winfo_toplevel()
    except Exception:
        target = master
    if target is None:
        target = master

    try:
        if not target.winfo_exists():
            return None
    except Exception:
        return None

    try:
        target.update_idletasks()
    except Exception:
        pass

    def _safe_query(fetch: Callable[[], int]) -> int:
        try:
            return int(fetch())
        except Exception:
            return 0

    master_x, master_y, master_w, master_h = 0, 0, 0, 0
    
    try:
        window_state = target.state()
    except Exception:
        window_state = 'normal'

    # SOLUÇÃO: Se a janela estiver maximizada ('zoomed'), usar
    # as dimensões do monitor, pois winfo_width/height retornam '1'.
    if window_state == 'zoomed':
        try:
            bounds = _get_monitor_bounds_for_window(target)
            master_x = bounds[0]
            master_y = bounds[1]
            master_w = bounds[2] - bounds[0]
            master_h = bounds[3] - bounds[1]
        except Exception:
            # Fallback se a API do monitor falhar
            master_w = _safe_query(target.winfo_screenwidth)
            master_h = _safe_query(target.winfo_screenheight)
            master_x = 0
            master_y = 0
    else:
        # Janela em estado 'normal', usar a lógica original (com a correção do geom_h)
        geom_w = geom_h = 0
        geom_x = geom_y = None
        try:
            geometry = target.winfo_geometry()
        except Exception:
            geometry = ""

        if geometry:
            try:
                match = re.match(r"^(\d+)x(\d+)([+-]\d+)([+-]\d+)$", geometry)
                if match:
                    geom_w = int(match.group(1))
                    geom_h = int(match.group(2))
                    geom_x = int(match.group(3))
                    geom_y = int(match.group(4))
                else:
                    size_part = geometry.split("+", 1)[0]
                    geom_w_str, geom_h_str = size_part.split("x", 1)
                    geom_w = int(geom_w_str)
                    geom_h = int(geom_h_str) # Correção da tentativa anterior
            except Exception:
                geom_w = geom_h = 0

        master_x = _safe_query(target.winfo_rootx)
        master_y = _safe_query(target.winfo_rooty)
        master_w = _safe_query(target.winfo_width)
        master_h = _safe_query(target.winfo_height)

        if master_w <= 1 or master_h <= 1:
            master_w = max(master_w, geom_w)
            master_h = max(master_h, geom_h)

        if master_w <= 1 or master_h <= 1:
            req_w = _safe_query(target.winfo_reqwidth)
            req_h = _safe_query(target.winfo_reqheight)
            master_w = max(master_w, req_w)
            master_h = max(master_h, req_h)

        if (master_x <= 1 or master_y <= 1) and geom_x is not None and geom_y is not None:
            master_x = geom_x
            master_y = geom_y
    
    if master_w <= 1 or master_h <= 1:
        return None

    centered_x = master_x + (master_w - width) // 2
    centered_y = master_y + (master_h - height) // 2
    return int(centered_x), int(centered_y)
_ICON_SYMBOLS = {
    "info": "ℹ️",
    "warning": "⚠️",
    "error": "❌",
    "question": "❓",
}


def _apply_dark_title_bar(win: tk.Misc) -> None:
    if sys.platform != "win32" or ctypes is None:
        return
    try:
        win.update_idletasks()
        hwnd = ctypes.windll.user32.GetParent(win.winfo_id())
        DWMWA_USE_IMMERSIVE_DARK_MODE = 20
        value = ctypes.c_int(1)
        ctypes.windll.dwmapi.DwmSetWindowAttribute(
            hwnd,
            DWMWA_USE_IMMERSIVE_DARK_MODE,
            ctypes.byref(value),
            ctypes.sizeof(value),
        )
    except Exception:
        pass


def _resolve_master(parent: Optional[tk.Misc]) -> tk.Misc:
    global _MSGBOX_FALLBACK_ROOT
    if parent is not None:
        try:
            candidate = parent.winfo_toplevel()
        except Exception:
            candidate = parent
        if candidate is None:
            candidate = parent
        try:
            if candidate.winfo_exists():
                return candidate
        except Exception:
            pass
    if _APP_ROOT is not None and _APP_ROOT.winfo_exists():
        return _APP_ROOT
    if _MSGBOX_FALLBACK_ROOT is None or not _MSGBOX_FALLBACK_ROOT.winfo_exists():
        try:
            _MSGBOX_FALLBACK_ROOT = ctk.CTk()
        except Exception:
            _MSGBOX_FALLBACK_ROOT = tk.Tk()
        _MSGBOX_FALLBACK_ROOT.withdraw()
    return _MSGBOX_FALLBACK_ROOT


def _show_messagebox(
    title: str,
    message: str,
    *,
    icon: str = "info",
    buttons: Tuple[str, ...] = ("OK",),
    default: Optional[str] = None,
    parent: Optional[tk.Misc] = None,
) -> str:
    master = _resolve_master(parent)
    dialog = ctk.CTkToplevel(master)
    dialog.title(title or "")
    dialog.transient(master)
    dialog.grab_set()
    dialog.resizable(False, False)

    try:
        dialog.configure(fg_color=THEME.color("background_alt"))
    except Exception:
        pass

    message_font = ctk.CTkFont(size=14)
    raw_lines = message.splitlines() if message else [""]
    measured = [message_font.measure(line) for line in raw_lines if line]
    longest_line = max(measured) if measured else message_font.measure(message or "")
    wrap_length = min(560, max(320, longest_line + 40))

    container = ctk.CTkFrame(dialog, fg_color="transparent")
    container.pack(fill="both", expand=True, padx=28, pady=24)

    icon_frame = ctk.CTkFrame(container, fg_color=THEME.color("surface"), corner_radius=12)
    icon_frame.pack(fill="both", expand=True, pady=(0, 20))

    symbol = _ICON_SYMBOLS.get(icon, _ICON_SYMBOLS["info"])
    ctk.CTkLabel(
        icon_frame,
        text=symbol,
        font=ctk.CTkFont(size=28),
        width=54,
        anchor="center",
    ).pack(side="left", padx=(18, 12), pady=18)

    message_label = ctk.CTkLabel(
        icon_frame,
        text=message,
        justify="left",
        anchor="w",
        wraplength=wrap_length,
        font=message_font,
    )
    message_label.pack(side="left", fill="both", expand=True, padx=(0, 20), pady=18)

    button_row = ctk.CTkFrame(container, fg_color="transparent")
    button_row.pack(fill="x")

    result = {"value": default or (buttons[0] if buttons else "")}

    primary_color = _solid_color()
    default_label = default or (buttons[0] if buttons else "")
    cancel_label = buttons[-1] if buttons else default_label

    def close_with(value: str) -> None:
        result["value"] = value
        dialog.destroy()

    for text in buttons:
        cmd = lambda v=text: close_with(v)
        btn = ctk.CTkButton(
            button_row,
            text=text,
            width=130,
            height=36,
            command=cmd,
        )
        if text != default_label:
            _apply_secondary_style(btn)
        else:
            btn.configure(fg_color=primary_color)
        btn.pack(side="right", padx=(8, 0))

    dialog.bind("<Return>", lambda _e: close_with(default_label))
    dialog.bind("<Escape>", lambda _e: close_with(cancel_label))
    dialog.protocol("WM_DELETE_WINDOW", lambda: close_with(cancel_label))

    dialog.update_idletasks()
    master.update_idletasks()
    width = dialog.winfo_width()
    height = dialog.winfo_height()
    desired_width = max(width, wrap_length + 170)
    desired_height = max(height, message_label.winfo_reqheight() + 160)
    if desired_width > width or desired_height > height:
        width = max(width, desired_width)
        height = max(height, desired_height)
        dialog.geometry(f"{width}x{height}")
        dialog.update_idletasks()
    dialog.minsize(width, height)

    try:
        anchor = master if master is not None and master.winfo_exists() else dialog
        bounds = _get_monitor_bounds_for_window(anchor)
        centered = _center_within(master, width, height)

        if centered is None:
            screen_w = dialog.winfo_screenwidth()
            screen_h = dialog.winfo_screenheight()
            x = (screen_w - width) // 2
            y = (screen_h - height) // 2
        else:
            x, y = centered

        x, y = _clamp_to_bounds(int(x), int(y), width, height, bounds)
    except Exception:
        screen_w = dialog.winfo_screenwidth()
        screen_h = dialog.winfo_screenheight()
        x = (screen_w - width) // 2
        y = (screen_h - height) // 2

    dialog.geometry(f"{width}x{height}+{x}+{y}")
    _apply_dark_title_bar(dialog)
    _animate_fade_in(dialog)

    dialog.focus_set()
    dialog.wait_window()
    return result["value"]


def _patch_messageboxes() -> None:
    def _show_info(title: str, message: str, **kwargs) -> str:
        _show_messagebox(title, message, icon="info", buttons=("OK",), parent=kwargs.get("parent"))
        return "ok"

    def _show_warning(title: str, message: str, **kwargs) -> str:
        _show_messagebox(title, message, icon="warning", buttons=("Entendi",), parent=kwargs.get("parent"))
        return "ok"

    def _show_error(title: str, message: str, **kwargs) -> str:
        _show_messagebox(title, message, icon="error", buttons=("Fechar",), parent=kwargs.get("parent"))
        return "ok"

    def _ask_yes_no(title: str, message: str, **kwargs) -> bool:
        default = kwargs.get("default", "yes").lower()
        default_label = "Sim" if default in {"yes", "y", "ok", "true"} else "Não"
        response = _show_messagebox(
            title,
            message,
            icon="question",
            buttons=("Sim", "Não"),
            default=default_label,
            parent=kwargs.get("parent"),
        )
        return response == "Sim"

    messagebox.showinfo = _show_info  # type: ignore[assignment]
    messagebox.showwarning = _show_warning  # type: ignore[assignment]
    messagebox.showerror = _show_error  # type: ignore[assignment]
    messagebox.askyesno = _ask_yes_no  # type: ignore[assignment]


_patch_messageboxes()

# ===== Modelo de dados =====
@dataclass
class Condition:
    src_field: str
    op: str
    value: str

@dataclass
class ObjectFieldDef:
    name: str
    ftype: str = "Texto"
    options: str = ""
    required: bool = False
    readonly: bool = True
    group: str = ""
    order: int = 0
    note: str = ""

@dataclass
class Field:
    id: str
    name: str = "Novo campo"
    ftype: str = "Texto"
    required: bool = False
    readonly: bool = False
    info: str = ""
    options: str = ""
    note: str = ""
    origin_task: Optional[str] = None
    origin_field: Optional[str] = None
    name_locked: bool = False
    name_lock_reason: str = ""   # "", "objeto", "origem"
    name_before_obj: str = ""
    name_before_origin: str = ""
    obj_type: str = ""
    cond: List[Condition] = dc_field(default_factory=list)

@dataclass
class Task:
    id: str
    name: str
    fields: List[Field] = dc_field(default_factory=list)

@dataclass
class ProjectModel:
    flow_name: str = "Novo fluxo"
    tasks: List[Task] = dc_field(default_factory=list)
    object_type: str = ""
    object_schema: List[ObjectFieldDef] = dc_field(default_factory=list)

    def to_dict(self) -> dict:
        return {
            "flow_name": self.flow_name,
            "object_type": self.object_type,
            "object_schema": [vars(ofd) for ofd in self.object_schema],
            "tasks": [
                {
                    "id": t.id, "name": t.name,
                    "fields": [
                        {
                            "id": f.id, "name": f.name, "ftype": f.ftype,
                            "required": f.required, "readonly": f.readonly,
                            "info": f.info, "options": f.options, "note": f.note,
                            "origin_task": f.origin_task, "origin_field": f.origin_field,
                            "name_locked": f.name_locked, "name_lock_reason": f.name_lock_reason,
                            "name_before_obj": f.name_before_obj, "name_before_origin": f.name_before_origin,
                            "obj_type": f.obj_type,
                            "cond": [{"src_field": c.src_field, "op": c.op, "value": c.value} for c in f.cond],
                        } for f in t.fields
                    ],
                } for t in self.tasks
            ]
        }

    @staticmethod
    def from_dict(d: dict) -> "ProjectModel":
        flow_name = d.get("flow_name", "Novo fluxo")
        object_type = d.get("object_type", "")
        object_schema = []
        for x in d.get("object_schema", []):
            try:
                object_schema.append(ObjectFieldDef(**x))
            except Exception:
                pass
        tasks: List[Task] = []
        for td in d.get("tasks", []):
            fields: List[Field] = []
            for fd in td.get("fields", []):
                fields.append(Field(
                    id=fd.get("id", _uid()),
                    name=fd.get("name", "Campo"),
                    ftype=fd.get("ftype", "Texto"),
                    required=bool(fd.get("required", False)),
                    readonly=bool(fd.get("readonly", False)),
                    info=fd.get("info", ""),
                    options=fd.get("options", ""),
                    note=fd.get("note", ""),
                    origin_task=fd.get("origin_task"),
                    origin_field=fd.get("origin_field"),
                    name_locked=bool(fd.get("name_locked", False)),
                    name_lock_reason=fd.get("name_lock_reason", ""),
                    name_before_obj=fd.get("name_before_obj", ""),
                    name_before_origin=fd.get("name_before_origin", ""),
                    obj_type=fd.get("obj_type", ""),
                    cond=[Condition(**c) for c in fd.get("cond", [])],
                ))
            tasks.append(Task(id=td.get("id", _uid()), name=td.get("name", "Tarefa"), fields=fields))
        return ProjectModel(flow_name=flow_name, tasks=tasks, object_type=object_type, object_schema=object_schema)

# ===== Templates locais =====
class TemplateStore:
    def __init__(self, path: str = TEMPLATES_DB_PATH):
        self.path = path
        self._ensure_file()

    def _ensure_file(self):
        if not os.path.exists(self.path):
            self._write({"templates": []})

    def _read(self) -> dict:
        try:
            with open(self.path, "r", encoding="utf-8") as f:
                return json.load(f)
        except Exception:
            return {"templates": []}

    def _write(self, data: dict):
        with open(self.path, "w", encoding="utf-8") as f:
            json.dump(data, f, ensure_ascii=False, indent=2)

    def list_all(self) -> List[dict]:
        data = self._read()
        items = data.get("templates", [])
        items.sort(key=lambda x: x.get("updated_at",""), reverse=True)
        return items

    def save_template(self, name: str, project: ProjectModel, replace: bool = False, create_copy_if_exists: bool = False) -> dict:
        data = self._read()
        items = data.get("templates", [])
        now = _now_iso()
        existing = next((t for t in items if t.get("name","") == name), None)
        if existing and replace:
            existing["project"] = project.to_dict()
            existing["updated_at"] = now
            self._write(data)
            return existing
        if existing and create_copy_if_exists:
            n = 2
            base = name
            new_name = f"{base} ({n})"
            while any(t.get("name","") == new_name for t in items):
                n += 1
                new_name = f"{base} ({n})"
            name = new_name
        entry = {"id": _uid(), "name": name, "created_at": now, "updated_at": now, "project": project.to_dict()}
        items.append(entry)
        data["templates"] = items
        self._write(data)
        return entry

    def rename(self, tmpl_id: str, new_name: str) -> bool:
        data = self._read(); items = data.get("templates", [])
        if any(t.get("name","") == new_name and t.get("id") != tmpl_id for t in items):
            return False
        for t in items:
            if t.get("id") == tmpl_id:
                t["name"] = new_name
                t["updated_at"] = _now_iso()
                self._write(data)
                return True
        return False

    def delete(self, tmpl_id: str) -> bool:
        data = self._read(); items = data.get("templates", [])
        n = len(items)
        items = [t for t in items if t.get("id") != tmpl_id]
        if len(items) == n:
            return False
        data["templates"] = items
        self._write(data); return True

    def get(self, tmpl_id: str) -> Optional[dict]:
        data = self._read()
        for t in data.get("templates", []):
            if t.get("id") == tmpl_id:
                return t
        return None

# ===== Menu de Contexto Customizado =====
class CustomContextMenu(ctk.CTkToplevel):
    def __init__(self, master, event, field: Field, app_instance: "App"):
        super().__init__(master)
        self.app = app_instance
        self.field = field

        self.withdraw()
        self.overrideredirect(True)
        self.attributes("-topmost", True)
        try:
            self.transient(app_instance)
        except Exception:
            pass

        try:
            self.configure(fg_color="transparent")
        except Exception:
            pass

        outer_padx = 4
        outer_pady = 6
        border_width = 1

        menu_frame = ctk.CTkFrame(
            self,
            corner_radius=8,
            fg_color=THEME.color("surface"),
        )
        _safe_configure(
            menu_frame,
            border_width=border_width,
            border_color=THEME.color("menu_border"),
        )
        menu_frame.pack(fill="both", expand=True, padx=outer_padx, pady=outer_pady)

        options = self._build_options()

        for entry in options:
            if entry is None:
                ctk.CTkFrame(menu_frame, height=1, fg_color=THEME.color("menu_separator")).pack(fill="x", padx=6, pady=4)
                continue

            text, command = entry
            btn = ctk.CTkButton(
                menu_frame,
                text=text,
                command=self._create_command(command),
                anchor="w",
                corner_radius=4,
                height=32,
            )
            THEME.apply_button(btn, "menu")
            if text == "Excluir Campo":
                btn.configure(text_color=THEME.color("danger_text"), hover_color=THEME.color("danger_hover"))
            btn.pack(fill="x", padx=6, pady=2)

        self.update_idletasks()
        content_width = menu_frame.winfo_reqwidth()
        content_height = menu_frame.winfo_reqheight()
        width = content_width + (outer_padx * 2) + (border_width * 2)
        height = content_height + (outer_pady * 2) + (border_width * 2)
        bounds = _get_monitor_bounds_for_point(event.x_root, event.y_root, self.app)
        x, y = _clamp_to_bounds(event.x_root, event.y_root, width, height, bounds)
        self.geometry(f"{width}x{height}+{int(x)}+{int(y)}")
        self.deiconify()
        self.lift()

        _animate_fade_in(self, duration=90, steps=5)
        self.after(10, self.focus_force)
        self.bind("<Escape>", lambda _e: self.destroy())

        self.bind_id = self.app.bind("<Button-1>", self._check_if_outside, add='+')
        self.bind_id_secondary = self.app.bind("<Button-3>", self._check_if_outside, add='+')
        self.bind("<FocusOut>", lambda _e: self.destroy())

    def _build_options(self):
        """Retorna a lista de ações exibidas no menu contextual."""

        base_actions = [
            ("Copiar", lambda: self.app._copy_single_field(self.field)),
            ("Recortar", lambda: self.app._cut_single_field(self.field)),
            ("Duplicar Campo", lambda: self.app._duplicate_field(self.field)),
        ]

        if self.field.ftype == "Anexo":
            base_actions.append(
                (
                    "Definir/Editar Tipos de Documento...",
                    lambda: self.app._open_attachment_type_editor(self.field),
                )
            )

        move_actions = [
            ("Mover para o Topo", lambda: self.app._move_field_to_top(self.field)),
            ("Mover para o Fim", lambda: self.app._move_field_to_end(self.field)),
        ]

        destructive_actions = [
            ("Excluir Campo", lambda: self.app._delete_field(self.field.id)),
        ]

        sections = [base_actions, move_actions, destructive_actions]
        flattened = []
        for section in sections:
            if not section:
                continue
            if flattened:
                flattened.append(None)
            flattened.extend(section)

        return flattened

    def _create_command(self, func):
        def wrapper():
            self.destroy()
            if func:
                func()
        return wrapper

    def _check_if_outside(self, event):
        x, y = self.winfo_x(), self.winfo_y()
        w, h = self.winfo_width(), self.winfo_height()
        if not (x < event.x_root < x + w and y < event.y_root < y + h):
            self.destroy()

    def destroy(self):
        if hasattr(self, 'bind_id'):
            self.app.unbind("<Button-1>", self.bind_id)
        if hasattr(self, 'bind_id_secondary'):
            self.app.unbind("<Button-3>", self.bind_id_secondary)
        if getattr(self.app, "context_menu", None) is self:
            self.app.context_menu = None
        super().destroy()


class ShortcutOverlay(ctk.CTkToplevel):
    def __init__(self, master: "App", sections):
        super().__init__(master)
        self.app = master
        self.sections = sections

        self.title("Atalhos do teclado")
        self.resizable(False, True)

        try:
            self.configure(fg_color=THEME.color("background_alt"))
        except Exception:
            pass

        header = ctk.CTkFrame(self, fg_color="transparent")
        header.pack(fill="x", padx=20, pady=(20, 10))
        ctk.CTkLabel(
            header,
            text="Atalhos úteis",
            font=ctk.CTkFont(size=20, weight="bold"),
            anchor="w",
        ).pack(fill="x")
        ctk.CTkLabel(
            header,
            text="Use estas combinações para acelerar o trabalho no designer.",
            text_color=THEME.color("text_subtle"),
            anchor="w",
        ).pack(fill="x", pady=(4, 0))

        body = ctk.CTkScrollableFrame(self, fg_color="transparent")
        body.pack(fill="both", expand=True, padx=20, pady=(0, 10))

        for title, entries in self.sections:
            card = ctk.CTkFrame(body, fg_color=THEME.color("surface"), corner_radius=12)
            card.pack(fill="x", pady=8)
            ctk.CTkLabel(
                card,
                text=title,
                font=ctk.CTkFont(size=16, weight="bold"),
                anchor="w",
            ).pack(fill="x", padx=16, pady=(14, 6))

            for keys, description in entries:
                row = ctk.CTkFrame(card, fg_color="transparent")
                row.pack(fill="x", padx=16, pady=4)
                ctk.CTkLabel(
                    row,
                    text=keys,
                    font=ctk.CTkFont(size=14, weight="bold"),
                    width=170,
                    anchor="w",
                ).pack(side="left")
                ctk.CTkLabel(
                    row,
                    text=description,
                    font=ctk.CTkFont(size=14),
                    justify="left",
                    anchor="w",
                    wraplength=320,
                    text_color=THEME.color("text_muted"),
                ).pack(side="left", fill="x", expand=True, padx=(12, 0))

        footer = ctk.CTkFrame(self, fg_color="transparent")
        footer.pack(fill="x", padx=20, pady=(0, 16))
        btn_close = ctk.CTkButton(footer, text="Fechar", command=self.destroy, width=140)
        _apply_secondary_style(btn_close)
        btn_close.pack(side="right")

        self.bind("<Escape>", lambda _e: self.destroy())

        self.update_idletasks()
        width = max(520, self.winfo_reqwidth())
        height = min(660, max(420, self.winfo_reqheight()))
        if hasattr(master, "_center_toplevel"):
            master._center_toplevel(self, width, height, transient=True)
        else:
            screen_w = self.winfo_screenwidth(); screen_h = self.winfo_screenheight()
            x = max(0, (screen_w - width) // 2)
            y = max(0, (screen_h - height) // 2)
            self.geometry(f"{width}x{height}+{x}+{y}")
            _apply_dark_title_bar(self)
            _animate_fade_in(self)

        self.after(50, lambda: self.focus_force())

    def destroy(self):
        if hasattr(self.app, "shortcuts_window") and self.app.shortcuts_window is self:
            self.app.shortcuts_window = None
        super().destroy()

# ===== UI: colunas =====
TYPE_VALUES = [
    "Texto",
    "Área de texto",
    "Numérico",
    "Lista",
    "Lista Vários",
    "Data",
    "Informativo",
    "Anexo",
    "Valores",
    "Componente do sistema",
    "Objeto",
]
LIST_FIELD_TYPES = {"Lista", "Lista Vários"}
MULTISELECT_FIELD_TYPES = {"Lista Vários"}
DEFAULT_COLS: List[Tuple[str, str, int]] = [
    ("move",  " ",                              60),
    ("sel",   "Sel.",                           48),
    ("campo", "Campo",                          260),
    ("tipo",  "Tipo",                           160),
    ("origem","Origem",                         240),
    ("regras","Regras (quando aparece)",        300),
    ("obrig", "Obrigatório",                    120),
    ("soleit","Só leitura",                     120),
    ("opts",  "Subtipo/Opções",                 260),
    ("obs",   "Observações",                    260),
    ("del",   "x",                               44),
]
MIN_W = {"move": 60, "sel": 44, "campo": 160, "tipo": 120, "origem": 160, "regras": 180, "obrig": 90, "soleit": 90, "opts": 180, "obs": 160, "del": 40}
HEADER_ALIGN = {"move":"center", "sel":"center","campo":"w","tipo":"w","origem":"w","regras":"w","obrig":"center","soleit":"center","opts":"w","obs":"w","del":"center"}
HEADER_PADX_LEFT = 10

# ===== Classes da UI do Importador de BPMN (Copiado do Simulador) =====
class EditItemDialog(ctk.CTkToplevel):
    def __init__(self, parent, item: Dict[str, Any], item_type: str):
        super().__init__(parent)
        self.transient(parent)
        self.grab_set()
        
        self.result = None
        self.item_type = item_type
        
        title = "Editar Tarefa" if item_type == "task" else "Editar Campo"
        self.title(title)
        
        # CORREÇÃO: A chamada para centralizar e estilizar foi movida para a classe que a invoca.
        # Isto previne o erro que deixava a janela vazia.

        main_frame = ctk.CTkFrame(self, fg_color="transparent")
        main_frame.pack(padx=20, pady=20)
        
        ctk.CTkLabel(main_frame, text="Nome:").grid(row=0, column=0, sticky="w", padx=(0, 10))
        self.name_entry = ctk.CTkEntry(main_frame, width=300)
        self.name_entry.grid(row=0, column=1)
        self.name_entry.insert(0, item.get("name", "") if item_type == "task" else item.get("campo", ""))

        if self.item_type == "field":
            ctk.CTkLabel(main_frame, text="Opções (separadas por ';'):").grid(row=1, column=0, sticky="w", pady=(10, 0))
            self.options_entry = ctk.CTkEntry(main_frame, width=300)
            self.options_entry.grid(row=1, column=1, pady=(10, 0))
            self.options_entry.insert(0, "; ".join(item.get("opcoes", [])))
            
        btn_frame = ctk.CTkFrame(self, fg_color="transparent")
        btn_frame.pack(pady=(10, 20))
        
        ctk.CTkButton(btn_frame, text="Salvar", command=self.on_save).pack(side="left", padx=5)
        btn_cancel = ctk.CTkButton(btn_frame, text="Cancelar", command=self.destroy)
        _apply_secondary_style(btn_cancel)
        btn_cancel.pack(side="left", padx=5)
        
        self.after(50, self.name_entry.focus_set)

    def on_save(self):
        new_name = self.name_entry.get().strip()
        if not new_name:
            messagebox.showwarning("Aviso", "O nome não pode ser vazio.", parent=self)
            return

        self.result = {"name": new_name}
        if self.item_type == "field":
            self.result["opcoes"] = [opt.strip() for opt in self.options_entry.get().split(";") if opt.strip()]
        
        self.destroy()

class BPMNImporterWindow(ctk.CTkToplevel):
    def __init__(self, master):
        super().__init__(master)
        self.grab_set()
        self.result = None
        self.app = master # CORREÇÃO: Guarda a referência da janela principal

        self.title("Importador de BPMN")
        try:
            self.configure(fg_color=DARK_BG2)
        except Exception:
            pass
        self.resizable(True, True)
        self.app._center_toplevel(self, 900, 620, transient=False)
        self.minsize(720, 520)
        # self.app._set_dark_title_bar(self) # A _center_toplevel já faz isso

        self.diagrams_data = {}
        self.diagram_list = []
        self.current_tasks = []
        self.current_fields_by_task = {}
        self.task_vars: Dict[str, tk.BooleanVar] = {}
        self.field_vars: Dict[str, tk.BooleanVar] = {}
        self.current_diagram_id: Optional[str] = None
        self.current_diagram_label: str = ""
        self.selected_task_ids: Set[str] = set()
        self.selected_field_ids: Set[str] = set()
        self.original_task_order: List[str] = []
        self.search_var = tk.StringVar()
        self.selection_badge: Optional[ctk.CTkLabel] = None

        self.main_frame = ctk.CTkFrame(self, fg_color="transparent")
        self.main_frame.pack(fill="both", expand=True, padx=15, pady=15)

        self.welcome_frame = ctk.CTkFrame(self.main_frame, fg_color="transparent")
        self.importer_frame = ctk.CTkFrame(self.main_frame, fg_color="transparent")
        
        self.setup_welcome_screen()
        self.setup_importer_screen()

        self.show_welcome_screen()
        self.after(250, self.run_import_flow)

    def _center_child_dialog(self, dialog: ctk.CTkToplevel, width: int, height: int) -> None:
        """Centraliza um diálogo em relação ao importador sem minimizar a janela atual."""
        try:
            dialog.transient(self)
        except Exception:
            pass

        self.update_idletasks()
        dialog.update_idletasks()

        width = max(width, dialog.winfo_reqwidth())
        height = max(height, dialog.winfo_reqheight())

        x = self.winfo_x() + max(0, (self.winfo_width() - width) // 2)
        y = self.winfo_y() + max(0, (self.winfo_height() - height) // 2)

        screen_w = self.winfo_screenwidth()
        screen_h = self.winfo_screenheight()
        x = max(0, min(x, screen_w - width))
        y = max(0, min(y, screen_h - height))

        dialog.geometry(f"{width}x{height}+{x}+{y}")

        try:
            dialog.configure(fg_color=DARK_BG2)
        except Exception:
            pass

        try:
            self.app._set_dark_title_bar(dialog)
        except Exception:
            pass

        _animate_fade_in(dialog)

    def setup_welcome_screen(self):
        ctk.CTkLabel(self.welcome_frame, text="A carregar ficheiro BPMN...", font=ctk.CTkFont(size=16, weight="bold")).pack(pady=(100, 15))
    
    def setup_importer_screen(self):
        self.importer_frame.grid_columnconfigure(0, weight=1)
        self.importer_frame.grid_rowconfigure(2, weight=1)

        header_card = ctk.CTkFrame(
            self.importer_frame,
            fg_color=THEME.color("surface"),
            corner_radius=16,
        )
        _safe_configure(
            header_card,
            border_width=1,
            border_color=THEME.color("surface_border"),
        )
        header_card.grid(row=0, column=0, sticky="ew")
        header_card.grid_columnconfigure(1, weight=1)
        header_card.grid_columnconfigure(2, weight=0)

        title_block = ctk.CTkFrame(header_card, fg_color="transparent")
        title_block.grid(row=0, column=0, columnspan=3, sticky="ew", padx=22, pady=(18, 6))
        ctk.CTkLabel(title_block, text="Importar Tarefas a partir do BPMN", font=ctk.CTkFont(size=18, weight="bold")).pack(anchor="w")
        ctk.CTkLabel(
            title_block,
            text="Selecione o diagrama, filtre itens específicos e ajuste a ordem antes de adicionar ao fluxo.",
            text_color=THEME.color("text_muted"),
        ).pack(anchor="w", pady=(2, 0))

        ctk.CTkLabel(header_card, text="Diagrama:", font=ctk.CTkFont(weight="bold")).grid(row=1, column=0, padx=(22, 12), pady=(4, 12), sticky="w")
        self.diag_combobox = ctk.CTkComboBox(
            header_card,
            width=340,
            command=lambda value: self.update_tree_preview(value, force_reload=True),
        )
        self.diag_combobox.grid(row=1, column=1, sticky="ew", pady=(4, 12))

        sort_buttons = ctk.CTkFrame(header_card, fg_color="transparent")
        sort_buttons.grid(row=1, column=2, padx=(12, 22), pady=(4, 12), sticky="e")
        btn_restore = ctk.CTkButton(
            sort_buttons,
            text="Restaurar ordem do BPMN",
            width=200,
            command=self.restore_original_order,
        )
        THEME.apply_button(btn_restore, "ghost")
        btn_restore.pack(side="top", fill="x")
        btn_alpha = ctk.CTkButton(
            sort_buttons,
            text="Ordenar A-Z",
            width=200,
            command=self.sort_tasks_alphabetically,
        )
        THEME.apply_button(btn_alpha, "ghost")
        btn_alpha.pack(side="top", fill="x", pady=(8, 0))

        ctk.CTkLabel(header_card, text="Filtro rápido:", font=ctk.CTkFont(weight="bold")).grid(row=2, column=0, padx=(22, 12), pady=(0, 22), sticky="w")
        search_entry = ctk.CTkEntry(
            header_card,
            textvariable=self.search_var,
            placeholder_text="Digite parte do nome da tarefa ou campo...",
        )
        search_entry.grid(row=2, column=1, sticky="ew", pady=(0, 22))
        self.search_var.trace_add("write", lambda *_: self.render_tree())

        actions_frame = ctk.CTkFrame(self.importer_frame, fg_color="transparent")
        actions_frame.grid(row=1, column=0, sticky="ew", pady=(18, 12))
        actions_frame.grid_columnconfigure(0, weight=1)
        self.selection_badge = ctk.CTkLabel(
            actions_frame,
            text="Nenhum diagrama carregado",
            text_color=THEME.color("text_muted"),
        )
        self.selection_badge.pack(side="left")
        ctk.CTkButton(actions_frame, text="Adicionar Selecionados ao Fluxo", command=self.confirm_selection, height=38, width=240).pack(side="right", padx=(8, 0))
        btn_cancel = ctk.CTkButton(actions_frame, text="Cancelar", command=self.destroy, height=38)
        _apply_secondary_style(btn_cancel)
        btn_cancel.pack(side="right")

        self.tree_scroll_frame = ctk.CTkScrollableFrame(
            self.importer_frame,
            fg_color=THEME.color("surface_low"),
            corner_radius=18,
        )
        self.tree_scroll_frame.grid(row=2, column=0, sticky="nsew")

    def show_welcome_screen(self):
        self.importer_frame.pack_forget()
        self.welcome_frame.pack(fill="both", expand=True)

    def show_importer_screen(self):
        self.welcome_frame.pack_forget()
        self.importer_frame.pack(fill="both", expand=True)
        
    def run_import_flow(self):
        path = filedialog.askopenfilename(title="Selecione um ficheiro Bizagi .bpm", filetypes=[("Bizagi Modeler (*.bpm)", "*.bpm")], parent=self)
        if not path:
            self.destroy()
            return

        try:
            diags, nodes, trans = parse_bizagi_group_by_diagram(path)
            if not diags:
                messagebox.showwarning("Aviso", "Nenhum diagrama válido foi encontrado no ficheiro.", parent=self)
                self.destroy()
                return

            self.diagram_list = diags
            self.diagrams_data = {
                diag_id: {"nodes": nodes.get(diag_id, {}), "transitions": trans.get(diag_id, [])} for diag_id, _ in diags
            }
            
            diag_labels = [label for _, label in self.diagram_list]
            self.diag_combobox.configure(values=diag_labels)
            self.diag_combobox.set(diag_labels[0])

            self.show_importer_screen()
            self.update_tree_preview()

        except Exception as e:
            messagebox.showerror("Erro na Importação", f"Ocorreu um erro ao processar o ficheiro:\n\n{e}", parent=self)
            self.destroy()

    def update_tree_preview(self, selected_label=None, *, force_reload: bool = False):
        if selected_label is None:
            selected_label = self.diag_combobox.get()

        diag_id = next((d_id for d_id, lbl in self.diagram_list if lbl == selected_label), None)
        if not diag_id:
            self.current_diagram_id = None
            self.current_diagram_label = selected_label or ""
            self.current_tasks = []
            self.current_fields_by_task = {}
            self.original_task_order = []
            self.selected_task_ids.clear()
            self.selected_field_ids.clear()
            self.render_tree()
            return

        needs_reload = force_reload or diag_id != self.current_diagram_id
        self.current_diagram_id = diag_id
        self.current_diagram_label = selected_label or ""

        if needs_reload:
            data = self.diagrams_data[diag_id]
            tasks, fields_by_task = build_task_fields_for_diagram(data["nodes"], data["transitions"])
            self.current_tasks = tasks
            self.current_fields_by_task = fields_by_task
            self.original_task_order = [t["id"] for t in tasks]
            self.selected_task_ids = {t["id"] for t in tasks}
            self.selected_field_ids = {
                field["id"]
                for task in tasks
                for field in self.current_fields_by_task.get(task["id"], [])
            }

        self.render_tree()

    def handle_row_click(self, event, item, item_type, checkbox):
        if not (checkbox.winfo_x() <= event.x < checkbox.winfo_x() + checkbox.winfo_width()):
             self.open_edit_dialog(item, item_type, checkbox)

    def _prune_selection(self) -> None:
        valid_task_ids = {t["id"] for t in self.current_tasks}
        valid_field_ids = {
            field["id"]
            for task in self.current_tasks
            for field in self.current_fields_by_task.get(task["id"], [])
        }
        self.selected_task_ids &= valid_task_ids
        self.selected_field_ids &= valid_field_ids

    def _update_selection_badge(self, visible_count: int) -> None:
        if not self.selection_badge:
            return
        total_tasks = len(self.current_tasks)
        total_fields = sum(len(self.current_fields_by_task.get(t["id"], [])) for t in self.current_tasks)
        selected_tasks = len(self.selected_task_ids & {t["id"] for t in self.current_tasks})
        selected_fields = len(
            self.selected_field_ids
            & {
                field["id"]
                for task in self.current_tasks
                for field in self.current_fields_by_task.get(task["id"], [])
            }
        )
        text = "Nenhuma tarefa disponível"
        if total_tasks:
            if total_fields:
                fields_part = f"{selected_fields}/{total_fields} campos"
            else:
                fields_part = "Nenhum campo de decisão"
            text = (
                f"{selected_tasks}/{total_tasks} tarefas selecionadas — "
                f"{fields_part} — "
                f"{visible_count} em exibição"
            )
        self.selection_badge.configure(text=text)

    def _filtered_tasks(self) -> List[Dict]:
        if not self.current_tasks:
            return []
        self._prune_selection()
        query = self.search_var.get().strip().lower()
        visible_tasks: List[Dict] = []
        for task in self.current_tasks:
            fields = self.current_fields_by_task.get(task["id"], [])
            if not query:
                visible_tasks.append(task)
                continue
            search_pool = [task.get("name", "")] + [f.get("campo", "") for f in fields]
            for f in fields:
                search_pool.extend(f.get("opcoes", []))
            if any(query in (item or "").lower() for item in search_pool):
                visible_tasks.append(task)
        return visible_tasks

    def render_tree(self) -> None:
        for widget in self.tree_scroll_frame.winfo_children():
            widget.destroy()
        self.task_vars.clear()
        self.field_vars.clear()

        if not self.current_tasks:
            ctk.CTkLabel(
                self.tree_scroll_frame,
                text="Carregue um diagrama para visualizar tarefas.",
                text_color=THEME.color("text_muted"),
            ).pack(pady=40)
            self._update_selection_badge(0)
            return

        visible_tasks = self._filtered_tasks()

        if not visible_tasks:
            ctk.CTkLabel(
                self.tree_scroll_frame,
                text="Nenhum resultado encontrado para o filtro informado.",
                text_color=THEME.color("text_muted"),
            ).pack(pady=40)
            self._update_selection_badge(0)
            return

        for idx, task in enumerate(visible_tasks, start=1):
            self.render_task_item(task, idx)

        self._update_selection_badge(len(visible_tasks))

    def render_task_item(self, task: Dict, position: int):
        task_id = task['id']
        task_var = tk.BooleanVar(value=task_id in self.selected_task_ids)
        self.task_vars[task_id] = task_var

        task_card = ctk.CTkFrame(
            self.tree_scroll_frame,
            fg_color=THEME.color("surface"),
            corner_radius=16,
        )
        _safe_configure(
            task_card,
            border_width=1,
            border_color=THEME.color("surface_border"),
        )
        task_card.pack(fill="x", pady=(0, 16), padx=18)

        task_header = ctk.CTkFrame(task_card, fg_color="transparent", cursor="hand2")
        task_header.pack(fill="x", pady=(14, 8), padx=18)

        badge = ctk.CTkLabel(
            task_header,
            text=f"{position:02d}",
            width=44,
            fg_color=THEME.color("badge_bg"),
            corner_radius=8,
            font=ctk.CTkFont(weight="bold"),
            anchor="center",
            text_color=THEME.color("badge_text"),
        )
        badge.pack(side="left", padx=(0, 12))

        chk_task = ctk.CTkCheckBox(
            task_header,
            text=task['name'],
            variable=task_var,
            font=ctk.CTkFont(size=15, weight="bold"),
            command=lambda t_id=task_id: self.on_task_toggle(t_id),
        )
        chk_task.pack(side="left", padx=(0, 12))

        task_fields = self.current_fields_by_task.get(task_id, [])
        ctk.CTkLabel(
            task_header,
            text=f"{len(task_fields)} campo(s)",
            text_color=THEME.color("text_muted"),
        ).pack(side="left")

        controls = ctk.CTkFrame(task_header, fg_color="transparent")
        controls.pack(side="right")
        btn_up = ctk.CTkButton(
            controls,
            text="▲",
            width=34,
            height=34,
            command=lambda tid=task_id: self.reorder_task(tid, -1),
        )
        THEME.apply_button(btn_up, "ghost")
        btn_up.pack(side="left", padx=(0, 6))
        btn_down = ctk.CTkButton(
            controls,
            text="▼",
            width=34,
            height=34,
            command=lambda tid=task_id: self.reorder_task(tid, 1),
        )
        THEME.apply_button(btn_down, "ghost")
        btn_down.pack(side="left")

        for widget in (task_card, task_header, badge):
            widget.bind("<Enter>", lambda _e, w=task_card: w.configure(fg_color=THEME.color("surface_hover")))
            widget.bind("<Leave>", lambda _e, w=task_card: w.configure(fg_color=THEME.color("surface")))

        task_header.bind("<Button-1>", lambda e, t=task, c=chk_task: self.handle_row_click(e, t, "task", c))

        if task_fields:
            fields_frame = ctk.CTkFrame(task_card, fg_color=THEME.color("surface_low"), corner_radius=12)
            fields_frame.pack(fill="x", padx=16, pady=(0, 14))
            for field in task_fields:
                self.render_field_item(field, task_id, fields_frame)
        else:
            ctk.CTkLabel(
                task_card,
                text="Nenhum campo de decisão encontrado neste passo.",
                text_color=THEME.color("text_muted"),
            ).pack(fill="x", padx=18, pady=(0, 16))

    def reorder_task(self, task_id: str, delta: int) -> None:
        index = next((i for i, t in enumerate(self.current_tasks) if t['id'] == task_id), -1)
        if index == -1:
            return
        new_index = index + delta
        if not (0 <= new_index < len(self.current_tasks)):
            return
        self.current_tasks.insert(new_index, self.current_tasks.pop(index))
        self.render_tree()

    def sort_tasks_alphabetically(self) -> None:
        self.current_tasks.sort(key=lambda t: (t.get('name', '') or '').lower())
        self.render_tree()

    def restore_original_order(self) -> None:
        if not self.original_task_order:
            return
        order_map = {tid: idx for idx, tid in enumerate(self.original_task_order)}
        self.current_tasks.sort(key=lambda t: order_map.get(t['id'], len(order_map)))
        self.render_tree()

    def render_field_item(self, field: Dict, task_id: str, parent_card: ctk.CTkFrame):
        field_id = field['id']
        field_var = tk.BooleanVar(value=field_id in self.selected_field_ids)
        self.field_vars[field_id] = field_var

        wrapper = ctk.CTkFrame(parent_card, fg_color="transparent")
        wrapper.pack(fill="x", padx=12, pady=6)

        field_row = ctk.CTkFrame(
            wrapper,
            fg_color=THEME.color("field_surface"),
            corner_radius=10,
            cursor="hand2",
        )
        _safe_configure(
            field_row,
            border_width=1,
            border_color=THEME.color("field_border"),
        )
        field_row.pack(fill="x")

        content_frame = ctk.CTkFrame(field_row, fg_color="transparent")
        content_frame.pack(fill="x", padx=14, pady=10)

        chk_field = ctk.CTkCheckBox(
            content_frame,
            text="",
            variable=field_var,
            command=lambda fid=field_id, tid=task_id: self.on_field_toggle(tid, fid),
        )
        chk_field.pack(side="left")

        label_frame = ctk.CTkFrame(content_frame, fg_color="transparent")
        label_frame.pack(side="left", fill="both", expand=True, padx=(10, 0))

        ctk.CTkLabel(label_frame, text=field['campo'], font=ctk.CTkFont(weight="bold")).pack(anchor="w")
        ctk.CTkLabel(
            content_frame,
            text=field['tipo'],
            fg_color=THEME.color("chip_bg"),
            text_color=THEME.color("chip_text"),
            corner_radius=8,
            padx=12,
            pady=4,
        ).pack(side="right")

        if field['opcoes']:
            ctk.CTkLabel(
                label_frame,
                text=f"Opções: {', '.join(field['opcoes'])}",
                text_color=THEME.color("text_muted"),
                wraplength=420,
                justify="left",
            ).pack(anchor="w", pady=(4, 0))

        def _open_editor(event):
            self.handle_row_click(event, field, "field", chk_field)

        for widget in (field_row, content_frame, label_frame):
            widget.bind("<Enter>", lambda _e, w=field_row: w.configure(fg_color=THEME.color("field_hover")))
            widget.bind("<Leave>", lambda _e, w=field_row: w.configure(fg_color=THEME.color("field_surface")))
            widget.bind("<Button-1>", _open_editor)


    def on_task_toggle(self, task_id: str):
        is_selected = self.task_vars[task_id].get()
        if is_selected:
            self.selected_task_ids.add(task_id)
        else:
            self.selected_task_ids.discard(task_id)

        self._update_selection_badge(len(self._filtered_tasks()))

    def on_field_toggle(self, task_id: str, field_id: str) -> None:
        selected = self.field_vars[field_id].get()
        if selected:
            self.selected_field_ids.add(field_id)
        else:
            self.selected_field_ids.discard(field_id)

        self._update_selection_badge(len(self._filtered_tasks()))

    def open_edit_dialog(self, item: Dict, item_type: str, checkbox_widget: ctk.CTkCheckBox):
        dialog = EditItemDialog(self, item, item_type)

        # CORREÇÃO: centraliza usando o importador como base para evitar que a janela
        # principal "pisque" ou seja minimizada durante a edição.
        self._center_child_dialog(dialog, 450, 220 if item_type == "field" else 160)
        
        self.wait_window(dialog)
        
        if dialog.result:
            if item_type == "task":
                item['name'] = dialog.result['name']
            elif item_type == "field":
                item['campo'] = dialog.result['name']
                item['opcoes'] = dialog.result['opcoes']

            # Atualiza a interface gráfica para mostrar as alterações
            self.render_tree()

    def confirm_selection(self):
        selected_tasks = [task for task in self.current_tasks if task['id'] in self.selected_task_ids]

        if not selected_tasks:
            messagebox.showwarning("Importador de BPMN", "Selecione ao menos uma tarefa antes de importar.", parent=self)
            return

        selected_fields_by_task = {}
        for task in selected_tasks:
            task_id = task['id']
            selected_fields_by_task[task_id] = [
                field
                for field in self.current_fields_by_task.get(task_id, [])
                if field['id'] in self.selected_field_ids
            ]

        self.result = {
            "tasks": selected_tasks,
            "fields_by_task": selected_fields_by_task,
            "diagram_id": self.current_diagram_id,
            "diagram_label": self.current_diagram_label,
        }

        self.destroy()


# ===== App principal =====
class App(ctk.CTk):
    def __init__(self):
        super().__init__()
        self._resize_timer = None
        global _APP_ROOT
        _APP_ROOT = self
        self.title(f"Designer de Campos — Fluxos [{APP_VERSION}]")
        self.geometry("1360x800"); self.minsize(1180, 680)

        try:
            self.tk_scale = float(self.tk.call("tk", "scaling"))
        except Exception:
            self.tk_scale = 1.0

        self.row_h = max(36, int(round(36 * self.tk_scale)))
        self.header_h = self.row_h

        self.cols = self._load_cols_config()
        self.col_gap = self._load_col_gap_config(default=max(5, int(round(5 * self.tk_scale))))

        self.project = ProjectModel()
        self.current_task_id: Optional[str] = None
        self.store = TemplateStore()

        self.selected_field_ids: Set[str] = set()
        self.overview_window: Optional[ctk.CTkToplevel] = None
        self.sim_window: Optional[ctk.CTkToplevel] = None
        self.context_menu: Optional[CustomContextMenu] = None
        self.shortcuts_window: Optional[ShortcutOverlay] = None
        self.validator_ignored: Set[str] = set()

        # --- Cache de Metadados ---
        self._field_id_to_name: Dict[str, str] = {}
        self._task_id_to_name: Dict[str, str] = {}
        self._field_id_to_task_id: Dict[str, str] = {}
        
        # estado da visão HTML (fases colapsadas)
        self._html_overview_collapsed: Set[str] = set()

        # Undo/Redo
        self._undo_stack: List[dict] = []
        self._redo_stack: List[dict] = []
        self._clipboard: Dict = {} # Alterado para Dict para armazenar a tarefa de origem
        self._UNDO_MAX = 50

        # registradores UI da grade principal
        self._resizers: List[tk.Frame] = []
        self._resizer_state: Optional[Dict[str, int]] = None
        self._resizer_guide: Optional[tk.Frame] = None
        self._rows: List[ctk.CTkFrame] = []
        self._row_cells: Dict[int, Dict[str, ctk.CTkFrame]] = {}
        self._header_cells: Dict[str, ctk.CTkFrame] = {}
        self._field_row_map: Dict[str, ctk.CTkFrame] = {}

        ctk.set_appearance_mode("dark")
        ctk.set_default_color_theme("blue")

        self._build_menubar()
        self._build_top_controls()
        self._build_table_area()
        self._bind_shortcuts()

        # --- CORREÇÃO DE "GHOSTING" APÓS BLOQUEIO/SUSPENSÃO ---
        # O evento <Map> dispara quando a janela é exibida/restaurada na tela.
        # O evento <FocusIn> dispara quando você clica na janela.
        def _wake_up_window(event=None):
            # Garante que a janela não esteja minimizada antes de forçar atributos
            if self.state() == 'normal':
                # Força o Windows a reconhecer a opacidade total
                if self.attributes("-alpha") < 1.0:
                    self.attributes("-alpha", 1.0)

                # Opcional: Força um update visual leve para "acordar" o renderizador
                # self.update_idletasks()

        # Vincula aos eventos de "acordar" e "focar"
        self.bind("<Map>", _wake_up_window)
        self.bind("<FocusIn>", _wake_up_window)
        # -------------------------------------------------------

        self.new_flow_blank(show_message=False)
        self._rebuild_metadata_cache() # Inicializa o cache

    # --- Cache de Metadados ---
    def _rebuild_metadata_cache(self):
        """Constrói caches para buscas O(1) de nomes de tarefas e campos."""
        self._field_id_to_name.clear()
        self._task_id_to_name.clear()
        self._field_id_to_task_id.clear()

        for task in self.project.tasks:
            self._task_id_to_name[task.id] = task.name
            for field in task.fields:
                self._field_id_to_name[field.id] = field.name
                self._field_id_to_task_id[field.id] = task.id

    # ===== Menus =====
    def _build_menubar(self):
        bar = ctk.CTkFrame(self, fg_color=DARK_BG, height=34, corner_radius=0)
        bar.pack(side="top", fill="x"); bar.pack_propagate(False)

        # --- Frame para menus à esquerda ---
        left_menus = ctk.CTkFrame(bar, fg_color="transparent")
        left_menus.pack(side="left")

        def menu(parent, title, items: Dict[str, callable]):
            def on_select(choice: str, w: ctk.CTkOptionMenu):
                try:
                    cb = items.get(choice)
                    if callable(cb): cb()
                finally:
                    w.set(title)
            om = ctk.CTkOptionMenu(parent, values=list(items.keys()), command=lambda *_: None)
            om.configure(command=lambda choice, w=om: on_select(choice, w))
            om.set(title)
            base = _solid_color(); om.configure(fg_color=base, button_color=base, button_hover_color=base)
            om.pack(side="left", padx=(10, 6), pady=4)
            return om

        menu(left_menus, "Arquivo", {
            "Novo vazio (Ctrl+Shift+N)": self.new_flow_blank,
            "Abrir... (Ctrl+O)": self.open_project,
            "Salvar (Ctrl+S)": self.save_project,
            "Exportar fluxo (XLSX) (Ctrl+E)": self.export_flow_to_xlsx,
            "Sair": self.destroy,
        })
        menu(left_menus, "Templates", {
            "Aplicar › Cadastro": self.apply_builtin_template_cadastro_confirm,
            "Meus templates... (Ctrl+T)": self.open_templates_dialog,
            "Salvar fluxo como template... (Ctrl+Shift-T)": self.save_flow_as_template,
            "Importar fluxo do Excel (.xlsx)...": self.import_flow_from_xlsx,
            "Importar Tarefas de BPMN...": self.open_bpmn_importer, # PONTO DE ENTRADA ADICIONADO
            "Baixar modelo de fluxo (.xlsx)": self.download_flow_template_xlsx,
        })
        menu(left_menus, "Objetos", {
            "Tipo do objeto do fluxo...": self.open_flow_object_type_dialog,
            "Esquema do objeto...": self.open_object_schema_editor,
            "Importar esquema (XLSX)...": self.import_object_schema_xlsx,
            "Exportar esquema (XLSX)...": self.export_object_schema_xlsx,
            "Baixar modelo de esquema (.xlsx)": self.download_object_schema_template_xlsx,
        })
        menu(left_menus, "Layout", {
            "Largura das colunas...": self.open_columns_dialog,
            "Espaçamento entre colunas...": self.open_col_gap_dialog,
            "Restaurar larguras padrão": self.restore_default_columns,
        })
        menu(left_menus, "Ajuda", {
            "Atalhos do teclado (F1)": lambda: self.open_shortcuts_overlay(),
        })
        
        # --- Frame para botões de ação à direita ---
        right_actions = ctk.CTkFrame(bar, fg_color="transparent")
        right_actions.pack(side="right", padx=10, pady=4)

        # ===== Botões com largura fixa pequena para estética =====
        self.btn_copy = ctk.CTkButton(right_actions, text="Copiar Sel.", command=self._copy_selected_fields, width=110)
        self.btn_copy.pack(side="left", padx=(0, 4))
        
        self.btn_paste = ctk.CTkButton(right_actions, text="Colar", command=self._paste_fields, state="disabled", width=70)
        self.btn_paste.pack(side="left", padx=(0, 4))
        
        self.btn_delete_selected = ctk.CTkButton(right_actions, text="Excluir Sel.", command=self.delete_selected_fields, width=110)
        self.btn_delete_selected.pack(side="left")

    # ===== Top controls =====
    def _build_top_controls(self):
        top = ctk.CTkFrame(self, fg_color="transparent"); top.pack(side="top", fill="x", padx=10, pady=(6, 6))

        # --- Controles da Esquerda ---
        left = ctk.CTkFrame(top, fg_color="transparent"); left.pack(side="left")
        self.lbl_flow = ctk.CTkLabel(left, text="Fluxo: -"); self.lbl_flow.pack(side="left", padx=(0, 8))
        ctk.CTkButton(left, text="Renomear...", command=self.rename_flow, width=110).pack(side="left", padx=(0, 6))

        ctk.CTkLabel(left, text="Tarefa:").pack(side="left", padx=(6, 6))
        self.cmb_task = ctk.CTkComboBox(left, values=[], width=380, command=lambda *_: self._on_task_change()); self.cmb_task.pack(side="left", padx=(0, 6))
        ctk.CTkButton(left, text="Tarefas...", command=self.open_tasks_dialog, width=90).pack(side="left", padx=(0, 6))
        ctk.CTkButton(left, text="+ Campo", command=self._add_field, width=90).pack(side="left", padx=(0, 6))
        
        # --- Controles da Direita (Layout Revertido) ---
        right = ctk.CTkFrame(top, fg_color="transparent"); right.pack(side="right")
        ctk.CTkButton(right, text="Validar Fluxo", command=self.open_flow_validator, width=120).pack(side="left", padx=(0, 4))
        ctk.CTkButton(right, text="Planilha", command=self.open_overview_html, width=90).pack(side="left", padx=(0, 4))
        ctk.CTkButton(right, text="Simular", command=self.open_simulator, width=90).pack(side="left")

    # ===== Lógica de Ponte do Importador de BPMN =====
    def open_bpmn_importer(self):
        importer_window = BPMNImporterWindow(self)
        self.wait_window(importer_window)

        if not importer_window.result:
            return # Utilizador cancelou

        # --- Iniciar transação segura ---
        self._push_undo() # CRÍTICO: Garante a reversibilidade

        imported_data = importer_window.result
        newly_added_tasks = []

        selected_diag_label = (imported_data.get("diagram_label") or "").strip()
        if selected_diag_label:
            self.project.flow_name = selected_diag_label

        for task_data in imported_data["tasks"]:
            # Cria a nova tarefa
            new_task = Task(id=_uid(), name=task_data["name"], fields=[])

            # Cria os campos para esta tarefa
            fields_for_this_task = imported_data["fields_by_task"].get(task_data["id"], [])
            for field_data in fields_for_this_task:
                new_field = Field(
                    id=_uid(),
                    name=field_data["campo"],
                    ftype="Lista", # BPMN só importa campos de decisão (Lista)
                    options="; ".join(field_data["opcoes"])
                )
                new_task.fields.append(new_field)

            newly_added_tasks.append(new_task)

        # Adiciona as novas tarefas ao projeto (operação aditiva)
        self.project.tasks.extend(newly_added_tasks)

        # Atualiza a interface de forma segura
        self._rebuild_metadata_cache()
        self._refresh_task_combo()
        self._refresh_rows()
        self._refresh_flow_label()

        messagebox.showinfo("Sucesso", f"{len(newly_added_tasks)} tarefa(s) importada(s) com sucesso.")

    # ===== Tabela de edição (grade principal) =====
    def _build_table_area(self):
        area = ctk.CTkFrame(self, fg_color="transparent")
        area.pack(fill="both", expand=True, padx=10, pady=(0, 10))

        # Header
        self.header_canvas = tk.Canvas(area, height=self.header_h, highlightthickness=0, bd=0, background=DARK_BG3)
        self.header_canvas.pack(side="top", fill="x")
        self.header_frame = ctk.CTkFrame(self.header_canvas, fg_color=DARK_BG3, height=self.header_h, corner_radius=0)
        self.header_window = self.header_canvas.create_window((0, 0), window=self.header_frame, anchor="nw")
        self.header_separator = ctk.CTkFrame(area, fg_color=DARK_BG3, height=1, corner_radius=0)
        self.header_separator.pack(side="top", fill="x")

        # Body + scroll
        body_container = ctk.CTkFrame(area, fg_color="transparent"); body_container.pack(fill="both", expand=True)
        self.body_canvas = tk.Canvas(body_container, highlightthickness=0, bd=0, background=DARK_BG2)
        self.body_canvas.pack(side="left", fill="both", expand=True)
        self.vscroll = ctk.CTkScrollbar(body_container, orientation="vertical", command=self.body_canvas.yview)
        self.vscroll.pack(side="right", fill="y", padx=(6, 0))
        self.body_canvas.configure(yscrollcommand=self.vscroll.set)

        self.rows_frame = ctk.CTkFrame(self.body_canvas, fg_color=DARK_BG2)
        self.body_window = self.body_canvas.create_window((0, 0), window=self.rows_frame, anchor="nw")

        # Scroll H
        self.xscroll = ctk.CTkScrollbar(area, orientation="horizontal", command=lambda *a: (self.header_canvas.xview(*a), self.body_canvas.xview(*a)))
        self.xscroll.pack(side="bottom", fill="x", pady=(6, 0))
        self.header_canvas.configure(xscrollcommand=lambda a,b: self.xscroll.set(a,b))
        self.body_canvas.configure(xscrollcommand=lambda a,b: self.xscroll.set(a,b))

        self.rows_frame.bind("<Configure>", lambda e: self._resize_rows())
        self.body_canvas.bind("<Configure>", lambda e: self._on_body_viewport_resize())

        self._bind_mousewheel(self.body_canvas)
        self._build_header(initial=True)

    # ===== Shortcuts =====
    def open_shortcuts_overlay(self, event=None):
        if self.shortcuts_window and self.shortcuts_window.winfo_exists():
            try:
                self.shortcuts_window.focus_set()
            except Exception:
                pass
            return "break"

        self.shortcuts_window = ShortcutOverlay(self, SHORTCUT_SECTIONS)
        return "break"

    def _bind_shortcuts(self):
        self.bind_all("<Control-s>", lambda e: self.save_project())
        self.bind_all("<Control-e>", lambda e: self.export_flow_to_xlsx())
        self.bind_all("<Control-o>", lambda e: self.open_project())
        self.bind_all("<Control-Shift-N>", lambda e: self.new_flow_blank())
        self.bind_all("<Control-t>", lambda e: self.open_templates_dialog())
        self.bind_all("<Control-Shift-T>", lambda e: self.save_flow_as_template())
        # Undo / Redo
        self.bind_all("<Control-z>", lambda e: self.undo_action())
        self.bind_all("<Control-y>", lambda e: self.redo_action())
        self.bind_all("<Control-Shift-Z>", lambda e: self.redo_action())
        self.bind_all("<F1>", self.open_shortcuts_overlay)

    # ===== Undo/Redo =====
    def _serialize_project(self) -> dict:
        current_task = self._get_task()
        current_task_id = current_task.id if current_task else None
        return {
            "project": self.project.to_dict(),
            "current_task_id": current_task_id,
            "selected_field_ids": list(self.selected_field_ids),
        }

    def _apply_project_dict(self, d: dict) -> bool:
        if not isinstance(d, dict):
            messagebox.showerror("Abrir projeto", "O arquivo selecionado não contém um projeto válido.")
            return False

        state_current_task = None
        project_dict = d
        state_selection: Set[str] = set()

        if "project" in d and isinstance(d.get("project"), dict):
            project_dict = d.get("project", {})
            state_current_task = d.get("current_task_id") or d.get("_current_task_id")
            selection_data = d.get("selected_field_ids")
            if isinstance(selection_data, (list, tuple, set)):
                state_selection = {str(fid) for fid in selection_data}

        try:
            new_project = ProjectModel.from_dict(project_dict)
        except Exception as exc:
            messagebox.showerror("Abrir projeto", f"Não foi possível carregar o projeto.\n\n{exc}")
            return False

        self.project = new_project

        valid_ids = {t.id for t in self.project.tasks}
        valid_field_ids = {f.id for t in self.project.tasks for f in t.fields}

        desired_task_id = None
        if state_current_task and state_current_task in valid_ids:
            desired_task_id = state_current_task
        elif self.current_task_id and self.current_task_id in valid_ids:
            desired_task_id = self.current_task_id

        self.current_task_id = desired_task_id or (self.project.tasks[0].id if self.project.tasks else None)
        self.selected_field_ids = {fid for fid in state_selection if fid in valid_field_ids}
        # ajustar campos Objeto
        for t in self.project.tasks:
            for f in t.fields:
                if f.ftype == "Objeto":
                    if not self.project.object_type:
                        self.project.object_type = f.obj_type or f.name
                    f.obj_type = self.project.object_type; f.name = self.project.object_type
                    f.name_lock_reason = "objeto"; f.name_locked = True
        
        self._rebuild_metadata_cache() # Atualiza o cache após carregar

        self._refresh_flow_label(); self._build_header(initial=True); self._refresh_task_combo(); self._refresh_rows()
        if self.sim_window and self.sim_window.winfo_exists():
            try: self.sim_window.on_model_changed()
            except Exception: pass

        return True

    def _push_undo(self):
        try:
            self._undo_stack.append(self._serialize_project())
            if len(self._undo_stack) > self._UNDO_MAX:
                self._undo_stack.pop(0)
            self._redo_stack.clear()
        except Exception:
            pass

    def undo_action(self):
        if not self._undo_stack:
            return
        try:
            self._redo_stack.append(self._serialize_project())
            state = self._undo_stack.pop()
            self._apply_project_dict(state)
        except Exception as e:
            messagebox.showerror("Undo", f"Falha ao desfazer.\n\n{e}")

    def redo_action(self):
        if not self._redo_stack:
            return
        try:
            self._undo_stack.append(self._serialize_project())
            state = self._redo_stack.pop()
            self._apply_project_dict(state)
        except Exception as e:
            messagebox.showerror("Redo", f"Falha ao refazer.\n\n{e}")
            
    # ===== Persistência de Edição =====
    def _commit_active_edits(self):
        """Salva explicitamente o conteúdo do widget focado no modelo de dados."""
        try:
            focused_widget = self.focus_get()
            allowed_widgets: Tuple[type, ...] = (ctk.CTkEntry,)
            if hasattr(ctk, "CTkTextbox"):
                allowed_widgets += (ctk.CTkTextbox,)
            allowed_widgets += (tk.Entry, tk.Text)

            if focused_widget is None or not isinstance(focused_widget, allowed_widgets):
                return

            parent = focused_widget.master
            while parent is not None and not hasattr(parent, '_grid_info'):
                if parent in self._rows:
                    break
                parent = parent.master

            if parent in self._rows:
                row_idx = self._rows.index(parent)
                task = self._get_task()
                if task and row_idx < len(task.fields):
                    field = task.fields[row_idx]

                    undo_triggered = False

                    def ensure_once():
                        nonlocal undo_triggered
                        if not undo_triggered:
                            self._push_undo()
                            undo_triggered = True

                    self._commit_row_data(field, ensure_undo=ensure_once)
        except Exception:
            pass

    def _commit_row_data(self, field_obj: Field, *, ensure_undo: Optional[Callable[[], None]] = None) -> bool:
        """Salva os dados dos widgets de entrada de uma linha diretamente no objeto Field."""
        changed = False
        need_rebuild = False
        try:
            task = self._get_task()
            if not task:
                return False

            row_idx = -1
            for i, f in enumerate(task.fields):
                if f.id == field_obj.id:
                    row_idx = i
                    break

            if row_idx == -1 or row_idx not in self._row_cells:
                return False

            row_widgets = self._row_cells[row_idx]

            if 'campo' in row_widgets and not field_obj.name_locked:
                children = row_widgets['campo'].winfo_children()
                if children:
                    widget = children[0]
                    if isinstance(widget, (ctk.CTkEntry, tk.Entry)):
                        new_name = widget.get()
                        if new_name != field_obj.name:
                            if ensure_undo:
                                ensure_undo()
                            field_obj.name = new_name
                            need_rebuild = True
                            changed = True
                            for task_iter in self.project.tasks:
                                for fld in task_iter.fields:
                                    if fld.origin_field == field_obj.id and fld.name_lock_reason == "origem":
                                        fld.name = new_name

            if 'opts' in row_widgets:
                children = row_widgets['opts'].winfo_children()
                if children:
                    widget = children[0]
                    if isinstance(widget, (ctk.CTkEntry, tk.Entry)):
                        current_opts = field_obj.options or ""
                        new_opts = widget.get()
                        if new_opts != current_opts:
                            if ensure_undo:
                                ensure_undo()
                            field_obj.options = new_opts
                            changed = True

            if 'obs' in row_widgets:
                children = row_widgets['obs'].winfo_children()
                if children:
                    widget = children[0]
                    if isinstance(widget, (ctk.CTkEntry, tk.Entry)):
                        current_note = field_obj.note or ""
                        new_note = widget.get()
                        if new_note != current_note:
                            if ensure_undo:
                                ensure_undo()
                            field_obj.note = new_note
                            changed = True

            if need_rebuild:
                self._rebuild_metadata_cache()
        except Exception:
            return changed

        return changed

    def _commit_all_visible_row_edits(self):
        """Garante que todos os campos visíveis tenham suas edições pendentes persistidas."""
        try:
            task = self._get_task()
            if not task:
                return

            undo_triggered = False

            def ensure_once():
                nonlocal undo_triggered
                if not undo_triggered:
                    self._push_undo()
                    undo_triggered = True

            for field in task.fields:
                self._commit_row_data(field, ensure_undo=ensure_once)
        except Exception:
            pass

    # ===== Persistência de layout =====
    def _load_json(self) -> dict:
        try:
            if os.path.exists(CONFIG_PATH):
                with open(CONFIG_PATH, "r", encoding="utf-8") as f:
                    return json.load(f)
        except Exception:
            pass
        return {}

    def _save_json(self, data: dict):
        try:
            with open(CONFIG_PATH, "w", encoding="utf-8") as f:
                json.dump(data, f, ensure_ascii=False, indent=2)
        except Exception as e:
            messagebox.showerror("Salvar configuração", f"Falha ao salvar config.\n\n{e}")

    def _load_cols_config(self) -> List[Tuple[str, str, int]]:
        data = self._load_json(); widths: Dict[str, int] = data.get("col_widths", {}) if isinstance(data, dict) else {}
        out = []
        for (k, lbl, w) in DEFAULT_COLS:
            try: v = int(widths.get(k, w))
            except Exception: v = w
            out.append((k, lbl, v))
        return out

    def _save_cols_config(self):
        data = self._load_json()
        widths = {k: int(w) for (k, _, w) in self.cols}
        data["col_widths"] = widths; data["col_gap"] = int(self.col_gap)
        self._save_json(data)

    def _load_col_gap_config(self, default: int) -> int:
        data = self._load_json()
        try: return int(data.get("col_gap", default))
        except Exception: return default

    # ===== Geometria/scroll =====
    def _col_positions(self, cols_override: Optional[List[Tuple[str, str, int]]] = None) -> List[Tuple[str, int, int]]:
        cols = cols_override or self.cols
        x = 0; gap = int(self.col_gap); out = []
        for key, _, w in cols:
            w = max(MIN_W.get(key, 60), int(w))
            out.append((key, x, w))
            x += w + gap
        return out

    def _total_table_width(self, cols_override: Optional[List[Tuple[str, str, int]]] = None) -> int:
        cols = cols_override or self.cols
        gap = int(self.col_gap)
        return sum(max(MIN_W.get(k, 60), int(w)) for k, _, w in cols) + max(0, len(cols)-1)*gap

    def _on_body_viewport_resize(self, event=None):
        # DEBOUNCE: Aguarda 20ms para evitar processamento excessivo durante arraste
        if self._resize_timer:
            self.after_cancel(self._resize_timer)
        self._resize_timer = self.after(20, self._perform_resize_layout)

    def _perform_resize_layout(self):
        self._resize_rows()
        self._resize_header()
        # rebuild_resizers=False evita recriar widgets desnecessariamente durante resize
        self._apply_positions(self._col_positions(), rebuild_resizers=False)
        self._resize_timer = None

    def _resize_header(self):
        total_w = self._total_table_width()
        viewport_w = max(1, self.body_canvas.winfo_width())
        inner_w = max(total_w, viewport_w)

        # GUARDA CONDICIONAL: Só aplica configure se o scrollregion mudou
        new_region = (0, 0, total_w, self.header_h)
        # Converte para string para comparação segura com o retorno do Tcl/Tk
        current_region_str = str(self.header_canvas.cget("scrollregion")).replace('"', '')
        new_region_str = f"0 0 {total_w} {self.header_h}"

        if current_region_str != new_region_str:
            self.header_canvas.configure(scrollregion=new_region)

        # GUARDA CONDICIONAL: Só redimensiona a janela interna se necessário
        current_inner_w = float(self.header_canvas.itemcget(self.header_window, "width"))
        if abs(current_inner_w - inner_w) > 1: # Tolerância de 1px
            self.header_canvas.itemconfig(self.header_window, width=inner_w, height=self.header_h)

        # Sincroniza largura da viewport
        if int(self.header_canvas.cget("width")) != int(viewport_w):
            self.header_canvas.configure(width=viewport_w)

    def _resize_rows(self):
        total_w = self._total_table_width()
        viewport_w = max(1, self.body_canvas.winfo_width())
        inner_w = max(total_w, viewport_w)

        bbox = self.body_canvas.bbox(self.body_window)
        height = max(bbox[3] if bbox else 0, self.body_canvas.winfo_height())

        # GUARDA CONDICIONAL: Impede loop de eventos no scrollregion
        new_region = (0, 0, total_w, height)
        current_region_str = str(self.body_canvas.cget("scrollregion")).replace('"', '')
        new_region_str = f"0 0 {total_w} {height}"

        if current_region_str != new_region_str:
            self.body_canvas.configure(scrollregion=new_region)

        # GUARDA CONDICIONAL: Impede loop de eventos na largura do item
        current_inner_w = float(self.body_canvas.itemcget(self.body_window, "width"))
        if abs(current_inner_w - inner_w) > 1:
            self.body_canvas.itemconfig(self.body_window, width=inner_w)

    def _bind_mousewheel(self, widget: tk.Widget):
        widget.bind_all("<MouseWheel>", lambda e: self._on_mousewheel(e), add="+")
        widget.bind_all("<Shift-MouseWheel>", lambda e: self._on_hwheel(e), add="+")
        # Linux
        widget.bind_all("<Button-4>", lambda e: self._on_mousewheel(e), add="+")
        widget.bind_all("<Button-5>", lambda e: self._on_mousewheel(e), add="+")
        widget.bind_all("<Shift-Button-4>", lambda e: self._on_hwheel(e), add="+")
        widget.bind_all("<Shift-Button-5>", lambda e: self._on_hwheel(e), add="+")
    def _on_mousewheel(self, e):
        delta = -1 if getattr(e, "delta", 0) > 0 or getattr(e, "num", None) == 4 else 1
        self.body_canvas.yview_scroll(delta, "units")
    def _on_hwheel(self, e):
        self.header_canvas.xview_scroll(-1 if getattr(e, "delta", 0) > 0 else 1, "units")
        self.body_canvas.xview_scroll(-1 if getattr(e, "delta", 0) > 0 else 1, "units")

    # ===== Header =====
    def _build_header(self, initial: bool = False):
        if initial:
            for key, x, w in self._col_positions():
                cell = ctk.CTkFrame(self.header_frame, fg_color=DARK_BG3, width=w, height=self.header_h, corner_radius=0)
                cell.place(x=x, y=0)
                align = HEADER_ALIGN.get(key, "w")
                lbl = ctk.CTkLabel(cell, text=self._label_of(key), anchor=align, fg_color=DARK_BG3)
                if align == "w":
                    lbl.pack(fill="both", expand=True, padx=(HEADER_PADX_LEFT, 0))
                else:
                    lbl.pack(fill="both", expand=True)
                self._header_cells[key] = cell
        self._apply_positions(self._col_positions(), rebuild_resizers=True)

    def _label_of(self, key: str) -> str:
        for k, label, _ in self.cols:
            if k == key: return label
        return key

    def _clear_resizers(self):
        if hasattr(self, "_resizers"):
            for r in self._resizers:
                if r and str(r): r.destroy()
            self._resizers.clear()

    def _build_resizers(self, positions: List[Tuple[str, int, int]]):
        self._clear_resizers()
        for i in range(len(positions) - 1):
            _, x, w = positions[i]; boundary_x = x + w
            rz = tk.Frame(self.header_frame, width=8, height=self.header_h, cursor="sb_h_double_arrow",
                          highlightthickness=0, bd=0, bg=DARK_BG3)
            rz.place(x=boundary_x - 4, y=0)
            rz.bind("<Button-1>",       lambda e, idx=i: self._on_resizer_press(e, idx))
            rz.bind("<B1-Motion>",      lambda e, idx=i: self._on_resizer_drag(e, idx))
            rz.bind("<ButtonRelease-1>",lambda e, idx=i: self._on_resizer_release(e, idx))
            rz.bind("<Double-Button-1>",lambda e, idx=i: self._on_resizer_autofit(idx))
            self._resizers.append(rz)

    def _apply_positions(self, positions: List[Tuple[str, int, int]], rebuild_resizers: bool):
        x0 = self.body_canvas.xview()[0] if self.body_canvas.winfo_ismapped() else 0.0
        for key, x, w in positions:
            cell = self._header_cells.get(key)
            if cell and str(cell): cell.place_configure(x=x, y=0, width=w, height=self.header_h)
        xw = {k: (x, w) for (k, x, w) in positions}
        for _, cells_map in self._row_cells.items():
            for key, cell in cells_map.items():
                if key in xw and str(cell):
                    x, w = xw[key]; cell.place_configure(x=x, y=0, width=w, height=self.row_h)
        if rebuild_resizers: self._build_resizers(positions)
        self._resize_rows(); self._resize_header()
        self.update_idletasks()
        self.header_canvas.xview_moveto(x0); self.body_canvas.xview_moveto(x0)

    def _on_resizer_press(self, e, idx: int):
        pos = self._col_positions(); _, x, w = pos[idx]
        boundary_x = x + w
        self._resizer_state = {"idx": idx, "x0": e.x_root, "left_x": x, "start_w": w}
        if self._resizer_guide is None or not str(self._resizer_guide):
            self._resizer_guide = tk.Frame(self.header_frame, bg="#5aa0ff", width=2, height=self.header_h)
        self._resizer_guide.place(x=boundary_x, y=0)

    def _on_resizer_drag(self, e, idx: int):
        if not self._resizer_state or self._resizer_state.get("idx") != idx: return
        dx = e.x_root - self._resizer_state["x0"]
        temp_cols = list(self.cols)
        key, label, start_w = temp_cols[idx]
        temp_cols[idx] = (key, label, max(MIN_W.get(key, 60), start_w + dx))
        temp_pos = self._col_positions(temp_cols)
        new_boundary = temp_pos[idx][1] + temp_cols[idx][2]
        if self._resizer_guide and str(self._resizer_guide): self._resizer_guide.place(x=new_boundary, y=0)
        self._apply_positions(temp_pos, rebuild_resizers=False)

    def _on_resizer_release(self, e, idx: int):
        if not self._resizer_state or self._resizer_state.get("idx") != idx: return
        dx = e.x_root - self._resizer_state["x0"]
        key, label, start_w = self.cols[idx]
        new_w = max(MIN_W.get(key, 60), start_w + dx)
        self.cols[idx] = (key, label, int(new_w))
        self._resizer_state = None
        if self._resizer_guide and str(self._resizer_guide): self._resizer_guide.destroy(); self._resizer_guide = None
        self._apply_positions(self._col_positions(), rebuild_resizers=True); self._save_cols_config()

    def _on_resizer_autofit(self, idx: int):
        key, label, _ = self.cols[idx]
        est = max(MIN_W.get(key, 60), int(len(label) * 7 + 32))
        self.cols[idx] = (key, label, est)
        self._apply_positions(self._col_positions(), rebuild_resizers=True); self._save_cols_config()

    # ===== Layout dialogs =====
    def open_columns_dialog(self):
        win = ctk.CTkToplevel(self); win.title("Largura das colunas")
        self._center_toplevel(win, 520, 520)
        win.grab_set()
        
        frm = ctk.CTkScrollableFrame(win); frm.pack(fill="both", expand=True, padx=12, pady=12)
        entries: Dict[str, ctk.CTkEntry] = {}
        for key, label, width in self.cols:
            row = ctk.CTkFrame(frm, fg_color="transparent"); row.pack(fill="x", pady=4)
            ctk.CTkLabel(row, text=label, width=260, anchor="w").pack(side="left")
            e = ctk.CTkEntry(row, width=120); e.insert(0, str(width)); e.pack(side="left", padx=8)
            entries[key] = e
        btns = ctk.CTkFrame(win, fg_color="transparent"); btns.pack(fill="x", padx=12, pady=(0, 12))
        import re as _re
        def parse_int(s: str, default: int = 100) -> int:
            m = _re.findall(r"\d+", s or "")
            try: v = int("".join(m)) if m else default; return max(40, v)
            except Exception: return default
        def aplicar():
            self._push_undo()
            new_cols = list(self.cols)
            for key, ent in entries.items():
                try: v = parse_int(ent.get(), 100)
                except Exception: v = 100
                for idx, (k, label, w) in enumerate(new_cols):
                    if k == key: new_cols[idx] = (k, label, v); break
            self.cols = new_cols
            self._apply_positions(self._col_positions(), rebuild_resizers=True)
            self._save_cols_config()
        ctk.CTkButton(btns, text="Aplicar", width=110, command=aplicar).pack(side="right")
        ctk.CTkButton(btns, text="Fechar", width=110, command=win.destroy).pack(side="right", padx=6)

    def open_col_gap_dialog(self):
        win = ctk.CTkToplevel(self); win.title("Espaçamento entre colunas")
        self._center_toplevel(win, 380, 160)
        win.grab_set()

        ctk.CTkLabel(win, text="Gap (px) entre colunas:").pack(pady=(14, 6))
        e = ctk.CTkEntry(win, width=120); e.pack(); e.insert(0, str(int(self.col_gap)))
        import re as _re
        def ok():
            self._push_undo()
            s = e.get().strip(); m = _re.findall(r"\d+", s)
            if not m: return
            self.col_gap = max(0, int("".join(m)))
            self._apply_positions(self._col_positions(), rebuild_resizers=True)
            self._save_cols_config()
            win.destroy()
        ctk.CTkButton(win, text="OK", command=ok).pack(pady=10)

    def restore_default_columns(self):
        self._push_undo()
        self.cols = list(DEFAULT_COLS)
        try:
            self.tk_scale = float(self.tk.call("tk", "scaling"))
        except Exception:
            self.tk_scale = 1.0
        self.col_gap = max(5, int(round(5 * self.tk_scale)))
        self._apply_positions(self._col_positions(), rebuild_resizers=True)
        self._save_cols_config()
        messagebox.showinfo("Layout", "Larguras e espaçamento restaurados.")

    # ===== Gerenciador de Tarefas =====
    def open_tasks_dialog(self):
        win = ctk.CTkToplevel(self)
        win.title("Tarefas do fluxo")
        self._center_toplevel(win, 680, 580)
        win.grab_set()
        
        win.grid_columnconfigure(0, weight=1)
        win.grid_rowconfigure(1, weight=1)

        top_frame = ctk.CTkFrame(win, fg_color="transparent")
        top_frame.grid(row=0, column=0, sticky="ew", padx=10, pady=(10, 6))
        ctk.CTkLabel(top_frame, text=f"Fluxo: {self.project.flow_name}", anchor="w").pack(side="left")

        list_frame = ctk.CTkScrollableFrame(win)
        list_frame.grid(row=1, column=0, sticky="nsew", padx=10, pady=(0, 10))

        add_frame = ctk.CTkFrame(win)
        add_frame.grid(row=2, column=0, sticky="ew", padx=10, pady=(4, 8))
        add_frame.grid_columnconfigure(0, weight=1)

        entry_new_task = ctk.CTkEntry(add_frame, placeholder_text="Nome da nova tarefa")
        entry_new_task.grid(row=0, column=0, sticky="ew", padx=(0, 8))

        btn_add_another = ctk.CTkButton(add_frame, text="Salvar e Adicionar Outra", width=180)
        btn_add_another.grid(row=0, column=1, padx=(0, 8))

        bottom_frame = ctk.CTkFrame(win, fg_color="transparent")
        bottom_frame.grid(row=3, column=0, sticky="ew", padx=10, pady=(0, 10))
        ctk.CTkButton(bottom_frame, text="Fechar", width=120, command=win.destroy).pack(side="right")

        def refresh_all_ui():
            self._rebuild_metadata_cache()
            self._refresh_task_combo()
            self._refresh_rows()
            if self.sim_window and self.sim_window.winfo_exists():
                try: self.sim_window.on_model_changed()
                except Exception: pass

        def cleanup_refs(deleted_field_ids: Set[str]):
            if not deleted_field_ids: return
            for t in self.project.tasks:
                for f in t.fields:
                    f.cond = [c for c in f.cond if c.src_field not in deleted_field_ids]
                    if f.origin_field in deleted_field_ids:
                        if f.name_lock_reason == "origem" and f.name_before_origin: f.name = f.name_before_origin
                        f.origin_task = None; f.origin_field = None
                        f.name_lock_reason = "" if f.name_lock_reason == "origem" else f.name_lock_reason
                        f.name_locked = (f.name_lock_reason != "")

        def add_task_from_entry(close_after: bool):
            name = entry_new_task.get().strip()
            if not name:
                messagebox.showwarning("Nova Tarefa", "O nome da tarefa não pode ser vazio.", parent=win)
                return
            
            self._push_undo()
            t = Task(id=_uid(), name=name, fields=[])
            self.project.tasks.append(t)
            if not self.current_task_id:
                self.current_task_id = t.id
            
            entry_new_task.delete(0, "end")
            render()
            refresh_all_ui()
            
            if close_after:
                win.destroy()
            else:
                entry_new_task.focus()

        btn_add_another.configure(command=lambda: add_task_from_entry(False))
        entry_new_task.bind("<Return>", lambda event: add_task_from_entry(False))
        bottom_frame.winfo_children()[0].configure(command=lambda: add_task_from_entry(True) if entry_new_task.get().strip() else win.destroy())

        def rename_task(t: Task):
            new = self._prompt_text("Renomear tarefa", "Novo nome:", t.name)
            if new is not None:
                self._push_undo()
                t.name = new
                render(); refresh_all_ui()

        def delete_task(t: Task):
            if not messagebox.askyesno("Excluir tarefa", f"Excluir a tarefa '{t.name}' e todos os seus campos?\n\nAs referências a campos desta tarefa serão limpas.", parent=win):
                return
            self._push_undo()
            deleted_ids = {f.id for f in t.fields}
            cleanup_refs(deleted_ids)
            self.project.tasks = [x for x in self.project.tasks if x.id != t.id]
            if self.current_task_id == t.id:
                self.current_task_id = self.project.tasks[0].id if self.project.tasks else None
            render(); refresh_all_ui()

        def move_task(t: Task, delta: int):
            tasks = self.project.tasks
            i = next((idx for idx, x in enumerate(tasks) if x.id == t.id), -1)
            if i < 0: return
            j = max(0, min(len(tasks)-1, i+delta))
            if j == i: return
            self._push_undo()
            tasks[i], tasks[j] = tasks[j], tasks[i]
            render(); refresh_all_ui()

        def render():
            for w in list_frame.winfo_children(): w.destroy()
            if not self.project.tasks:
                ctk.CTkLabel(list_frame, text="Nenhuma tarefa. Use o campo abaixo para adicionar.").pack(pady=8)
                return
            for idx, t in enumerate(self.project.tasks):
                row = ctk.CTkFrame(list_frame, fg_color=THEME.color("surface"), corner_radius=6)
                row.pack(fill="x", padx=0, pady=4)
                ctk.CTkLabel(row, text=f"{idx+1}. {t.name}", anchor="w").pack(side="left", padx=8, pady=8)
                right = ctk.CTkFrame(row, fg_color="transparent"); right.pack(side="right", padx=6, pady=6)
                btn_up = ctk.CTkButton(
                    right,
                    text="▲",
                    width=36,
                    command=lambda tt=t: move_task(tt, -1),
                    state="normal" if idx > 0 else "disabled",
                )
                THEME.apply_button(btn_up, "ghost")
                btn_up.pack(side="left", padx=(0, 4))
                btn_down = ctk.CTkButton(
                    right,
                    text="▼",
                    width=36,
                    command=lambda tt=t: move_task(tt, +1),
                    state="normal" if idx < len(self.project.tasks) - 1 else "disabled",
                )
                THEME.apply_button(btn_down, "ghost")
                btn_down.pack(side="left", padx=(0, 8))
                ctk.CTkButton(right, text="Renomear", width=110, command=lambda tt=t: rename_task(tt)).pack(side="left", padx=(0, 8))
                btn_delete = ctk.CTkButton(right, text="Excluir", width=90, command=lambda tt=t: delete_task(tt))
                btn_delete.configure(text_color=THEME.color("danger_text"), hover_color=THEME.color("danger_hover"))
                btn_delete.pack(side="left")

        render()
        win.after(100, entry_new_task.focus)

    # ===== Otimização de Performance ao mover campos =====
    def _move_field(self, field_id: str, delta: int):
        task = self._get_task()
        if not task:
            return

        fields = task.fields
        idx = next((i for i, f in enumerate(fields) if f.id == field_id), -1)

        if idx == -1:
            return

        new_idx = idx + delta
        if 0 <= new_idx < len(fields):
            self._push_undo()

            # 1. Move no modelo de dados
            fields[idx], fields[new_idx] = fields[new_idx], fields[idx]

            # 2. Move o widget da linha na lista de widgets da UI
            row_widget = self._rows.pop(idx)
            self._rows.insert(new_idx, row_widget)

            # 3. Re-empacota os widgets na nova ordem (muito mais rápido que destruir e recriar)
            for w in self._rows:
                w.pack_forget()
            for w in self._rows:
                w.pack(fill="x", pady=0)

            # 4. Reconstrói o mapeamento de células de forma segura
            temp_cells = list(self._row_cells.values())
            moved_cell_dict = temp_cells.pop(idx)
            temp_cells.insert(new_idx, moved_cell_dict)
            self._row_cells = {i: d for i, d in enumerate(temp_cells)}

            # O rebuild do cache não é estritamente necessário aqui, mas não custa.
            self._rebuild_metadata_cache()


    # ===== Edição de linhas =====
    def _refresh_rows(self):
        for w in self.rows_frame.winfo_children(): w.destroy()
        self._rows.clear(); self._row_cells.clear(); self._field_row_map.clear()
        if not self.project.tasks:
            self._resize_rows(); self._resize_header(); return
        t = self._get_task()
        if not t: return
        
        positions = self._col_positions()

        for idx, f in enumerate(t.fields):
            self._add_row_widget(idx, f, t, positions)

        self._apply_positions(self._col_positions(), rebuild_resizers=True)

    def _update_single_row_widgets(self, f: Field):
        """Otimizado: Atualiza valores simples e RECRIANDO widgets complexos se necessário."""
        row = self._field_row_map.get(f.id)
        if not row or row not in self._rows: return
        idx = self._rows.index(row)
        widgets = self._row_cells.get(idx, {})

        # 1. Atualiza Cores e Textos Simples (Sem custo)
        bg_color = "#171a1f" if (f.origin_task and f.origin_field) else "transparent"
        try: row.configure(fg_color=bg_color)
        except: pass

        if "origem" in widgets:
            for child in widgets["origem"].winfo_children():
                if isinstance(child, ctk.CTkButton):
                    origem_label = self._origin_summary(f)
                    if f.origin_task and f.origin_field: origem_label = "🔗 " + origem_label
                    child.configure(text=origem_label)

        if "regras" in widgets:
            summary = self._cond_summary(f)
            for frame in widgets["regras"].winfo_children():
                if isinstance(frame, ctk.CTkFrame):
                    for lbl in frame.winfo_children():
                        if isinstance(lbl, ctk.CTkLabel):
                            lbl.configure(text=summary)

        # 2. Atualiza Checkboxes
        for key, attr in [("obrig", "required"), ("soleit", "readonly")]:
            if key in widgets:
                for child in widgets[key].winfo_children():
                    if isinstance(child, ctk.CTkCheckBox):
                        val = getattr(f, attr)
                        if val: child.select()
                        else: child.deselect()
                        if key == "soleit" and f.ftype == "Informativo": child.configure(state="disabled")
                        elif key == "soleit": child.configure(state="normal")

        # 3. Atualiza Nome e Observações (Sempre Entry, seguro atualizar)
        for key, attr in [("obs", "note"), ("campo", "name")]:
            if key in widgets:
                for child in widgets[key].winfo_children():
                    if isinstance(child, (ctk.CTkEntry, tk.Entry)):
                        if self.focus_get() != child:
                            current = child.get()
                            val = getattr(f, attr) or ""
                            if current != val:
                                child.delete(0, "end")
                                child.insert(0, val)

        # 4. CRÍTICO: Recria a Célula de Opções ('opts') se o widget não bater com o tipo
        # Isso evita o lag de tentar adaptar widgets incompatíveis
        if "opts" in widgets:
            cell_frame = widgets["opts"]
            children = cell_frame.winfo_children()
            needs_rebuild = False
            
            if not children:
                needs_rebuild = True
            else:
                widget = children[0]
                is_button = isinstance(widget, ctk.CTkButton)
                is_entry = isinstance(widget, (ctk.CTkEntry, tk.Entry))
                
                if f.ftype == "Objeto" and not is_button: needs_rebuild = True
                elif f.ftype != "Objeto" and not is_entry: needs_rebuild = True

            if needs_rebuild:
                # Limpa célula
                for child in children: child.destroy()
                
                # Recria widget correto
                if f.ftype == "Objeto":
                    label = f"Objeto do fluxo: {self.project.object_type or '(defina em Objetos > Tipo...)'}"
                    btn = ctk.CTkButton(cell_frame, text=label + "  (Esquema…)", command=self.open_object_schema_editor)
                    btn.pack(fill="both", expand=True)
                    # Rebind menu context
                    btn.bind("<Button-3>", lambda e: self._show_context_menu(e, f))
                else:
                    eopt = ctk.CTkEntry(cell_frame)
                    eopt.insert(0, f.options or "")
                    eopt.pack(fill="both", expand=True)
                    eopt.bind("<FocusOut>", lambda _=None, w=eopt: (self._push_undo(), setattr(f, "options", w.get())))
                    eopt.bind("<Button-3>", lambda e: self._show_context_menu(e, f))
            else:
                # Se o widget já é do tipo certo, só atualiza o valor
                if children and isinstance(children[0], (ctk.CTkEntry, tk.Entry)):
                    w = children[0]
                    if self.focus_get() != w:
                        curr = w.get()
                        val = f.options or ""
                        if curr != val:
                            w.delete(0, "end")
                            w.insert(0, val)

        # 5. Sincroniza o Menu de Tipo
        if "tipo" in widgets:
            for child in widgets["tipo"].winfo_children():
                if isinstance(child, ctk.CTkOptionMenu):
                    if child.get() != f.ftype: child.set(f.ftype)

    def _append_row_widget(self, idx_row: int, f: Field, t: Task):
        positions = self._col_positions()
        self._add_row_widget(idx_row, f, t, positions)
        self._apply_positions(positions, rebuild_resizers=False)

    def _add_row_widget(self, idx_row: int, f: Field, t: Task, positions: List[Tuple[str, int, int]]):
        bg = HILIGHT_ORIGIN_BG if (f.origin_task and f.origin_field) else "transparent"
        row = ctk.CTkFrame(self.rows_frame, fg_color=bg, height=self.row_h, corner_radius=0)
        row.pack(fill="x", pady=0)
        self._rows.append(row); self._row_cells[idx_row] = {}
        if f.id:
            self._field_row_map[f.id] = row
        
        # Função de callback para o menu de contexto. "Congela" o campo atual.
        show_menu_func = lambda e, field_obj=f: self._show_context_menu(e, field_obj)
        row.bind("<Button-3>", show_menu_func)

        def cell(key: str) -> ctk.CTkFrame:
            col_data = next((tup for tup in positions if tup[0] == key), None)
            if not col_data: return ctk.CTkFrame(row)
            _, x, w = col_data
            cont = ctk.CTkFrame(row, fg_color="transparent", width=w, height=self.row_h, corner_radius=0)
            cont.place(x=x, y=0); self._row_cells[idx_row][key] = cont
            cont.bind("<Button-3>", show_menu_func)
            return cont

        # move
        cmove = cell("move")
        btn_frm = ctk.CTkFrame(cmove, fg_color="transparent")
        btn_frm.pack(expand=True)
        btn_frm.bind("<Button-3>", show_menu_func)
        btn_up = ctk.CTkButton(btn_frm, text="▲", width=24, command=lambda fid=f.id: self._move_field(fid, -1))
        btn_up.pack(side="left", padx=(0,2))
        btn_up.bind("<Button-3>", show_menu_func)
        btn_down = ctk.CTkButton(btn_frm, text="▼", width=24, command=lambda fid=f.id: self._move_field(fid, 1))
        btn_down.pack(side="left")
        btn_down.bind("<Button-3>", show_menu_func)

        # sel
        cs = cell("sel")
        v = tk.BooleanVar(value=(f.id in self.selected_field_ids))
        chk = ctk.CTkCheckBox(cs, text="", variable=v, command=lambda fid=f.id, var=v: self._toggle_select(fid, var.get()))
        chk.pack(expand=True)
        chk.bind("<Button-3>", show_menu_func)

        # nome
        ce = cell("campo")
        name_disabled = (f.name_lock_reason in ("objeto","origem")) or f.name_locked
        e = ctk.CTkEntry(ce); e.insert(0, f.name or "")
        if name_disabled: e.configure(state="disabled")
        e.pack(fill="both", expand=True)
        e.bind("<Button-3>", show_menu_func)
        if not name_disabled:
            e.bind("<FocusOut>", lambda _=None, w=e: self._on_field_name_changed(t.id, f, w.get()))

        # tipo
        ct = cell("tipo")
        om = ctk.CTkOptionMenu(ct, values=TYPE_VALUES, command=lambda *_: self._on_change_type(f, om.get()))
        base = _solid_color(); om.configure(fg_color=base, button_color=base, button_hover_color=base)
        om.set(f.ftype); om.pack(fill="both", expand=True)
        om.bind("<Button-3>", show_menu_func)
        # O ctk.CTkOptionMenu é complexo, então vinculamos seus filhos também por segurança
        for child in om.winfo_children():
            child.bind("<Button-3>", show_menu_func)

        # origem
        co = cell("origem")
        origem_label = self._origin_summary(f)
        if f.origin_task and f.origin_field: origem_label = "🔗 " + origem_label
        btn_origin = ctk.CTkButton(co, text=origem_label, command=lambda: self.open_origin_picker(f))
        btn_origin.pack(fill="both", expand=True)
        btn_origin.bind("<Button-3>", show_menu_func)

        # regras
        cr = cell("regras")
        summary = self._cond_summary(f)
        rule_frame = ctk.CTkFrame(cr, fg_color="transparent", corner_radius=6)
        rule_frame.pack(fill="both", expand=True)
        col_w = next((c[2] for c in self.cols if c[0] == 'regras'), 300)
        rule_label = ctk.CTkLabel(rule_frame, text=summary, anchor="w", justify="left", wraplength=col_w - 20)
        rule_label.pack(fill="both", expand=True, padx=10)

        def on_enter(e): rule_frame.configure(fg_color=DARK_BG3)
        def on_leave(e): rule_frame.configure(fg_color="transparent")
        
        rule_frame.bind("<Enter>", on_enter); rule_frame.bind("<Leave>", on_leave)
        rule_frame.bind("<Button-1>", lambda e: self.open_cond_builder(f))
        rule_label.bind("<Button-1>", lambda e: self.open_cond_builder(f))
        rule_frame.bind("<Button-3>", show_menu_func)
        rule_label.bind("<Button-3>", show_menu_func)

        # flags
        cobr = cell("obrig")
        var_req = tk.BooleanVar(value=f.required)
        chk_req = ctk.CTkCheckBox(cobr, text="", variable=var_req, command=lambda: (self._push_undo(), setattr(f, "required", var_req.get())))
        chk_req.pack(expand=True)
        chk_req.bind("<Button-3>", show_menu_func)

        csol = cell("soleit")
        var_ro = tk.BooleanVar(value=f.readonly or (f.ftype == "Informativo"))
        chk_ro = ctk.CTkCheckBox(csol, text="", variable=var_ro, command=lambda: self._set_readonly(f, var_ro.get()))
        if f.ftype == "Informativo":
            f.readonly = True
            chk_ro.configure(state="disabled")
        chk_ro.pack(expand=True)
        chk_ro.bind("<Button-3>", show_menu_func)

        # opções
        copts = cell("opts")
        if f.ftype == "Objeto":
            label = f"Objeto do fluxo: {self.project.object_type or '(defina em Objetos > Tipo...)'}"
            btn_opts = ctk.CTkButton(copts, text=label + "  (Esquema…)", command=self.open_object_schema_editor)
            btn_opts.pack(fill="both", expand=True)
            btn_opts.bind("<Button-3>", show_menu_func)
        else:
            eopt = ctk.CTkEntry(copts); eopt.insert(0, f.options or ""); eopt.pack(fill="both", expand=True)
            eopt.bind("<FocusOut>", lambda _=None, w=eopt: (self._push_undo(), setattr(f, "options", w.get())))
            eopt.bind("<Button-3>", show_menu_func)

        # observações
        cobs = cell("obs")
        eobs = ctk.CTkEntry(cobs); eobs.insert(0, f.note or ""); eobs.pack(fill="both", expand=True)
        eobs.bind("<FocusOut>", lambda _=None, w=eobs: (self._push_undo(), setattr(f, "note", w.get())))
        eobs.bind("<Button-3>", show_menu_func)

        # delete
        cdel = cell("del")
        btn_del = ctk.CTkButton(cdel, text="x", command=lambda fid=f.id: self._delete_field(fid))
        btn_del.pack(expand=True)
        btn_del.bind("<Button-3>", show_menu_func)


    def _toggle_select(self, fid: str, val: bool):
        if val: self.selected_field_ids.add(fid)
        else: self.selected_field_ids.discard(fid)

    def delete_selected_fields(self):
        self._commit_active_edits()
        if not self.selected_field_ids:
            messagebox.showinfo("Excluir", "Nenhum campo selecionado."); return
        if not messagebox.askyesno("Excluir", f"Excluir {len(self.selected_field_ids)} campo(s) selecionado(s)?"):
            return
        self._push_undo()
        t = self._get_task()
        to_del = set(self.selected_field_ids)
        for task in self.project.tasks:
            for f in task.fields:
                f.cond = [c for c in f.cond if c.src_field not in to_del]
                if f.origin_field in to_del:
                    if f.name_lock_reason == "origem" and f.name_before_origin:
                        f.name = f.name_before_origin
                    f.origin_task = None; f.origin_field = None
                    f.name_lock_reason = "" if f.name_lock_reason == "origem" else f.name_lock_reason
                    f.name_locked = (f.name_lock_reason != "")
        t.fields = [f for f in t.fields if f.id not in to_del]
        self.selected_field_ids.clear()
        self._rebuild_metadata_cache()
        self._refresh_rows()
        if self.sim_window and self.sim_window.winfo_exists():
            try: self.sim_window.on_model_changed()
            except Exception: pass

    # ===== Copiar/Colar Campos =====
    def _copy_selected_fields(self):
        self._commit_active_edits()
        if not self.selected_field_ids:
            messagebox.showinfo("Copiar", "Nenhum campo selecionado para copiar.")
            return

        current_task = self._get_task()
        if not current_task:
            return

        self._clipboard.clear()
        
        fields_to_copy = []
        for field in current_task.fields:
            if field.id in self.selected_field_ids:
                field_data = {
                    "id_origem": field.id,
                    "name": field.name, "ftype": field.ftype,
                    "required": field.required, "readonly": field.readonly,
                    "info": field.info, "options": field.options, "note": field.note,
                    "cond": [{"src_field": c.src_field, "op": c.op, "value": c.value} for c in field.cond],
                }
                fields_to_copy.append(field_data)
        
        if fields_to_copy:
            # Armazena a tarefa de origem junto com os campos
            self._clipboard = {
                "source_task_id": current_task.id,
                "fields": fields_to_copy
            }
            self.btn_paste.configure(state="normal")
            messagebox.showinfo("Copiar", f"{len(fields_to_copy)} campo(s) copiado(s) para a área de transferência.")
    
    def _open_paste_dialog(self) -> Optional[str]:
        """Abre um diálogo para o usuário escolher o tipo de colagem. Retorna 'copy', 'origin', ou None."""
        win = ctk.CTkToplevel(self); win.title("Opções de Colagem")
        win.transient(self) # Garante que a janela abra no monitor correto
        self._center_toplevel(win, 480, 200)
        win.grab_set()

        ctk.CTkLabel(win, text="Como você deseja colar os campos?", font=("Segoe UI", 14, "bold")).pack(pady=(20, 10))
        
        result = {"value": None}
        
        def set_choice(choice: str):
            result["value"] = choice
            win.destroy()

        ctk.CTkButton(win, text="Colar como Cópia Independente", command=lambda: set_choice("copy")).pack(fill="x", padx=20, pady=5)
        ctk.CTkButton(win, text="Colar Vinculado à Origem", command=lambda: set_choice("origin")).pack(fill="x", padx=20, pady=5)
        
        self.wait_window(win)
        return result["value"]

    def _paste_fields(self):
        self._commit_active_edits()
        if not self._clipboard or not self._clipboard.get("fields"):
            messagebox.showwarning("Colar", "Nenhum campo na área de transferência para colar.")
            return
            
        target_task = self._get_task()
        if not target_task:
            messagebox.showerror("Colar", "Nenhuma tarefa de destino selecionada.")
            return
        
        choice = self._open_paste_dialog()
        if not choice:
            return

        self._push_undo()
        
        if choice == "copy":
            self._execute_paste_as_copy(target_task)
        elif choice == "origin":
            self._execute_paste_with_origin(target_task)

        self._rebuild_metadata_cache()
        self.selected_field_ids.clear()
        self._refresh_rows()

    def _execute_paste_as_copy(self, target_task: Task):
        fields_to_paste = self._clipboard.get("fields", [])
        original_to_new_id = {}
        new_fields_to_add = []

        for field_data in fields_to_paste:
            new_id = _uid()
            original_id = field_data.get("id_origem")
            is_object = field_data["ftype"] == "Objeto"
            
            new_field = Field(
                id=new_id, name=field_data["name"], ftype=field_data["ftype"],
                required=field_data["required"], readonly=field_data["readonly"],
                info=field_data["info"], options=field_data["options"], note=field_data["note"],
                name_locked=is_object, name_lock_reason="objeto" if is_object else "",
                obj_type=self.project.object_type if is_object else "", cond=[]
            )
            
            new_fields_to_add.append((new_field, field_data.get("cond", [])))
            if original_id:
                original_to_new_id[original_id] = new_id

        target_task.fields.extend([f for f, _ in new_fields_to_add])

        all_fields_map = {f.id: f for t in self.project.tasks for f in t.fields}
        missing_dependencies: List[str] = []

        for new_field, original_conds in new_fields_to_add:
            for cond_data in original_conds:
                src_id_orig = cond_data["src_field"]
                new_src_id = original_to_new_id.get(src_id_orig)
                if not new_src_id and src_id_orig in all_fields_map:
                    new_src_id = src_id_orig
                if new_src_id:
                    new_field.cond.append(Condition(src_field=new_src_id, op=cond_data["op"], value=cond_data["value"]))
                else:
                    new_field.cond.append(Condition(src_field=src_id_orig, op=cond_data["op"], value=cond_data["value"]))
                    missing_dependencies.append(src_id_orig)

        pasted_count = len(fields_to_paste)
        if missing_dependencies:
            missing_list = "\n".join(f"- [id {fid}]" for fid in sorted(set(missing_dependencies)))
            messagebox.showwarning(
                "Colar",
                f"{pasted_count} campo(s) colado(s) como cópia, porém algumas regras mantiveram referências a campos removidos.\n\n"
                f"Revise as condições associadas aos identificadores abaixo:\n{missing_list}",
            )
        else:
            messagebox.showinfo("Colar", f"{pasted_count} campo(s) colado(s) como cópia.")

    def _execute_paste_with_origin(self, target_task: Task):
        source_task_id = self._clipboard.get("source_task_id")
        fields_to_paste = self._clipboard.get("fields", [])
        
        if not source_task_id:
            messagebox.showerror("Colar com Origem", "Erro: A tarefa de origem não foi encontrada no clipboard.")
            return

        all_fields_map = {f.id: f for t in self.project.tasks for f in t.fields}
        
        for field_data in fields_to_paste:
            origin_field_id = field_data.get("id_origem")
            if not origin_field_id or origin_field_id not in all_fields_map:
                continue

            origin_field_obj = all_fields_map[origin_field_id]

            new_field = Field(
                id=_uid(),
                name=origin_field_obj.name,
                ftype=origin_field_obj.ftype,
                options=origin_field_obj.options,
                origin_task=source_task_id,
                origin_field=origin_field_id,
                name_locked=True,
                name_lock_reason="origem",
                readonly=True, # Comportamento padrão seguro
                required=False, # Comportamento padrão seguro
                cond=[] # Condições não são herdadas ao colar com origem
            )
            target_task.fields.append(new_field)
            
        messagebox.showinfo("Colar", f"{len(fields_to_paste)} campo(s) colado(s) com vínculo à origem.")

    # ===== Helpers de dados =====
    def _refresh_flow_label(self):
        obj = f" | Objeto: {self.project.object_type}" if self.project.object_type else ""
        self.lbl_flow.configure(text=f"Fluxo: {self.project.flow_name}{obj}")

    def _focus_on_field(self, task_id: Optional[str], field_id: Optional[str]):
        if not task_id:
            return
        target_task = next((t for t in self.project.tasks if t.id == task_id), None)
        if not target_task:
            return

        self.deiconify()
        try:
            self.focus_set()
        except Exception:
            pass

        self.current_task_id = target_task.id
        self._refresh_task_combo()
        self._refresh_rows()

        if not field_id:
            return

        row = self._field_row_map.get(field_id)
        if not row:
            return

        self.update_idletasks()
        bbox = self.body_canvas.bbox(self.body_window)
        total_height = max(1, (bbox[3] - bbox[1]) if bbox else self.rows_frame.winfo_height())
        canvas_height = max(1, self.body_canvas.winfo_height())
        target_top = row.winfo_y()
        target_bottom = target_top + (row.winfo_height() or self.row_h)
        view_top = self.body_canvas.canvasy(0)
        view_bottom = view_top + canvas_height

        if target_top < view_top or target_bottom > view_bottom:
            new_top = max(0, target_top - max(0, (canvas_height - (target_bottom - target_top)) // 2))
            self.body_canvas.yview_moveto(min(new_top / total_height, 1))

        self._flash_row(field_id)

    def _flash_row(self, field_id: str, flashes: int = 4):
        row = self._field_row_map.get(field_id)
        if not row:
            return
        try:
            original_color = row.cget("fg_color")
        except Exception:
            original_color = getattr(row, "fg_color", "transparent")

        highlight_color = "#2a3642"

        def toggle(step: int = 0):
            if not row.winfo_exists():
                return
            if step >= flashes * 2:
                try:
                    row.configure(fg_color=original_color)
                except Exception:
                    pass
                return
            try:
                row.configure(fg_color=highlight_color if step % 2 == 0 else original_color)
            except Exception:
                return
            row.after(130, lambda: toggle(step + 1))

        toggle()

    def _get_task(self, task_id: Optional[str] = None) -> Optional[Task]:
        tid = task_id or self.current_task_id
        if tid is None and self.project.tasks:
            self.current_task_id = self.project.tasks[0].id
            return self.project.tasks[0]
        for t in self.project.tasks:
            if t.id == tid: return t
        return None

    def _refresh_task_combo(self):
        display_names = [f"{i+1}. {t.name}" for i, t in enumerate(self.project.tasks)]
        self.cmb_task.configure(values=display_names)
        
        current_task = self._get_task()
        if current_task:
            try:
                current_idx = self.project.tasks.index(current_task)
                self.cmb_task.set(display_names[current_idx])
            except ValueError:
                if display_names: self.cmb_task.set(display_names[0])
                else: self.cmb_task.set("")
        elif display_names:
            self.cmb_task.set(display_names[0])
            self.current_task_id = self.project.tasks[0].id
        else:
            self.cmb_task.set("")
            self.current_task_id = None

    def _on_task_change(self):
        selected_display_name = self.cmb_task.get()
        if not selected_display_name:
            return

        try:
            task_index = int(selected_display_name.split('.')[0]) - 1
            if 0 <= task_index < len(self.project.tasks):
                self.current_task_id = self.project.tasks[task_index].id
        except (ValueError, IndexError):
            name_part = selected_display_name.split('. ', 1)[-1]
            for t in self.project.tasks:
                if t.name == name_part:
                    self.current_task_id = t.id
                    break
        
        self._refresh_rows()

    def _on_field_name_changed(self, task_id: str, f: Field, new_name: str):
        self._push_undo()
        f.name = new_name
        self._rebuild_metadata_cache() 
        for t in self.project.tasks:
            for fld in t.fields:
                if fld.origin_field == f.id and fld.name_lock_reason == "origem":
                    fld.name = new_name

    def _origin_summary(self, f: Field) -> str:
        if f.origin_task and f.origin_field:
            task_name = self._get_task_name(f.origin_task)
            field_name = self._get_field_name(f.origin_field)
            return f"{task_name} › {field_name}"
        return "Sem origem"

    def _get_task_name(self, task_id: Optional[str]) -> str:
        if not task_id: return ""
        return self._task_id_to_name.get(task_id, "")

    def _get_field_name(self, field_id: Optional[str]) -> str:
        if not field_id: return ""
        return self._field_id_to_name.get(field_id, "")

    def _cond_summary_for_task(self, task: Task, f: Field) -> str:
        if not f.cond: return ""
        parts = []
        for c in f.cond:
            base_name = self._get_field_name(c.src_field)
            parts.append(f"{base_name or f'[id {c.src_field}]'} {c.op} {c.value!r}")
        return " OU ".join(parts)

    def _format_field_subtype(self, field: Field) -> str:
        if field.ftype in LIST_FIELD_TYPES and field.options:
            return f"Opções: {field.options}"
        if field.ftype == "Informativo" and field.options:
            return f"Texto: {field.options}"
        if field.ftype == "Objeto":
            return f"Objeto: {self.project.object_type or '-'}"
        return field.ftype

    def _overview_xlsx_initial_name(self) -> str:
        base = self.project.flow_name or "Fluxo"
        normalized = unicodedata.normalize("NFKD", base)
        ascii_name = normalized.encode("ascii", "ignore").decode("ascii")
        safe = re.sub(r"[^A-Za-z0-9]+", "_", ascii_name).strip("_")
        if not safe:
            safe = "Fluxo"
        return f"Campos_{safe}.xlsx"

    def _cond_summary(self, f: Field) -> str:
        t = self._get_task()
        if not t: return "Sem regras"
        return ("Sem regras" if not f.cond else "Exibir quando " + self._cond_summary_for_task(t, f))

    # ===== Tipo/ReadOnly =====
    def _on_change_type(self, f: Field, new_type: str):
        self._commit_row_data(f)
        prev = f.ftype
        if new_type == prev: return

        # Logica de negócio
        self._push_undo()
        
        # Regras especiais de Objeto
        if new_type == "Objeto":
            if f.origin_task or f.origin_field:
                messagebox.showwarning("Objeto", "Campo Objeto não pode ter origem."); self._update_single_row_widgets(f); return
            t = self._get_task()
            if any(x.ftype == "Objeto" and x.id != f.id for x in t.fields):
                messagebox.showwarning("Objeto", "Só 1 Objeto por tarefa."); self._update_single_row_widgets(f); return
            if not self.project.object_type:
                if not self.open_flow_object_type_dialog():
                    self._update_single_row_widgets(f); return
            
            f.ftype = "Objeto"
            if f.name_lock_reason != "objeto": f.name_before_obj = f.name
            f.name = self.project.object_type or "Objeto"
            f.name_lock_reason = "objeto"; f.name_locked = True; f.obj_type = self.project.object_type
        
        elif prev == "Objeto":
            if f.name_lock_reason == "objeto" and f.name_before_obj: f.name = f.name_before_obj
            f.name_lock_reason = "" if f.name_lock_reason == "objeto" else f.name_lock_reason
            f.name_locked = (f.name_lock_reason != ""); f.obj_type = ""
            f.ftype = new_type
        
        else:
            f.ftype = new_type

        if new_type == "Informativo": f.readonly = True
        if new_type not in LIST_FIELD_TYPES and new_type != "Informativo": f.options = ""

        # Se virou Anexo, abre editor (opcional)
        if new_type == "Anexo" and prev != "Anexo":
            self._open_attachment_type_editor(f)

        # Reconstrói cache se o nome mudou (ex: virou Objeto)
        if f.name_locked or prev == "Objeto":
            self._rebuild_metadata_cache()

        # CHAMA A NOVA ATUALIZAÇÃO OTIMIZADA
        self._update_single_row_widgets(f)

    def _set_readonly(self, f: Field, val: bool):
        self._commit_row_data(f)
        self._push_undo()
        if f.ftype == "Informativo":
            f.readonly = True
        else:
            f.readonly = val

    # ===== Operações de adicionar/excluir campo =====
    def _add_field(self):
        self._commit_active_edits()
        if not self.project.tasks:
            if messagebox.askyesno("Sem tarefas", "Este fluxo está vazio. Deseja abrir o gerenciador de tarefas para criar uma?"):
                self.open_tasks_dialog()
            return
        self._push_undo()
        t = self._get_task()
        f = Field(id=_uid())
        t.fields.append(f)
        self._rebuild_metadata_cache()
        self._append_row_widget(len(t.fields)-1, f, t)

    def _delete_field(self, field_id: str):
        if not self.project.tasks:
            return
        self._push_undo()
        current_task = self._get_task()
        
        for task in self.project.tasks:
            for f in task.fields:
                f.cond = [c for c in f.cond if c.src_field != field_id]
                if f.origin_field == field_id:
                    if f.name_lock_reason == "origem" and f.name_before_origin:
                        f.name = f.name_before_origin
                    f.origin_task = None
                    f.origin_field = None
                    f.name_lock_reason = "" if f.name_lock_reason == "origem" else f.name_lock_reason
                    f.name_locked = (f.name_lock_reason != "")

        field_to_remove_idx = -1
        for i, f in enumerate(current_task.fields):
            if f.id == field_id:
                field_to_remove_idx = i
                break
                
        current_task.fields = [f for f in current_task.fields if f.id != field_id]

        if field_to_remove_idx != -1 and field_to_remove_idx < len(self._rows):
            row_widget_to_remove = self._rows.pop(field_to_remove_idx)
            row_widget_to_remove.destroy()
            self._row_cells.pop(field_to_remove_idx, None)
            # Reindexa o dicionário de células
            new_row_cells = {}
            for i, (old_idx, value) in enumerate(self._row_cells.items()):
                new_row_cells[i] = value
            self._row_cells = new_row_cells
        else:
            self._refresh_rows() # Fallback para garantir consistência

        self._rebuild_metadata_cache()
        
        if self.sim_window and self.sim_window.winfo_exists():
            try: self.sim_window.on_model_changed()
            except Exception: pass

    # ===== Ações do Menu de Contexto =====
    def _show_context_menu(self, event, target_field: Field):
        """Exibe o menu de contexto para um campo específico."""
        self._commit_active_edits()
        if self.context_menu and self.context_menu.winfo_exists():
            self.context_menu.destroy()
        self.context_menu = CustomContextMenu(self, event, target_field, self)

    def _copy_single_field(self, field: Field):
        """Prepara o clipboard com um único campo para ser colado."""
        self._clipboard.clear()
        current_task = self._get_task()
        if not current_task: return

        field_data = {
            "id_origem": field.id, "name": field.name, "ftype": field.ftype,
            "required": field.required, "readonly": field.readonly,
            "info": field.info, "options": field.options, "note": field.note,
            "cond": [{"src_field": c.src_field, "op": c.op, "value": c.value} for c in field.cond],
        }
        self._clipboard = {"source_task_id": current_task.id, "fields": [field_data]}
        self.btn_paste.configure(state="normal")

    def _cut_single_field(self, field: Field):
        """Copia um único campo e depois o exclui (Recortar)."""
        self._copy_single_field(field)
        self._delete_field(field.id)

    def _duplicate_field(self, field: Field):
        """Cria uma cópia de um campo logo abaixo do original."""
        self._push_undo()
        task = self._get_task()
        if not task: return

        try:
            original_index = task.fields.index(field)
        except ValueError:
            original_index = len(task.fields) - 1

        new_field = Field(
            id=_uid(), name=f"{field.name} (Cópia)", ftype=field.ftype,
            required=field.required, readonly=field.readonly,
            info=field.info, options=field.options, note=field.note,
            origin_task=None, origin_field=None, # Duplicatas são independentes
            name_locked=field.name_locked, name_lock_reason=field.name_lock_reason,
            obj_type=field.obj_type,
            cond=[Condition(**vars(c)) for c in field.cond] # Cópia profunda das condições
        )
        task.fields.insert(original_index + 1, new_field)
        self._rebuild_metadata_cache()
        self._refresh_rows()

    def _move_field_to_top(self, field: Field):
        """Move um campo para o início da lista na tarefa atual."""
        self._push_undo()
        task = self._get_task()
        if not task: return
        try:
            task.fields.remove(field)
            task.fields.insert(0, field)
            self._refresh_rows()
        except ValueError: pass # Campo não encontrado

    def _move_field_to_end(self, field: Field):
        """Move um campo para o fim da lista na tarefa atual."""
        self._push_undo()
        task = self._get_task()
        if not task: return
        try:
            task.fields.remove(field)
            task.fields.append(field)
            self._refresh_rows()
        except ValueError: pass # Campo não encontrado

    def _open_attachment_type_editor(self, field: Field):
        """Abre um diálogo para editar a tag [Tipo de Doc.:] na nota de um campo."""
        pattern = r"\[Tipo de Doc\.:\s*(.*?)\s*\]"
        match = re.search(pattern, field.note)
        initial_content = match.group(1) if match else ""

        new_content = self._prompt_attachment_types(initial_content)
        if new_content is None:
            return  # Usuário cancelou

        normalized_content = new_content.strip()

        self._push_undo()
        new_tag = f"[Tipo de Doc.: {normalized_content}]" if normalized_content else ""

        if match: # Tag já existe
            if not new_tag: # Novo conteúdo está vazio, remove a tag antiga
                field.note = (field.note[:match.start()] + field.note[match.end():]).strip()
            else: # Substitui a tag antiga pela nova
                field.note = field.note[:match.start()] + new_tag + field.note[match.end():]
        elif new_tag: # Tag não existe e novo conteúdo foi fornecido
            field.note = (new_tag + " " + field.note).strip()
        
        self._update_single_row_widgets(field)

    def _prompt_attachment_types(self, initial: str) -> Optional[str]:
        sem_limitation_label = "Sem Limitação"
        initial_clean = (initial or "").strip()
        sem_lim_initial = initial_clean.lower() == sem_limitation_label.lower()

        win = ctk.CTkToplevel(self)
        win.title("Tipos de Documento para Anexo")
        win.transient(self)
        self._center_toplevel(win, 520, 210)
        win.grab_set()

        container = ctk.CTkFrame(win, fg_color="transparent")
        container.pack(fill="both", expand=True, padx=18, pady=16)

        ctk.CTkLabel(
            container,
            text=(
                "Informe os tipos permitidos separados por ';' ou marque "
                "'Sem Limitação' para aceitar qualquer documento."
            ),
            justify="left",
            anchor="w",
            wraplength=460,
        ).pack(fill="x")

        input_row = ctk.CTkFrame(container, fg_color="transparent")
        input_row.pack(fill="x", pady=(12, 0))

        var_sem_lim = tk.BooleanVar(value=sem_lim_initial)

        entry = ctk.CTkEntry(input_row)
        entry.pack(side="right", fill="x", expand=True)
        if not sem_lim_initial:
            entry.insert(0, initial_clean)

        def toggle_state() -> None:
            if var_sem_lim.get():
                entry.delete(0, tk.END)
                entry.configure(state="disabled")
            else:
                entry.configure(state="normal")
                entry.focus_set()

        chk = ctk.CTkCheckBox(
            input_row,
            text=sem_limitation_label,
            variable=var_sem_lim,
            command=toggle_state,
        )
        chk.pack(side="left", padx=(0, 12))

        toggle_state()

        output: Dict[str, Optional[str]] = {"value": None}

        def finish(ok: bool) -> None:
            if ok:
                if var_sem_lim.get():
                    output["value"] = sem_limitation_label
                else:
                    output["value"] = entry.get() if entry.get() is not None else ""
            else:
                output["value"] = None
            win.destroy()

        buttons = ctk.CTkFrame(container, fg_color="transparent")
        buttons.pack(fill="x", pady=18)

        ctk.CTkButton(buttons, text="OK", width=140, command=lambda: finish(True)).pack(side="left")
        btn_cancel = ctk.CTkButton(buttons, text="Cancelar", width=140, command=lambda: finish(False))
        _apply_secondary_style(btn_cancel)
        btn_cancel.pack(side="left", padx=(12, 0))

        entry.bind("<Return>", lambda _event: finish(True))
        win.bind("<Escape>", lambda _event: finish(False))

        if not var_sem_lim.get():
            entry.focus_set()
        else:
            chk.focus_set()

        self.wait_window(win)
        return output["value"]
        
    # ===== Visão Planilha (HTML) =====
    def _build_overview_html(self, query: str, collapsed: Set[str]) -> str:
        q = (query or "").lower()
        
        tasks_with_rows: List[Tuple[Task, List[List[str]]]] = []
        
        for t in self.project.tasks:
            task_rows = []
            for f in t.fields:
                origem = f"{self._get_task_name(f.origin_task)} › {self._get_field_name(f.origin_field)}" if (f.origin_task and f.origin_field) else ""
                regras = self._cond_summary_for_task(t, f)
                subtipo = self._format_field_subtype(f)
                
                row_data = [f.name, f.ftype, origem, regras,
                       "Sim" if f.required else "Não", "Sim" if f.readonly else "Não", subtipo, f.note or ""]
                
                full_row_for_search = [t.name] + row_data
                if not q or any(q in (c or "").lower() for c in full_row_for_search):
                    task_rows.append(row_data)
            
            tasks_with_rows.append((t, task_rows))

        css = '''
        <style>
          :root { color-scheme: dark; }
          html, body { background:#0f1012; color:#e6e6e6; font-family:Segoe UI, Arial, sans-serif; margin:0; min-height:100%; }
          body { display:flex; flex-direction:column; }
          .toolbar { position:sticky; top:0; background:#121416; padding:8px 12px; z-index:2; border-bottom:1px solid #1e2126; }
          .wrap { flex:1; padding: 8px 12px 16px; box-sizing:border-box; }
          .table-container { overflow:auto; border:1px solid #1e2126; border-radius:8px; background:#0f1012; }
          table { width:100%; border-collapse:separate; border-spacing:0; table-layout:fixed; }
          thead th { position:sticky; top:0; background:#1b1d22; color:#cbd5e1; text-align:left; padding:10px; font-weight:600; border-bottom:1px solid #2a2f37; }
          tbody td { padding:10px; border-bottom:1px solid #1e2126; vertical-align:top; word-wrap:break-word; overflow-wrap:break-word; }
          tbody tr:nth-child(even) td { background:#15171b; }
          .phase { background:#0f172a; color:#dbeafe; padding:6px 10px; border-radius:999px; font-weight:600; display:inline-block; text-decoration:none; cursor: pointer; }
          .phase:hover { filter:brightness(1.1); }
          .center { text-align:center; }
          .w1{width:15%}.w2{width:18%}.w3{width:10%}.w4{width:16%}.w5{width:16%}.w6{width:6%}.w7{width:6%}.w8{width:13%}.w9{width:16%}
          .muted { color:#9aa4b2; }
          a { color:#93c5fd; text-decoration:none; }
          a:hover { text-decoration:underline; }
        </style>
        '''

        header = '''
        <div class="wrap">
          <div class="table-container">
            <table>
              <colgroup>
                <col class="w1"><col class="w2"><col class="w3"><col class="w4"><col class="w5"><col class="w6"><col class="w7"><col class="w8"><col class="w9">
              </colgroup>
              <thead>
                <tr>
                  <th>Fase/Tarefa</th>
                  <th>Campo</th>
                  <th>Tipo</th>
                  <th>Origem</th>
                  <th>Regras (quando aparece)</th>
                  <th>Obrig.</th>
                  <th>Só leit.</th>
                  <th>Subtipo/Opções</th>
                  <th>Obs.</th>
                </tr>
              </thead>
              <tbody>
        '''

        rows_html = []
        total_records = 0
        for task, rows in tasks_with_rows:
            tname = task.name; tid = task.id
            
            if not rows and not (q and q in tname.lower()):
                 if not q and not rows: pass
                 else: continue
            
            total_records += len(rows)

            if not rows:
                rows_html.append(f'''<tr><td>{tname}</td><td colspan="8" class="muted">(Nenhum campo nesta tarefa)</td></tr>''')
                continue

            enc_id = quote(tid, safe="")
            if tid in self._html_overview_collapsed:
                rows_html.append(f'''<tr><td><a class="phase" href="app://toggle?t={enc_id}&a=expand">▶ {tname}</a></td><td colspan="8" class="muted">({len(rows)} campos)</td></tr>''')
                continue

            rowspan = len(rows); first = rows[0]
            rows_html.append(f'''
              <tr>
                <td rowspan="{rowspan}"><a class="phase" href="app://toggle?t={enc_id}&a=collapse">▼ {tname}</a></td>
                <td>{first[0]}</td><td>{first[1]}</td><td>{first[2]}</td><td>{first[3]}</td>
                <td class="center">{first[4]}</td><td class="center">{first[5]}</td>
                <td>{first[6]}</td><td>{first[7]}</td>
              </tr>
            ''')
            for r in rows[1:]:
                rows_html.append(f'''
                  <tr>
                    <td>{r[0]}</td><td>{r[1]}</td><td>{r[2]}</td><td>{r[3]}</td>
                    <td class="center">{r[4]}</td><td class="center">{r[5]}</td>
                    <td>{r[6]}</td><td>{r[7]}</td>
                  </tr>
                ''')

        footer = '</tbody></table></div></div>'
        tools = f'''
        <div class="toolbar">
          <b>Fluxo:</b> {self.project.flow_name or '-'} &nbsp;&nbsp;|&nbsp;&nbsp;
          <b>Objeto:</b> {self.project.object_type or '-'} &nbsp;&nbsp;|&nbsp;&nbsp;
          <span class="muted">Registros: {total_records}</span>
        </div>
        '''
        return "<html><head><meta charset='utf-8'>" + css + "</head><body>" + tools + header + "".join(rows_html) + footer + "</body></html>"
    
    def open_overview_html(self):
        if HtmlFrame is None:
            messagebox.showwarning("Planilha (HTML)", "Dependência ausente: tkinterweb.\n\nUse o .bat fornecido ou rode: pip install tkinterweb")
            return

        win = ctk.CTkToplevel(self); win.title("Visão geral — Planilha (HTML)")
        self._center_toplevel(win, 1200, 720, transient=False)
        win.resizable(True, True) # CORREÇÃO: Permite que a janela seja maximizada
        win.grab_set()

        win.grid_columnconfigure(0, weight=1); win.grid_rowconfigure(2, weight=1)

        top = ctk.CTkFrame(win, fg_color="transparent"); top.grid(row=0, column=0, sticky="ew", padx=10, pady=(10, 4))
        ctk.CTkLabel(top, text="Buscar:").pack(side="left")
        e_search = ctk.CTkEntry(top, width=480); e_search.pack(side="left", padx=(6, 8))

        btns = ctk.CTkFrame(win, fg_color="transparent"); btns.grid(row=1, column=0, sticky="ew", padx=10, pady=(0, 8))
        ctk.CTkButton(btns, text="Exportar visão (XLSX)", width=160,
                      command=lambda: self._export_overview_xlsx(e_search.get(), self._html_overview_collapsed)).pack(side="right")
        ctk.CTkButton(btns, text="Expandir todas", width=140,
                      command=lambda: (self._html_overview_collapsed.clear(), render())).pack(side="left")
        ctk.CTkButton(btns, text="Colapsar todas", width=140,
                      command=lambda: (self._html_overview_collapsed.update([t.id for t in self.project.tasks]), render())).pack(side="left", padx=(8,0))

        frame = ctk.CTkFrame(win, fg_color="transparent"); frame.grid(row=2, column=0, sticky="nsew", padx=10, pady=(0,10))
        frame.grid_columnconfigure(0, weight=1); frame.grid_rowconfigure(0, weight=1)
        viewer = HtmlFrame(frame, messages_enabled=False); viewer.grid(row=0, column=0, sticky="nsew")

        def _apply_dark_backdrop(widget: tk.Misc, color: str) -> None:
            """Remove artefatos claros ajustando o fundo do HtmlFrame e filhos."""
            _safe_configure(
                widget,
                background=color,
                bg=color,
                highlightbackground=color,
                highlightcolor=color,
                highlightthickness=0,
                borderwidth=0,
                relief="flat",
            )
            for child in widget.winfo_children():
                _apply_dark_backdrop(child, color)

        _apply_dark_backdrop(viewer, DARK_BG2)

        def render():
            html = self._build_overview_html(e_search.get(), self._html_overview_collapsed)
            try: viewer.load_html(html)
            except Exception as ex: messagebox.showerror("Planilha (HTML)", f"Falha ao renderizar HTML.\n\n{ex}", parent=win)
        
        def _get_hovered_toggle_info():
            try:
                element = viewer.get_currently_hovered_element()
                if not isinstance(element, dict): return None
                tag = element.get("tag"); attrs = element.get("attrs", {}); href = attrs.get("href")
                if tag != "a" or not href or not href.startswith("app://toggle"): return None
                url = href; qs = {}
                if "?" in url:
                    for part in url.split("?", 1)[1].split("&"):
                        if "=" in part: k, v = part.split("=", 1); qs[k] = v
                t_id = unquote(qs.get("t", "")); action = qs.get("a", "")
                if not t_id: return None
                if action == "expand": self._html_overview_collapsed.discard(t_id)
                elif action == "collapse": self._html_overview_collapsed.add(t_id)
                else:
                    if t_id in self._html_overview_collapsed: self._html_overview_collapsed.discard(t_id)
                    else: self._html_overview_collapsed.add(t_id)
                render()
                return {"task_id": t_id, "action": action}
            except Exception: return None

        def handle_click(event):
            info = _get_hovered_toggle_info()
            if not info: return
            t_id = info["task_id"]; action = info["action"]
            if action == "expand": self._html_overview_collapsed.discard(t_id)
            elif action == "collapse": self._html_overview_collapsed.add(t_id)
            else:
                if t_id in self._html_overview_collapsed: self._html_overview_collapsed.discard(t_id)
                else: self._html_overview_collapsed.add(t_id)
            render(); return "break"

        def handle_release(event):
            if _get_hovered_toggle_info(): return "break"

        viewer.bind("<Button-1>", handle_click)
        viewer.bind("<ButtonRelease-1>", handle_release)

        e_search.bind("<KeyRelease>", lambda *_: render())
        render()

    # ===== Exportação da visão geral para XLSX =====
    def _export_overview_xlsx(self, q: str, collapsed: Set[str]):
        try:
            import openpyxl
            from openpyxl.styles import Font, Alignment
            from openpyxl.utils import get_column_letter
        except Exception:
            messagebox.showwarning("Exportar XLSX", "Instale openpyxl. Ex.: py -m pip install openpyxl"); return
        initial = self._overview_xlsx_initial_name()
        path = filedialog.asksaveasfilename(defaultextension=".xlsx", filetypes=[("Excel","*.xlsx")], initialfile=initial)
        if not path: return

        wb = openpyxl.Workbook(); ws = wb.active; ws.title = "Visão geral"
        headers = ["ID Tarefa", "Fase/Tarefa","Campo","Tipo","Origem","Regras (quando aparece)","Obrigatório","Só leitura","Subtipo/Opções","Observações"]
        ws.append(headers)
        bold = Font(bold=True); align = Alignment(horizontal="center", vertical="center", wrap_text=True)
        for col, h in enumerate(headers, start=1):
            c = ws.cell(row=1, column=col, value=h); c.font = bold; c.alignment = align

        query = (q or "").lower()
        r = 2
        for idx, t in enumerate(self.project.tasks, start=1):
            task_rows = []
            for f in t.fields:
                origem = f"{self._get_task_name(f.origin_task)} › {self._get_field_name(f.origin_field)}" if (f.origin_task and f.origin_field) else ""
                regras = self._cond_summary_for_task(t, f)
                subtipo = self._format_field_subtype(f)
                row = [t.name, f.name, f.ftype, origem, regras, "Sim" if f.required else "Não", "Sim" if f.readonly else "Não", subtipo, f.note or ""]
                
                if not query or any(query in (c or "").lower() for c in row):
                    task_rows.append(row[1:])

            # Lógica similar à da renderização HTML
            if not task_rows and not (query and query in t.name.lower()):
                 if not query and not task_rows: pass
                 else: continue

            if not task_rows:
                ws.append([idx, t.name, "(Nenhum campo nesta tarefa)"] + [""]*7); r += 1; continue
            
            if t.id in self._html_overview_collapsed:
                ws.append([idx, t.name, f"({len(task_rows)} campos)"] + [""]*7); r += 1; continue
            
            start = r
            for i, rowvals in enumerate(task_rows):
                if i == 0: ws.append([idx, t.name] + rowvals)
                else: ws.append(["", ""] + rowvals)
                r += 1

            if len(task_rows) > 1:
                ws.merge_cells(start_row=start, start_column=1, end_row=r-1, end_column=1)
                ws.merge_cells(start_row=start, start_column=2, end_row=r-1, end_column=2)
            
            ws.cell(row=start, column=1, value=idx).alignment = Alignment(vertical="center", horizontal="left", wrap_text=True)
            ws.cell(row=start, column=2, value=t.name).alignment = Alignment(vertical="center", horizontal="left", wrap_text=True)

        widths = [15, 26, 28, 16, 30, 36, 14, 14, 28, 30]
        for i, w in enumerate(widths, start=1):
            ws.column_dimensions[get_column_letter(i)].width = w
        ws.freeze_panes = "A2"
        wb.save(path); messagebox.showinfo("Exportar", "Visão geral exportada (XLSX).")

    # ===== Objetos / Esquema =====
    def open_flow_object_type_dialog(self) -> bool:
        win = ctk.CTkToplevel(self); win.title("Tipo do objeto do fluxo")
        win.transient(self) # Fix para múltiplos monitores
        self._center_toplevel(win, 520, 190)
        win.grab_set()

        ctk.CTkLabel(win, text="Nome do Objeto deste fluxo (ex.: Pagamento, Obrigação de Fazer):").pack(anchor="w", padx=10, pady=(12, 6))
        ent = ctk.CTkEntry(win, width=420); ent.pack(padx=10); ent.insert(0, self.project.object_type or "")
        ok_pressed = {"v": False}
        def ok():
            name = (ent.get() or "").strip()
            if not name: messagebox.showwarning("Objeto", "Informe um nome.", parent=win); return
            self._push_undo()
            self.project.object_type = name
            for t in self.project.tasks:
                for f in t.fields:
                    if f.ftype == "Objeto":
                        f.obj_type = name; f.name = name; f.name_lock_reason="objeto"; f.name_locked=True
            self._rebuild_metadata_cache()
            self._refresh_flow_label(); self._refresh_rows(); ok_pressed["v"] = True; win.destroy()
        ctk.CTkButton(win, text="OK", width=120, command=ok).pack(pady=12)
        self.wait_window(win); return ok_pressed["v"]

    def open_object_schema_editor(self):
        if not self.project.object_type:
            ok = self.open_flow_object_type_dialog()
            if not ok: return
        win = ctk.CTkToplevel(self); win.title(f"Esquema do objeto — {self.project.object_type}")
        win.transient(self) # Fix para múltiplos monitores
        self._center_toplevel(win, 820, 560)
        win.grab_set()

        win.grid_columnconfigure(0, weight=1); win.grid_rowconfigure(1, weight=1)
        ctk.CTkLabel(win, text=f"Objeto do fluxo: {self.project.object_type}", anchor="w").grid(row=0, column=0, sticky="ew", padx=12, pady=(12, 6))
        listfrm = ctk.CTkScrollableFrame(win); listfrm.grid(row=1, column=0, sticky="nsew", padx=12, pady=(0, 10))

        def render():
            for w in listfrm.winfo_children(): w.destroy()
            for ofd in self.project.object_schema:
                row = ctk.CTkFrame(listfrm, fg_color="#1e1e1e", corner_radius=6); row.pack(fill="x", pady=2)
                extra = f" · {ofd.options}" if (ofd.ftype in LIST_FIELD_TYPES and ofd.options) else ""
                left = ctk.CTkLabel(row, text=f"{ofd.name}  ·  {ofd.ftype}{extra}", anchor="w")
                left.pack(side="left", padx=8, pady=6)
                flags = []
                if ofd.required: flags.append("Obrig.")
                if ofd.readonly: flags.append("Só leit.")
                if flags: ctk.CTkLabel(row, text=" | ".join(flags), anchor="w").pack(side="left", padx=6)
                ctk.CTkButton(row, text="Editar", width=80, command=lambda ref=ofd: edit(ref)).pack(side="right", padx=4)
                ctk.CTkButton(row, text="x", width=34, command=lambda ref=ofd: remove(ref)).pack(side="right")

        def add(): edit(None)

        def edit(ref: Optional[ObjectFieldDef]):
            dlg = ctk.CTkToplevel(win); dlg.title("Campo do objeto")
            dlg.transient(win) # Fix para múltiplos monitores
            self._center_toplevel(dlg, 640, 360)
            dlg.grab_set()

            dlg.grid_columnconfigure(0, weight=1)
            frm = ctk.CTkFrame(dlg, fg_color="transparent"); frm.grid(row=0, column=0, sticky="nsew", padx=12, pady=(12, 6))
            frm.grid_columnconfigure(0, weight=1); frm.grid_columnconfigure(1, weight=1)
            colL = ctk.CTkFrame(frm, fg_color="transparent"); colL.grid(row=0, column=0, sticky="nsew", padx=(0, 8))
            colR = ctk.CTkFrame(frm, fg_color="transparent"); colR.grid(row=0, column=1, sticky="nsew", padx=(8, 0))

            ctk.CTkLabel(colL, text="Nome:").pack(anchor="w", pady=(0, 4))
            e_name = ctk.CTkEntry(colL); e_name.pack(fill="x"); e_name.insert(0, ref.name if ref else "")

            ctk.CTkLabel(colL, text="Tipo:").pack(anchor="w", pady=(12, 4))
            cmb = ctk.CTkComboBox(
                colL,
                values=[
                    "Texto",
                    "Área de texto",
                    "Numérico",
                    "Lista",
                    "Lista Vários",
                    "Data",
                    "Informativo",
                    "Anexo",
                    "Valores",
                    "Componente do sistema",
                ],
            )
            cmb.pack(fill="x"); cmb.set(ref.ftype if ref else "Texto")

            ctk.CTkLabel(colL, text="Opções (para Lista/Lista Vários)").pack(anchor="w", pady=(12, 4))
            e_opt = ctk.CTkEntry(colL); e_opt.pack(fill="x"); e_opt.insert(0, ref.options if ref else "")

            flags = ctk.CTkFrame(colR, fg_color="transparent"); flags.pack(fill="x", pady=(0,0))
            v_req = tk.BooleanVar(value=(ref.required if ref else False))
            v_ro  = tk.BooleanVar(value=(ref.readonly if ref else True))
            ctk.CTkCheckBox(flags, text="Obrigatório", variable=v_req).pack(side="left")
            ctk.CTkCheckBox(flags, text="Somente leitura", variable=v_ro).pack(side="left", padx=12)

            ctk.CTkLabel(colR, text="Observações:").pack(anchor="w", pady=(12, 4))
            e_note = ctk.CTkEntry(colR); e_note.pack(fill="x"); e_note.insert(0, ref.note if ref else "")

            btns = ctk.CTkFrame(dlg, fg_color="transparent"); btns.grid(row=1, column=0, sticky="ew", padx=12, pady=(10, 12))
            def ok():
                self._push_undo()
                name = e_name.get().strip()
                if not name: messagebox.showwarning("Objeto", "Informe o nome do campo.", parent=dlg); return
                if ref is None:
                    self.project.object_schema.append(ObjectFieldDef(
                        name=name, ftype=cmb.get().strip(), options=e_opt.get().strip(),
                        required=v_req.get(), readonly=v_ro.get(), order=len(self.project.object_schema), note=e_note.get().strip()
                    ))
                else:
                    ref.name=name; ref.ftype=cmb.get().strip(); ref.options=e_opt.get().strip()
                    ref.required=v_req.get(); ref.readonly=v_ro.get(); ref.note=e_note.get().strip()
                dlg.destroy(); render()
            ctk.CTkButton(btns, text="Salvar", width=120, command=ok).pack(side="right", padx=6)
            btn_cancel = ctk.CTkButton(btns, text="Cancelar", width=120, command=dlg.destroy)
            _apply_secondary_style(btn_cancel)
            btn_cancel.pack(side="right")

        def remove(ref: ObjectFieldDef):
            if messagebox.askyesno("Objeto", f"Remover '{ref.name}'?", parent=win):
                self._push_undo()
                self.project.object_schema = [x for x in self.project.object_schema if x is not ref]
                for i, ofd in enumerate(self.project.object_schema): ofd.order = i
                render()

        btns2 = ctk.CTkFrame(win, fg_color="transparent"); btns2.grid(row=2, column=0, sticky="ew", padx=12, pady=(0,10))
        ctk.CTkButton(btns2, text="+ Campo do objeto", width=160, command=add).pack(side="left")
        ctk.CTkButton(btns2, text="Fechar", width=120, command=win.destroy).pack(side="right")
        for i, ofd in enumerate(self.project.object_schema): ofd.order = i
        render()

    # ===== Import/Export esquema =====
    def import_object_schema_xlsx(self):
        try: import openpyxl
        except Exception: messagebox.showwarning("Importar XLSX", "Instale openpyxl. Ex.: py -m pip install openpyxl"); return
        path = filedialog.askopenfilename(filetypes=[("Excel","*.xlsx")])
        if not path: return
        if not self.project.object_type:
            ok = self.open_flow_object_type_dialog()
            if not ok: return
        try:
            wb = openpyxl.load_workbook(path, data_only=True); sh = wb.active
            rows = [[("" if c is None else str(c)) for c in row] for row in sh.iter_rows(values_only=True)]
            if not rows: raise ValueError("Planilha vazia.")
            header = [str(h).strip().lower() for h in rows[0]]; data = rows[1:]
            def idx(name): 
                try: return header.index(name)
                except ValueError: return None
            i_campo=idx("campo"); i_tipo=idx("tipo"); i_opts=idx("opções") if "opções" in header else idx("opcoes")
            i_obr=idx("obrigatório") if "obrigatório" in header else idx("obrigatorio")
            i_sol=idx("só leitura") if "só leitura" in header else idx("so leitura")
            i_obs=idx("observações") if "observações" in header else idx("observacoes")
            new_schema=[]
            for r in data:
                nome = (str(r[i_campo]).strip() if i_campo is not None and i_campo < len(r) else "")
                if not nome: continue
                ftype = (str(r[i_tipo]).strip() if i_tipo is not None and i_tipo < len(r) else "Texto")
                opts = (str(r[i_opts]).strip() if i_opts is not None and i_opts < len(r) else "")
                obr = ((str(r[i_obr]).strip().lower().startswith("s")) if i_obr is not None and i_obr < len(r) else False)
                sol = ((str(r[i_sol]).strip().lower().startswith("s")) if i_sol is not None and i_sol < len(r) else True)
                obs = (str(r[i_obs]).strip() if i_obs is not None and i_obs < len(r) else "")
                new_schema.append(ObjectFieldDef(name=nome, ftype=ftype, options=opts, required=obr, readonly=sol, note=obs))
            for i, ofd in enumerate(new_schema): ofd.order = i
            self._push_undo()
            self.project.object_schema = new_schema
            self._rebuild_metadata_cache()
            messagebox.showinfo("Importar", "Esquema do Objeto importado de XLSX.")
        except Exception as e: messagebox.showerror("Importar XLSX", f"Falha ao importar.\n\n{e}")

    def export_object_schema_xlsx(self):
        if not self.project.object_schema: messagebox.showinfo("Exportar", "Não há campos no esquema do Objeto."); return
        try:
            import openpyxl
            from openpyxl.styles import Font, Alignment
            from openpyxl.utils import get_column_letter
        except Exception: messagebox.showwarning("Exportar XLSX", "Instale openpyxl. Ex.: py -m pip install openpyxl"); return
        path = filedialog.asksaveasfilename(defaultextension=".xlsx", filetypes=[("Excel","*.xlsx")], initialfile=f"esquema_{self.project.object_type or 'objeto'}.xlsx")
        if not path: return
        wb = openpyxl.Workbook(); ws = wb.active; ws.title = "Esquema"
        headers = ["Tipo de Objeto","Ordem","Campo","Tipo","Opções","Obrigatório","Só leitura","Observações"]
        ws.append(headers)
        bold = Font(bold=True); align = Alignment(horizontal="center", vertical="center", wrap_text=True)
        for col, h in enumerate(headers, start=1):
            c = ws.cell(row=1, column=col, value=h); c.font = bold; c.alignment = align
        for i, ofd in enumerate(self.project.object_schema):
            ofd.order = i
            ws.append([self.project.object_type, ofd.order, ofd.name, ofd.ftype, ofd.options,
                       "Sim" if ofd.required else "Não", "Sim" if ofd.readonly else "Não", ofd.note])
        for col, w in enumerate([22,10,28,18,30,14,14,36], start=1): ws.column_dimensions[openpyxl.utils.get_column_letter(col)].width = w
        wb.save(path); messagebox.showinfo("Exportar", "Esquema exportado (XLSX).")

    def download_object_schema_template_xlsx(self):
        try:
            import openpyxl
            from openpyxl.worksheet.datavalidation import DataValidation
            from openpyxl.styles import Font, Alignment, PatternFill
            from openpyxl.utils import get_column_letter
        except Exception: messagebox.showwarning("Modelo XLSX", "Instale openpyxl. Ex.: py -m pip install openpyxl"); return
        path = filedialog.asksaveasfilename(defaultextension=".xlsx", filetypes=[("Excel","*.xlsx")], initialfile="modelo_esquema_objeto.xlsx")
        if not path: return
        wb = openpyxl.Workbook(); ws = wb.active; ws.title = "Esquema de Objeto"
        headers = ["Campo", "Tipo", "Opções", "Obrigatório", "Só leitura", "Observações"]
        ws.append(headers)
        example = ["Nome do Cliente", "Texto", "", "Sim", "Não", "Nome completo do cliente, sem abreviações."]
        ws.append(example)
        help_row = ["Nome do campo (ex: Valor da Causa)", "Texto, Área de texto, Lista, Data, etc.", "Para tipo 'Lista', separe as opções com ponto e vírgula (ex: Opção A;Opção B)", "Sim/Não", "Sim/Não", "Instruções ou comentários sobre o campo."]
        ws.append(help_row)

        header_fill = PatternFill("solid", fgColor="1f2937"); header_font = Font(bold=True, color="FFFFFF")
        for col, h in enumerate(headers, start=1):
            c = ws.cell(row=1, column=col, value=h); c.fill = header_fill; c.font = header_font; c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        
        widths = [30, 20, 30, 15, 15, 40]
        for i, w in enumerate(widths, start=1):
            ws.column_dimensions[get_column_letter(i)].width = w
        ws.freeze_panes = "A2"
        dv_bin = DataValidation(type="list", formula1='"Sim,Não"', allow_blank=True, showDropDown=True); ws.add_data_validation(dv_bin)
        dv_bin.add("D2:E1048576")
        tipos = [
            "Texto",
            "Área de texto",
            "Numérico",
            "Lista",
            "Lista Vários",
            "Data",
            "Informativo",
            "Anexo",
            "Valores",
            "Componente do sistema",
        ]
        dv_tipo = DataValidation(type="list", formula1=f'"{",".join(tipos)}"', allow_blank=True, showDropDown=True); ws.add_data_validation(dv_tipo)
        dv_tipo.add("B2:B1048576")
        for r in (2,3):
            for c in range(1, len(headers)+1): ws.cell(row=r, column=c).alignment = Alignment(vertical="top", wrap_text=True)

        wb.save(path); messagebox.showinfo("Modelo", "Modelo de esquema de objeto salvo (XLSX).")

    # ===== Arquivo / Fluxo =====
    def rename_flow(self):
        new = self._prompt_text("Renomear fluxo", "Nome do fluxo:", self.project.flow_name)
        if new is not None:
            self._push_undo()
            self.project.flow_name = new; self._refresh_flow_label()

    def new_flow_blank(self, show_message: bool = True):
        if (self.project.tasks or self.project.object_type or show_message):
            if not messagebox.askyesno("Novo vazio", "Criar novo fluxo em branco? (o atual não salvo será perdido)"):
                return
        if self.project.tasks or self.project.object_type:
            self._push_undo()
        self.project = ProjectModel(flow_name="Novo fluxo", tasks=[], object_schema=[], object_type="")
        self.current_task_id = None; self.selected_field_ids.clear()
        self._rebuild_metadata_cache()
        self._refresh_flow_label(); self._refresh_task_combo(); self._refresh_rows()
        if show_message: messagebox.showinfo("Novo", "Fluxo em branco criado. Use 'Tarefas...' para adicionar tarefas.")

    def save_project(self):
        path = filedialog.asksaveasfilename(defaultextension=".json", filetypes=[("Projeto","*.json")])
        if not path: return
        with open(path, "w", encoding="utf-8") as f: json.dump(self.project.to_dict(), f, ensure_ascii=False, indent=2)
        messagebox.showinfo("Salvar", "Projeto salvo.")

    def open_project(self):
        path = filedialog.askopenfilename(filetypes=[("Projeto","*.json")])
        if not path: return
        try:
            with open(path, "r", encoding="utf-8") as f:
                raw_data = f.read()
        except (OSError, UnicodeDecodeError) as exc:
            messagebox.showerror("Abrir projeto", f"Não foi possível ler o arquivo selecionado.\n\n{exc}")
            return

        try:
            data = json.loads(raw_data)
        except json.JSONDecodeError as exc:
            messagebox.showerror("Abrir projeto", f"O arquivo não é um JSON válido.\n\n{exc}")
            return

        if not isinstance(data, dict):
            messagebox.showerror("Abrir projeto", "O arquivo selecionado não contém um projeto válido.")
            return

        try:
            if self._apply_project_dict(data):
                self._undo_stack.clear()
                self._redo_stack.clear()
        except Exception as exc:
            messagebox.showerror("Abrir projeto", f"Falha ao carregar o projeto.\n\n{exc}")

    # ===== XLSX: exportar/importar fluxo =====
    def export_flow_to_xlsx(self):
        try:
            import openpyxl
            from openpyxl.styles import Font, Alignment
            from openpyxl.utils import get_column_letter
        except Exception: messagebox.showwarning("Exportar XLSX", "Instale openpyxl. Ex.: py -m pip install openpyxl"); return
        path = filedialog.asksaveasfilename(defaultextension=".xlsx", filetypes=[("Excel","*.xlsx")], initialfile=f"{self.project.flow_name or 'fluxo'}.xlsx")
        if not path: return
        wb = openpyxl.Workbook(); ws = wb.active; ws.title = "Fluxo"
        headers = ["Fluxo","ID Tarefa","Fase","Campo","Obrigatório","Somente Leitura","Sub Tipo","Observações","Origem (Tarefa › Campo)","Aparece quando"]
        ws.append(headers)
        bold = Font(bold=True); align = Alignment(horizontal="center", vertical="center", wrap_text=True)
        for col, h in enumerate(headers, start=1):
            c = ws.cell(row=1, column=col, value=h); c.font = bold; c.alignment = align
        for idx, t in enumerate(self.project.tasks, start=1):
            for fld in t.fields:
                origem=""
                if fld.origin_task and fld.origin_field:
                    origem=f"{self._get_task_name(fld.origin_task)} › {self._get_field_name(fld.origin_field)}"
                conds=[]
                for c in fld.cond:
                    base=self._get_field_name(c.src_field) or f"[id {c.src_field}]"
                    conds.append(f"{base} {c.op} {c.value!r}")
                subtipo=fld.ftype
                if fld.ftype=="Lista" and fld.options: subtipo=f"Lista: {fld.options}"
                if fld.ftype=="Informativo" and fld.options: subtipo=f"Informativo: {fld.options}"
                if fld.ftype=="Objeto": subtipo=f"Objeto: {self.project.object_type or '-'}"
                ws.append([self.project.flow_name, idx, t.name, fld.name,
                           "Sim" if fld.required else "Não", "Sim" if fld.readonly else "Não",
                           subtipo, fld.note or "", origem, " OU ".join(conds)])
        widths = [18,12,36,38,14,16,40,42,40,42]
        for i, w in enumerate(widths, start=1): ws.column_dimensions[get_column_letter(i)].width = w
        ws.freeze_panes = "A2"
        wb.save(path); messagebox.showinfo("Exportar", "Fluxo exportado (XLSX).")

    def import_flow_from_xlsx(self):
        try: import openpyxl
        except Exception: messagebox.showwarning("Importar XLSX", "Instale openpyxl. Ex.: py -m pip install openpyxl"); return
        path = filedialog.askopenfilename(filetypes=[("Excel","*.xlsx")])
        if not path: return
        try:
            wb = openpyxl.load_workbook(path, data_only=True); sh = wb.active
            rows = [[("" if c is None else str(c)) for c in row] for row in sh.iter_rows(values_only=True)]
            if not rows: raise ValueError("Planilha vazia.")
            header = [h.strip().lower() for h in rows[0]]; data = rows[1:]
            def idx(name:str) -> Optional[int]:
                try: return header.index(name)
                except ValueError: return None
            i_fluxo = idx("fluxo"); i_fase = idx("fase") if "fase" in header else idx("fase/tarefa")
            i_campo = idx("campo"); i_obr = idx("obrigatório") if "obrigatório" in header else idx("obrigatorio")
            i_sol = idx("somente leitura") if "somente leitura" in header else (idx("só leitura") if "só leitura" in header else None)
            i_sub = idx("sub tipo"); i_obs = idx("observações") if "observações" in header else idx("observacoes")
            i_org = idx("origem (tarefa › campo)") or idx("origem (tarefa > campo)")
            i_apq = idx("aparece quando")
            if i_fase is None or i_campo is None: raise ValueError("A planilha precisa ter ao menos 'Fase' e 'Campo'.")
            
            proj = ProjectModel()
            if i_fluxo is not None:
                first_name = next((str(r[i_fluxo]).strip() for r in data if i_fluxo < len(r) and str(r[i_fluxo]).strip()), "")
                proj.flow_name = first_name or "Fluxo importado"
            
            tasks_by_name: Dict[str, Task] = {}; pending_origins: List[Tuple[Field, str, str]] = []
            import re as _re
            for r in data:
                fase = (str(r[i_fase]).strip() if i_fase is not None and i_fase < len(r) else "")
                if not fase: continue
                if fase not in tasks_by_name: tasks_by_name[fase] = Task(id=_uid(), name=fase, fields=[])
                t = tasks_by_name[fase]
                campo = (str(r[i_campo]).strip() if i_campo is not None and i_campo < len(r) else "")
                if not campo: continue
                raw = (str(r[i_sub]).strip() if i_sub is not None and i_sub < len(r) else "")
                ftype="Texto"; options=""
                if raw:
                    low = raw.lower()
                    if low.startswith("lista"): ftype="Lista"; parts=raw.split(":",1); options=parts[1].strip() if len(parts)==2 else ""
                    elif low.startswith("informativo"): ftype="Informativo"; parts=raw.split(":",1); options=parts[1].strip() if len(parts)==2 else ""
                    elif low.startswith("área de texto") or low.startswith("area de texto"): ftype="Área de texto"
                    elif low.startswith("data"): ftype="Data"
                    elif low.startswith("anexo"): ftype="Anexo"
                    elif low.startswith("valores"): ftype="Valores"
                    elif low.startswith("componente do sistema"): ftype="Componente do sistema"
                    elif low.startswith("objeto"): ftype="Objeto"
                    elif low in ["texto","text"]: ftype="Texto"
                required = ((str(r[i_obr]).strip().lower().startswith("s")) if i_obr is not None and i_obr < len(r) else False)
                readonly = ((str(r[i_sol]).strip().lower().startswith("s")) if i_sol is not None and i_sol < len(r) else False)
                obs = (str(r[i_obs]).strip() if i_obs is not None and i_obs < len(r) else "")
                f = Field(id=_uid(), name=campo, ftype=ftype, required=required, readonly=readonly, options=options, note=obs); t.fields.append(f)
                if i_org is not None and i_org < len(r):
                    ref = str(r[i_org]).strip()
                    if ref:
                        parts = _re.split(r"[›>]", ref)
                        if len(parts) >= 2: pending_origins.append((f, parts[0].strip(), parts[1].strip()))
                if i_apq is not None and i_apq < len(r):
                    when = str(r[i_apq]).strip()
                    if when:
                        separators = _re.split(r"\s+(?:AND|E|OU|OR)\s+", when)
                        for clause in separators:
                            m = _re.match(r"\s*(.+?)\s*==\s*'(.*)'\s*$", clause)
                            if m:
                                base_name = m.group(1).strip(); value = m.group(2)
                                base_field = next((fld for fld in t.fields if fld.name == base_name), None)
                                if base_field: f.cond.append(Condition(src_field=base_field.id, op="==", value=value))
            proj.tasks = list(tasks_by_name.values())
            
            # Rebuild cache for origin mapping
            temp_app = App(); temp_app.project = proj; temp_app._rebuild_metadata_cache()
            
            for f, tname, fname in pending_origins:
                src_t = next((tx for tx in proj.tasks if tx.name == tname), None)
                if not src_t: continue
                src_f = next((fx for fx in src_t.fields if fx.name == fname), None)
                if not src_f: continue
                f.origin_task = src_t.id; f.origin_field = src_f.id
                f.name_before_origin = f.name; f.name = src_f.name
                f.name_lock_reason = "origem"; f.name_locked = True
                if f.ftype not in ("Informativo","Componente do sistema","Objeto"): f.readonly = True
            
            res = messagebox.askyesno("Importar fluxo", "Fluxo importado do Excel.\n\nAplicar agora substituindo o atual?\n\nSim = Aplicar | Não = Salvar como template")
            if res:
                self._push_undo()
                self._apply_project_dict(proj.to_dict())
                messagebox.showinfo("Importar", "Fluxo aplicado.")
            else:
                name = self._prompt_text("Salvar como template", "Nome do template:", proj.flow_name or "Fluxo importado")
                if name is not None:
                    exists = next((x for x in self.store.list_all() if x.get("name","") == name), None)
                    if exists:
                        rep = messagebox.askyesno("Templates", f"Já existe '{name}'. Substituir?\n\nSim = Substituir | Não = Criar cópia")
                        if rep: self.store.save_template(name, proj, replace=True)
                        else:   self.store.save_template(name, proj, replace=False, create_copy_if_exists=True)
                    else:
                        self.store.save_template(name, proj, replace=False)
                    messagebox.showinfo("Templates", "Template salvo.")
        except Exception as e:
            messagebox.showerror("Importar XLSX", f"Falha ao importar.\n\n{e}")

    def download_flow_template_xlsx(self):
        try:
            import openpyxl
            from openpyxl.worksheet.datavalidation import DataValidation
            from openpyxl.styles import Font, Alignment, PatternFill
            from openpyxl.utils import get_column_letter
        except Exception: messagebox.showwarning("Modelo XLSX", "Instale openpyxl. Ex.: py -m pip install openpyxl"); return
        path = filedialog.asksaveasfilename(defaultextension=".xlsx", filetypes=[("Excel","*.xlsx")], initialfile="modelo_fluxo.xlsx")
        if not path: return
        wb = openpyxl.Workbook(); ws = wb.active; ws.title = "Fluxo"
        headers = ["Fluxo","ID Tarefa","Fase","Campo","Obrigatório","Somente Leitura","Sub Tipo","Observações","Origem (Tarefa › Campo)","Aparece quando"]
        ws.append(headers)
        example = ["Cadastro","1","Aceitação de Patrocínio","Patrocínio aceito?","Sim","Não","Lista: Sim;Não","","",""]
        ws.append(example)
        help_row = ["Use o mesmo nome de fluxo", "Inteiro crescente por fase (1,2,3…)", "Nome da Fase/Tarefa", "Nome do campo", "Sim/Não", "Sim/Não", "Tipo; para Lista/Informativo inclua após ':' as opções ou o texto", "Observações livres", "Ex.: Aceitação de Patrocínio › Patrocínio aceito?", "Ex.: Campo == 'Sim' OU Outro == 'X'"]
        ws.append(help_row)
        header_fill = PatternFill("solid", fgColor="1f2937"); header_font = Font(bold=True, color="FFFFFF")
        for col, h in enumerate(headers, start=1):
            c = ws.cell(row=1, column=col, value=h); c.fill = header_fill; c.font = header_font; c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        widths = [18,12,36,38,14,16,40,42,40,42]
        for i, w in enumerate(widths, start=1): ws.column_dimensions[get_column_letter(i)].width = w
        ws.freeze_panes = "A2"
        dv_bin = DataValidation(type="list", formula1='"Sim,Não"', allow_blank=True, showDropDown=True); ws.add_data_validation(dv_bin)
        for col in (5,6): dv_bin.add(f"{get_column_letter(col)}2:{get_column_letter(col)}1048576")
        tipos = [
            "Texto",
            "Área de texto",
            "Numérico",
            "Lista: Sim;Não",
            "Lista Vários: Opção A;Opção B",
            "Data",
            "Informativo: Mensagem aqui",
            "Anexo",
            "Valores",
            "Componente do sistema",
            "Objeto",
        ]
        dv_tipo = DataValidation(type="list", formula1=f'"{",".join(tipos)}"', allow_blank=True, showDropDown=True); ws.add_data_validation(dv_tipo)
        dv_tipo.add("G2:G1048576")
        for r in (2,3):
            for c in range(1, len(headers)+1): ws.cell(row=r, column=c).alignment = Alignment(vertical="top", wrap_text=True)
        wb.save(path); messagebox.showinfo("Modelo", "Modelo salvo (XLSX).")

    # ===== Templates (biblioteca) =====
    def open_templates_dialog(self):
        win = ctk.CTkToplevel(self); win.title("Meus templates")
        win.transient(self) # Fix para múltiplos monitores
        self._center_toplevel(win, 720, 580)
        win.grab_set()

        win.grid_columnconfigure(0, weight=1); win.grid_rowconfigure(2, weight=1)

        top = ctk.CTkFrame(win, fg_color="transparent"); top.grid(row=0, column=0, sticky="ew", padx=10, pady=(10, 4))
        ctk.CTkLabel(top, text="Buscar:").pack(side="left")
        e_search = ctk.CTkEntry(top, width=380); e_search.pack(side="left", padx=(6, 0))
        ctk.CTkLabel(top, text=f"Armazenados em: {TEMPLATES_DB_PATH}", anchor="w").pack(side="right")

        listfrm = ctk.CTkScrollableFrame(win); listfrm.grid(row=2, column=0, sticky="nsew", padx=10, pady=(6, 8))
        selected_id = tk.StringVar(value="")

        def render():
            for w in listfrm.winfo_children(): w.destroy()
            q = (e_search.get() or "").lower(); items = self.store.list_all(); shown = 0
            for t in items:
                name = t.get("name","")
                if q and q not in name.lower(): continue
                row = ctk.CTkFrame(listfrm, fg_color="#1e1e1e", corner_radius=6); row.pack(fill="x", pady=4)
                ctk.CTkRadioButton(row, text=name, value=t.get("id",""), variable=selected_id).pack(side="left", padx=8, pady=6)
                meta = f"atualizado: {t.get('updated_at','')}"
                ctk.CTkLabel(row, text=meta).pack(side="left", padx=8); shown += 1
            if shown == 0: ctk.CTkLabel(listfrm, text="Nenhum template encontrado.").pack(pady=10)
        e_search.bind("<KeyRelease>", lambda *_: render()); render()

        btns = ctk.CTkFrame(win, fg_color="transparent"); btns.grid(row=3, column=0, sticky="ew", padx=10, pady=(0,10))
        def require_sel() -> Optional[dict]:
            tid = selected_id.get()
            if not tid: messagebox.showinfo("Templates", "Selecione um template.", parent=win); return None
            t = self.store.get(tid)
            if not t: messagebox.showwarning("Templates", "Template não encontrado.", parent=win)
            return t
        def apply_sel():
            t = require_sel();
            if not t: return
            if not messagebox.askyesno("Aplicar template", f"Substituir o fluxo atual por '{t.get('name','(sem nome)')}'?", parent=win): return
            self._apply_project_dict(t.get("project", {}))
            messagebox.showinfo("Templates", "Template aplicado.", parent=win)
        def rename_sel():
            t = require_sel(); 
            if not t: return
            new = self._prompt_text("Renomear template", "Novo nome:", t.get("name",""))
            if new is not None:
                if not self.store.rename(t.get("id",""), new):
                    messagebox.showwarning("Templates", "Já existe um template com esse nome.", parent=win)
                render()
        def delete_sel():
            t = require_sel();
            if not t: return
            if not messagebox.askyesno("Apagar template", f"Excluir '{t.get('name','')}' da biblioteca?", parent=win): return
            self.store.delete(t.get("id","")); render()
        def export_sel():
            t = require_sel();
            if not t: return
            path = filedialog.asksaveasfilename(defaultextension=".template.json", filetypes=[("Template JSON","*.template.json")], initialfile=f"{t.get('name','template')}.template.json")
            if not path: return
            data = {"name": t.get("name",""), "project": t.get("project", {}), "exported_at": _now_iso(), "version": APP_VERSION}
            with open(path, "w", encoding="utf-8") as f: json.dump(data, f, ensure_ascii=False, indent=2)
            messagebox.showinfo("Templates", "Template exportado.", parent=win)
        def import_template():
            path = filedialog.askopenfilename(filetypes=[("Template JSON","*.template.json"), ("JSON","*.json")])
            if not path: return
            try:
                with open(path, "r", encoding="utf-8") as f: data = json.load(f)
                name = (data.get("name") or "Template importado").strip(); project_dict = data.get("project", {})
                items = self.store.list_all(); exists = next((x for x in items if x.get("name","") == name), None)
                if exists:
                    res = messagebox.askyesno("Templates", f"Já existe '{name}'. Substituir?\n\nSim = Substituir | Não = Criar cópia", parent=win)
                    if res: self.store.save_template(name, ProjectModel.from_dict(project_dict), replace=True)
                    else:   self.store.save_template(name, ProjectModel.from_dict(project_dict), replace=False, create_copy_if_exists=True)
                else:
                    self.store.save_template(name, ProjectModel.from_dict(project_dict), replace=False)
                messagebox.showinfo("Templates", "Template importado.", parent=win); render()
            except Exception as e: messagebox.showerror("Templates", f"Falha ao importar.\n\n{e}", parent=win)
        
        ctk.CTkButton(btns, text="Aplicar (Substituir)", width=160, command=apply_sel).pack(side="left")
        ctk.CTkButton(btns, text="Renomear", width=120, command=rename_sel).pack(side="left", padx=6)
        ctk.CTkButton(btns, text="Excluir", width=110, command=delete_sel).pack(side="left", padx=6)
        ctk.CTkButton(btns, text="Exportar", width=110, command=export_sel).pack(side="left", padx=6)
        ctk.CTkButton(btns, text="Importar", width=110, command=import_template).pack(side="left", padx=6)
        ctk.CTkButton(btns, text="Fechar", width=120, command=win.destroy).pack(side="right")

    def save_flow_as_template(self):
        name = self._prompt_text("Salvar como template", "Nome do template:", self.project.flow_name or "Fluxo")
        if name is not None:
            exists = next((x for x in self.store.list_all() if x.get("name","") == name), None)
            if exists:
                rep = messagebox.askyesno("Templates", f"Já existe '{name}'. Substituir?\n\nSim = Substituir | Não = Criar cópia")
                if rep: self.store.save_template(name, self.project, replace=True)
                else: self.store.save_template(name, self.project, replace=False, create_copy_if_exists=True)
            else:
                self.store.save_template(name, self.project, replace=False)
            messagebox.showinfo("Templates", "Template salvo.")

    # ===== Template embutido: Cadastro =====
    def apply_builtin_template_cadastro_confirm(self):
        if not messagebox.askyesno("Aplicar template", "Substituir o fluxo atual por 'Cadastro'?"): return
        self.apply_builtin_template_cadastro()

    def apply_builtin_template_cadastro(self):
        proj = ProjectModel(flow_name="Cadastro")
        t1 = Task(id=_uid(), name="Aceitação de Patrocínio")
        f1_pa = Field(id=_uid(), name="Patrocínio aceito?", ftype="Lista", required=True, options="Sim;Não")
        f1_pf = Field(id=_uid(), name="Prazo Fatal", ftype="Data", required=True); f1_pf.cond.append(Condition(src_field=f1_pa.id, op="==", value="Sim"))
        f1_ss = Field(id=_uid(), name="Deseja solicitar subsídios?", ftype="Lista", required=True, options="Sim;Não"); f1_ss.cond.append(Condition(src_field=f1_pa.id, op="==", value="Sim"))
        t1.fields = [f1_pa, f1_pf, f1_ss]
        t2 = Task(id=_uid(), name="Indicar novo escritório responsável")
        f2_pa = Field(id=_uid(), name="Patrocínio aceito?", ftype="Lista", readonly=True, origin_task=t1.id, origin_field=f1_pa.id, name_lock_reason="origem", name_locked=True, name_before_origin="Patrocínio aceito?")
        f2_er = Field(id=_uid(), name="Escritório Responsável", ftype="Texto", required=True)
        t2.fields = [f2_pa, f2_er]
        t3 = Task(id=_uid(), name="Complementar cadastro")
        f3_inf = Field(id=_uid(), name="Informativo", ftype="Informativo", readonly=True, options='Para efetivar o cadastro do processo, clicar no botão "Cadastrar" localizado no menu superior horizontal. Somente confirmar esta tarefa após a realização deste procedimento.')
        t3.fields = [f3_inf]
        t4 = Task(id=_uid(), name="Inclusão de valores do processo")
        f4_comp = Field(id=_uid(), name="Componente Instância Pedido", ftype="Componente do sistema", required=True, readonly=False, note="Componente Instância Pedido")
        t4.fields = [f4_comp]
        t5 = Task(id=_uid(), name="Elaborar Peça Pertinente")
        f5_anx = Field(id=_uid(), name="Anexar Peça Pertinente", ftype="Anexo", required=True, note="[Tipo de Doc.: Peça Processual;Defesa;Petição;Outros] Nome do doc: peça pertinente.")
        f5_obs = Field(id=_uid(), name="Observações Peça Pertinente", ftype="Área de texto", required=True)
        t5.fields = [f5_anx, f5_obs]
        t6 = Task(id=_uid(), name="Validação da peça pertinente")
        f6_val = Field(id=_uid(), name="Peça Validada?", ftype="Lista", required=True, options="Sim;Não")
        f6_ref = Field(id=_uid(), name="Anexar Peça Pertinente", ftype="Anexo", readonly=True, origin_task=t5.id, origin_field=f5_anx.id, name_lock_reason="origem", name_locked=True, name_before_origin="Peça Pertinente (referência)")
        f6_obs = Field(id=_uid(), name="Observações Sobre a não Validação da Peça", ftype="Área de texto", required=False); f6_obs.cond.append(Condition(src_field=f6_val.id, op="==", value="Não"))
        t6.fields = [f6_val, f6_ref, f6_obs]
        t7 = Task(id=_uid(), name="Corrigir peça pertinente")
        f7_corr = Field(id=_uid(), name="Peça Pertinente — Corrigida", ftype="Anexo", required=True, note="[Tipo de Doc.: Peça Corrigida] Nome do doc: peça corrigida.")
        f7_ref = Field(id=_uid(), name="Anexar Peça Pertinente", ftype="Anexo", readonly=True, origin_task=t5.id, origin_field=f5_anx.id, name_lock_reason="origem", name_locked=True, name_before_origin="Peça Pertinente (referência)")
        t7.fields = [f7_corr, f7_ref]
        t8 = Task(id=_uid(), name="Realizar protocolo da peça e anexar comprovante")
        f8_proto = Field(id=_uid(), name="Protocolo da Peça", ftype="Anexo", required=True, note="[Tipo de Doc.: Protocolo;Protocolo Realizado] Comprovante do protocolo no tribunal.")
        f8_ref1 = Field(id=_uid(), name="Anexar Peça Pertinente", ftype="Anexo", readonly=True, origin_task=t5.id, origin_field=f5_anx.id, name_lock_reason="origem", name_locked=True, name_before_origin="Peça Pertinente (referência)")
        f8_ref2 = Field(id=_uid(), name="Peça Pertinente — Corrigida", ftype="Anexo", readonly=True, origin_task=t7.id, origin_field=f7_corr.id, name_lock_reason="origem", name_locked=True, name_before_origin="Peça Corrigida (referência)")
        t8.fields = [f8_proto, f8_ref1, f8_ref2]
        t9 = Task(id=_uid(), name="Ciência de peça protocolada")
        f9_ref = Field(id=_uid(), name="Protocolo da Peça", ftype="Anexo", readonly=True, origin_task=t8.id, origin_field=f8_proto.id, name_lock_reason="origem", name_locked=True, name_before_origin="Protocolo Anexado")
        t9.fields = [f9_ref]

        self._push_undo()
        self.project = ProjectModel(flow_name="Cadastro", tasks=[t1,t2,t3,t4,t5,t6,t7,t8,t9], object_type=self.project.object_type, object_schema=self.project.object_schema)
        self.current_task_id = t1.id
        self._rebuild_metadata_cache()
        self._refresh_flow_label(); self._build_header(initial=True); self._refresh_task_combo(); self._refresh_rows()
        messagebox.showinfo("Templates", "Template 'Cadastro' aplicado.")

    # ===== Origem/Regras/Simulador =====
    def open_origin_picker(self, f: Field):
        self._commit_row_data(f)
        if f.ftype == "Objeto":
            messagebox.showwarning("Origem", "Campos do tipo Objeto não suportam origem."); return
        win = ctk.CTkToplevel(self); win.title("Origem do valor")
        win.transient(self) # Fix para múltiplos monitores
        self._center_toplevel(win, 520, 420)
        win.grab_set()

        ctk.CTkLabel(win, text="Tarefa:").pack(anchor="w", padx=10, pady=(10, 4))
        cmb_t = ctk.CTkComboBox(win, values=[t.name for t in self.project.tasks], width=420); cmb_t.pack(padx=10, pady=(0, 8))
        
        current_task_name = self._get_task_name(f.origin_task) if f.origin_task else (self._get_task().name if self.project.tasks else "")
        cmb_t.set(current_task_name)
        
        ctk.CTkLabel(win, text="Campo:").pack(anchor="w", padx=10)
        cmb_f = ctk.CTkComboBox(win, values=["-"], width=420); cmb_f.pack(padx=10, pady=(0, 10))

        # --- Flexibilidade para Campos com "Origem" ---
        flags_frame = ctk.CTkFrame(win, fg_color="transparent")
        flags_frame.pack(pady=4, anchor="w", padx=10)
        
        is_ro = f.readonly if f.origin_field else True
        is_req = f.required if f.origin_field else False

        var_force_ro = tk.BooleanVar(value=is_ro)
        chk_ro = ctk.CTkCheckBox(flags_frame, text="Manter como 'Só Leitura' (Recomendado)", variable=var_force_ro)
        chk_ro.pack(anchor="w")

        var_force_req = tk.BooleanVar(value=is_req)
        chk_req = ctk.CTkCheckBox(flags_frame, text="Manter como 'Obrigatório'", variable=var_force_req)
        chk_req.pack(anchor="w", pady=(6,0))

        def refresh_fields(*_):
            tname = cmb_t.get()
            t = next((t for t in self.project.tasks if t.name == tname), None)
            if not t: return
            names = [fld.name for fld in t.fields] if t else ["-"]
            cmb_f.configure(values=names); cmb_f.set(names[0] if names else "-")
        cmb_t.configure(command=lambda *_: refresh_fields()); refresh_fields()

        btns = ctk.CTkFrame(win, fg_color="transparent"); btns.pack(fill="x", padx=10, pady=10)
        def ok():
            self._commit_active_edits()
            sel_t = cmb_t.get(); target_tid = next((t.id for t in self.project.tasks if t.name == sel_t), None)
            sel_f = cmb_f.get()
            if not target_tid or not sel_f or sel_f == "-":
                clear(); return

            t = self._get_task(target_tid)
            if not t: return
            origin_field = next((fld for fld in t.fields if fld.name == sel_f), None)
            if not origin_field:
                messagebox.showwarning("Origem", "Campo de origem não encontrado.", parent=win); return
            
            self._push_undo()
            f.origin_task = target_tid
            f.origin_field = origin_field.id
            if f.name_lock_reason != "origem": f.name_before_origin = f.name
            f.name = origin_field.name
            f.ftype = origin_field.ftype
            f.options = origin_field.options
            f.name_lock_reason = "origem"
            f.name_locked = True
            
            f.readonly = var_force_ro.get()
            f.required = var_force_req.get()

            self._rebuild_metadata_cache()
            win.destroy(); self._update_single_row_widgets(f)
            
        def clear():
            self._push_undo()
            if f.name_lock_reason == "origem" and f.name_before_origin: f.name = f.name_before_origin
            f.origin_task=None; f.origin_field=None
            f.name_lock_reason = "" if f.name_lock_reason == "origem" else f.name_lock_reason
            f.name_locked = (f.name_lock_reason != ""); 
            self._rebuild_metadata_cache()
            win.destroy(); self._update_single_row_widgets(f)

        ctk.CTkButton(btns, text="Limpar", width=110, command=clear).pack(side="left", padx=(0, 8))
        ctk.CTkButton(btns, text="OK", width=140, command=ok).pack(side="right")

    def _list_options_for_field(self, fld: Field) -> List[str]:
        if fld.ftype in LIST_FIELD_TYPES:
            opts = [o.strip() for o in (fld.options or "").split(";") if o.strip()]
            return opts
        return []

    def open_cond_builder(self, f: Field):
        self._commit_row_data(f)
        win = ctk.CTkToplevel(self); win.title("Regras — quando este campo aparece")
        win.transient(self) # Fix para múltiplos monitores
        self._center_toplevel(win, 740, 540)
        win.grab_set()

        win.grid_rowconfigure(1, weight=1); win.grid_columnconfigure(0, weight=1)

        cur = self._get_task()
        if not cur: return
        preview = ctk.CTkLabel(win, text=self._cond_summary(f), anchor="w")
        preview.grid(row=0, column=0, sticky="ew", padx=10, pady=(10, 4))

        listfrm = ctk.CTkScrollableFrame(win)
        listfrm.grid(row=1, column=0, sticky="nsew", padx=10, pady=(6, 4))

        def render_list():
            for w in listfrm.winfo_children(): w.destroy()
            if not f.cond: ctk.CTkLabel(listfrm, text="Sem regras").pack(pady=8)
            preview.configure(text=self._cond_summary(f))
            for i, c in enumerate(f.cond):
                base_name = self._get_field_name(c.src_field) or f"[id {c.src_field}]"
                row = ctk.CTkFrame(listfrm, fg_color="transparent")
                row.pack(fill="x", pady=4)
                ctk.CTkLabel(row, text=f"Exibir quando {base_name} {c.op} {c.value!r}", anchor="w").pack(side="left")
                ctk.CTkButton(row, text="x", width=34, command=lambda idx=i: (self._push_undo(), f.cond.pop(idx), render_list())).pack(side="right")
        render_list()

        newfrm = ctk.CTkFrame(win)
        newfrm.grid(row=2, column=0, sticky="ew", padx=10, pady=(8, 6))
        newfrm.grid_columnconfigure(1, weight=1)

        ctk.CTkLabel(newfrm, text="Campo base:").grid(row=0, column=0, sticky="w")
        base_names = [fld.name for fld in cur.fields if fld.id != f.id] or ["-"]
        cmb_base = ctk.CTkComboBox(newfrm, values=base_names, width=380)
        cmb_base.grid(row=0, column=1, sticky="ew", padx=(6, 0))
        cmb_base.set(base_names[0] if base_names else "-")

        ctk.CTkLabel(newfrm, text="Valor igual a:").grid(row=1, column=0, sticky="w", pady=(6, 2))
        value_container = ctk.CTkFrame(newfrm)
        value_container.grid(row=1, column=1, sticky="ew", padx=(6, 0))

        value_var = tk.StringVar(); value_widget: Optional[tk.Widget] = None
        def current_base_field() -> Optional[Field]:
            sel = cmb_base.get()
            for fld in cur.fields:
                if fld.name == sel: return fld

        def render_value_input():
            nonlocal value_widget
            for w in value_container.winfo_children(): w.destroy()
            fld = current_base_field(); value_widget = None
            if not fld: return
            opts = self._list_options_for_field(fld)
            if opts:
                value_var.set(opts[0])
                om = ctk.CTkOptionMenu(value_container, values=opts, variable=value_var, command=lambda *_: None)
                base = _solid_color()
                om.configure(fg_color=base, button_color=base, button_hover_color=base); om.pack(fill="x")
                value_widget = om
            else:
                entry = ctk.CTkEntry(value_container); entry.pack(fill="x"); value_widget = entry

        cmb_base.configure(command=lambda *_: render_value_input()); render_value_input()

        def add_cond():
            fld = current_base_field()
            if not fld: messagebox.showwarning("Regras", "Escolha um campo base.", parent=win); return
            if isinstance(value_widget, ctk.CTkOptionMenu): val = value_var.get().strip()
            else: val = value_container.winfo_children()[0].get().strip() if value_container.winfo_children() else ""
            if not val: messagebox.showwarning("Regras", "Informe um valor.", parent=win); return
            if any(c.src_field == fld.id and c.op == "==" and c.value == val for c in f.cond):
                messagebox.showinfo("Regras", "Esta regra já existe.", parent=win); return
            self._push_undo()
            f.cond.append(Condition(src_field=fld.id, op="==", value=val)); render_list()

        btn_row = ctk.CTkFrame(win, fg_color="transparent")
        btn_row.grid(row=3, column=0, sticky="ew", padx=10, pady=(2, 10))
        ctk.CTkButton(btn_row, text="Adicionar regra", width=160, command=add_cond).pack(side="right")
        ctk.CTkButton(btn_row, text="Fechar", width=140, command=lambda: (win.destroy(), self._update_single_row_widgets(f))).pack(side="left")

    # ===== Validador de Fluxo =====
    def _validation_issue_key(self, task: Task, field: Field, code: str) -> str:
        task_id = getattr(task, "id", "") or ""
        field_id = getattr(field, "id", "") or ""
        return f"{code}|{task_id}|{field_id}"

    def _run_validation(self) -> List[Tuple[Task, Field, str, str]]:
        """Executa a lógica de validação e retorna uma lista de problemas."""
        issues = []
        all_fields: Dict[str, Field] = {f.id: f for t in self.project.tasks for f in t.fields}

        for task in self.project.tasks:
            field_names_in_task: Set[str] = set()
            for field in task.fields:
                if not field.required and not field.readonly:
                    issues.append((task, field, "OPT_EDIT", f"O campo é opcional e editável. É a configuração desejada?"))
                if field.required and field.readonly and not field.origin_field:
                    issues.append((task, field, "REQ_RO", f"O campo é obrigatório e 'Só Leitura', mas não tem origem. O usuário não poderá preenchê-lo."))
                if field.ftype in LIST_FIELD_TYPES and not (field.options or "").strip() and not field.origin_field:
                    issues.append((task, field, "LIST_NO_OPTS", f"O campo é do tipo '{field.ftype}', mas não tem opções definidas."))
                if field.ftype == "Informativo" and not (field.options or "").strip():
                    issues.append((task, field, "INFO_NO_TEXT", f"O campo é 'Informativo', mas não tem texto para exibir."))
                if field.ftype == "Anexo" and "[Tipo de Doc.:" not in field.note:
                    issues.append((task, field, "ANX_NO_TYPE", f"Sugestão: defina os tipos de documento para este anexo."))
                if (field.name or "").strip().lower() == "novo campo" or not (field.name or "").strip():
                    issues.append((task, field, "NO_NAME", f"O campo não tem um nome definido."))
                if field.name in field_names_in_task:
                    issues.append((task, field, "DUP_NAME", f"Existe mais de um campo com este nome na mesma tarefa."))
                field_names_in_task.add(field.name)
                if field.origin_field and field.origin_field not in all_fields:
                    issues.append((task, field, "BAD_ORIGIN", f"A origem deste campo aponta para um campo que foi excluído."))
                for cond in field.cond:
                    if cond.src_field not in all_fields:
                        issues.append((task, field, "BAD_RULE", f"Uma regra de visibilidade depende de um campo que foi excluído."))
                        break
        
        has_obj_field = any(f.ftype == "Objeto" for t in self.project.tasks for f in t.fields)
        if has_obj_field and not self.project.object_type:
            pseudo_task = self.project.tasks[0] if self.project.tasks else Task(id="", name="N/A")
            issues.append((pseudo_task, Field(id=""), "OBJ_NO_TYPE", "Existem campos 'Objeto', mas o tipo de objeto do fluxo não foi definido."))
        
        return issues


    def open_flow_validator(self):
        self.focus_set()
        self.update_idletasks()

        initial_issues = self._run_validation()
        
        if not initial_issues:
            messagebox.showinfo("Validação", "Verificação concluída. Nenhum aviso encontrado.")
            return

        win = ctk.CTkToplevel(self); win.title("Validador de Fluxo")
        win.transient(self)
        self._center_toplevel(win, 900, 600)
        win.grab_set()
        
        win.grid_columnconfigure(0, weight=1)

        header_label = ctk.CTkLabel(win, text=f"{len(initial_issues)} avisos encontrados:", anchor="w")
        header_label.grid(row=0, column=0, padx=10, pady=(10, 4), sticky="ew")
        
        list_frame = ctk.CTkScrollableFrame(win)
        list_frame.grid(row=1, column=0, padx=10, pady=(0, 10), sticky="nsew")

        footer_frame = ctk.CTkFrame(win, fg_color="transparent")
        footer_frame.grid(row=2, column=0, padx=10, pady=10, sticky="ew")
        
        ctk.CTkButton(footer_frame, text="Fechar", width=120, command=lambda: (win.destroy(), self._refresh_rows())).pack(side="right")
        reset_ignored_btn = ctk.CTkButton(
            footer_frame,
            text="Reexibir ignorados",
            width=160,
            command=lambda: (self.validator_ignored.clear(), render_issues()),
        )

        # Apenas a lista de avisos deve expandir
        win.grid_rowconfigure(1, weight=1)

        def render_issues():
            raw_issues = self._run_validation()

            for w in list_frame.winfo_children():
                w.destroy()

            visible_issues: List[Tuple[Task, Field, str, str, str]] = []
            ignored_count = 0
            for task, field, code, message in raw_issues:
                issue_key = self._validation_issue_key(task, field, code)
                if issue_key in self.validator_ignored:
                    ignored_count += 1
                    continue
                visible_issues.append((task, field, code, message, issue_key))

            header_text = f"{len(visible_issues)} avisos restantes:"
            if ignored_count:
                header_text += f" ({ignored_count} ignorado(s))"
            header_label.configure(text=header_text)

            if self.validator_ignored:
                if not reset_ignored_btn.winfo_manager():
                    reset_ignored_btn.pack(side="left", padx=(0, 10))
            elif reset_ignored_btn.winfo_manager():
                reset_ignored_btn.pack_forget()

            if not visible_issues:
                msg = "Todos os avisos foram resolvidos!"
                if ignored_count:
                    msg = "Todos os avisos foram resolvidos ou ignorados."
                ctk.CTkLabel(list_frame, text=msg).pack(pady=20)
                return

            def exec_and_refresh(action: Callable[[], Any], *, push_undo: bool = False, rebuild_cache: bool = False):
                self._commit_active_edits()
                if push_undo:
                    self._push_undo()
                try:
                    action()
                finally:
                    if rebuild_cache:
                        self._rebuild_metadata_cache()
                    self._refresh_task_combo()
                    self._refresh_rows()
                    try:
                        self.update_idletasks()
                    except Exception:
                        pass
                    render_issues()

            for task, field, code, message, issue_key in visible_issues:
                row = ctk.CTkFrame(list_frame, fg_color="#1e1e1e", corner_radius=8)
                row.pack(fill="x", pady=4, padx=6)
                row.grid_columnconfigure(0, weight=1)

                info_frame = ctk.CTkFrame(row, fg_color="transparent")
                info_frame.grid(row=0, column=0, sticky="nsew", padx=8, pady=8)

                actions_frame = ctk.CTkFrame(row, fg_color="transparent")
                actions_frame.grid(row=0, column=1, sticky="ne", padx=8, pady=8)

                ctk.CTkLabel(info_frame, text=f"Tarefa: {task.name}", anchor="w").pack(fill="x", pady=(0, 2))
                if field.id:
                    ctk.CTkLabel(info_frame, text=f"Campo: {field.name}", anchor="w", font=("Segoe UI", 12, "bold")).pack(fill="x", pady=(0, 4))
                ctk.CTkLabel(info_frame, text=message, anchor="w", wraplength=520, justify="left").pack(fill="x")

                button_wrapper = ctk.CTkFrame(actions_frame, fg_color="transparent")
                button_wrapper.pack(anchor="ne")

                action_buttons: List[Tuple[str, Callable[[], Any]]] = []

                if field.id:
                    def go_to_field(t_id=task.id, f_id=field.id):
                        self._focus_on_field(t_id, f_id if f_id else None)
                        win.after(200, win.lift)

                    action_buttons.append(("Ir para o campo", go_to_field))

                if code == "OPT_EDIT":
                    action_buttons.append((
                        "Tornar Obrig.",
                        lambda f=field: exec_and_refresh(lambda: setattr(f, 'required', True), push_undo=True),
                    ))
                    action_buttons.append((
                        "Tornar Só Leit.",
                        lambda f=field: exec_and_refresh(lambda: setattr(f, 'readonly', True), push_undo=True),
                    ))
                elif code == "REQ_RO":
                    action_buttons.append((
                        "Permitir Edição",
                        lambda f=field: exec_and_refresh(lambda: setattr(f, 'readonly', False), push_undo=True),
                    ))
                elif code == "BAD_ORIGIN":
                    action_buttons.append((
                        "Limpar Origem",
                        lambda f=field: exec_and_refresh(
                            lambda: (setattr(f, 'origin_field', None), setattr(f, 'origin_task', None)),
                            push_undo=True,
                            rebuild_cache=True,
                        ),
                    ))
                elif code == "BAD_RULE":
                    action_buttons.append((
                        "Limpar Regras",
                        lambda f=field: exec_and_refresh(lambda: setattr(f, 'cond', []), push_undo=True),
                    ))
                elif code == "OBJ_NO_TYPE":
                    action_buttons.append((
                        "Definir Tipo...",
                        lambda: exec_and_refresh(self.open_flow_object_type_dialog),
                    ))

                if code == "ANX_NO_TYPE" and field.ftype == "Anexo" and field.id:
                    action_buttons.append((
                        "Definir tipos de doc...",
                        lambda f=field: exec_and_refresh(lambda: self._open_attachment_type_editor(f)),
                    ))

                action_buttons.append(("Ignorar aviso", lambda key=issue_key: (self.validator_ignored.add(key), render_issues())))

                for text, cmd in action_buttons:
                    ctk.CTkButton(button_wrapper, text=text, width=150, command=cmd).pack(fill="x", pady=2)

        render_issues()

    class SimWindow(ctk.CTkToplevel):
        def __init__(self, master: "App", project: ProjectModel, start_task_id: Optional[str] = None):
            super().__init__(master)
            self.master = master
            self.title("Simulador de Workflow")
            self.transient(master)  # Fix para múltiplos monitores
            self._center_toplevel(self, 900, 600, respect_req_size=False)
            self.grab_set()

            self.project = project
            self.answers: Dict[str, Any] = {}
            self.controller_field_ids: Set[str] = self._collect_controller_fields()
            top = ctk.CTkFrame(self, fg_color="transparent"); top.pack(side="top", fill="x", padx=10, pady=8)
            ctk.CTkLabel(top, text="Tarefa:").pack(side="left", padx=(0, 6))

            display_names = [f"{i+1}. {t.name}" for i, t in enumerate(project.tasks)]
            self.cmb = ctk.CTkComboBox(top, values=display_names, width=380, command=lambda *_: self._render())
            self.cmb.pack(side="left")

            start_task_display_name = ""
            start_task_idx = -1
            if start_task_id:
                for i, task in enumerate(project.tasks):
                    if task.id == start_task_id: start_task_idx = i; break

            if start_task_idx != -1: start_task_display_name = display_names[start_task_idx]
            elif display_names: start_task_display_name = display_names[0]

            self.cmb.set(start_task_display_name)

            ctk.CTkButton(top, text="Anterior", width=120, command=lambda: self._step(-1)).pack(side="left", padx=6)
            ctk.CTkButton(top, text="Próxima", width=120, command=lambda: self._step(+1)).pack(side="left")
            ctk.CTkButton(top, text="Escolher...", width=140, command=self._goto_dialog).pack(side="left", padx=6)

            self.body = ctk.CTkScrollableFrame(self); self.body.pack(fill="both", expand=True, padx=10, pady=(0, 10))
            self._render()

        def _collect_controller_fields(self) -> Set[str]:
            controllers: Set[str] = set()
            try:
                for task in self.project.tasks:
                    for field in task.fields:
                        for cond in getattr(field, "cond", []) or []:
                            src = getattr(cond, "src_field", None)
                            if src:
                                controllers.add(src)
            except Exception:
                pass
            return controllers

        def _maybe_rerender(self, field_id: Optional[str]) -> None:
            if field_id and field_id in self.controller_field_ids:
                self._render()

        def _center_toplevel(self, win, width, height, *, respect_req_size: bool = True):
            if hasattr(self.master, "_center_toplevel"):
                self.master._center_toplevel(win, width, height, respect_req_size=respect_req_size)

        def on_model_changed(self):
            current_ids: Set[str] = {f.id for t in self.project.tasks for f in t.fields}
            self.controller_field_ids = self._collect_controller_fields()

            new_answers: Dict[str, Any] = {}
            for k, v in self.answers.items():
                field_id_base = k.split("__", 1)[0]
                if field_id_base in current_ids: new_answers[k] = v
            self.answers = new_answers

            display_names = [f"{i+1}. {t.name}" for i, t in enumerate(self.project.tasks)]
            self.cmb.configure(values=display_names)

            try:
                current_idx = self._cur_index()
                if 0 <= current_idx < len(display_names): self.cmb.set(display_names[current_idx])
                elif display_names: self.cmb.set(display_names[0])
                else: self.cmb.set("")
            except Exception:
                if display_names: self.cmb.set(display_names[0])
                else: self.cmb.set("")

            self._render()

        def _cur_index(self) -> int:
            selected_display_name = self.cmb.get()
            if not selected_display_name: return 0
            try: return int(selected_display_name.split('.')[0]) - 1
            except (ValueError, IndexError): return 0

        def _step(self, delta: int):
            i = self._cur_index() + delta; i = max(0, min(len(self.project.tasks) - 1, i))
            display_names = self.cmb.cget("values")
            if 0 <= i < len(display_names): self.cmb.set(display_names[i]); self._render()

        def _goto_dialog(self):
            win = ctk.CTkToplevel(self); win.title("Ir para a tarefa...")
            win.transient(self) # Fix para múltiplos monitores
            self._center_toplevel(win, 460, 460);
            win.grab_set()

            ctk.CTkLabel(win, text="Buscar:").pack(anchor="w", padx=10, pady=(10, 4))
            e = ctk.CTkEntry(win, width=380); e.pack(padx=10)
            lst = ctk.CTkScrollableFrame(win); lst.pack(fill="both", expand=True, padx=10, pady=10)
            def render():
                for w in lst.winfo_children(): w.destroy()
                q = (e.get() or "").lower()
                for i, t in enumerate(self.project.tasks):
                    display_name = f"{i+1}. {t.name}"
                    if q in t.name.lower():
                        ctk.CTkButton(lst, text=display_name, command=lambda n=display_name: (win.destroy(), self.cmb.set(n), self._render())).pack(fill="x", pady=2)
            e.bind("<KeyRelease>", lambda *_: render()); render()
            ctk.CTkButton(win, text="Cancelar", command=win.destroy).pack(side="left", padx=10, pady=(0, 10))

        def _is_visible(self, f: Field) -> bool:
            if not f.cond: return True
            for c in f.cond:
                val = self.answers.get(c.src_field)
                if val is None or c.op != "==":
                    continue
                if isinstance(val, (list, tuple, set)):
                    if c.value in {str(x) for x in val}:
                        return True
                else:
                    if str(val) == c.value:
                        return True
            return False

        def _render(self):
            for w in self.body.winfo_children(): w.destroy()
            task = None
            try:
                task_idx = self._cur_index()
                if 0 <= task_idx < len(self.project.tasks): task = self.project.tasks[task_idx]
            except Exception: pass

            if not task: return
            visible_fields = [f for f in task.fields if self._is_visible(f)]
            if not visible_fields:
                ctk.CTkLabel(self.body, text="Nenhum campo visível nesta tarefa com as respostas atuais.").pack(pady=10)
                return

            def build_multiselect(parent, options, answer_key, *, instruction=True, width=220, expand=False):
                if not options:
                    ctk.CTkLabel(parent, text="(Sem opções definidas)", text_color="#9aa4b2").pack(side="left")
                    return

                pack_opts = {"side": "left", "padx": (0, 8)}
                if expand:
                    pack_opts.update({"fill": "x", "expand": True})

                container = ctk.CTkFrame(parent, fg_color="#1f2937", corner_radius=6)
                container.pack(**pack_opts)

                chips_frame = ctk.CTkFrame(container, fg_color="transparent")
                chips_frame.pack(fill="x", expand=True, padx=8, pady=(8, 4))
                chips_frame.grid_columnconfigure(0, weight=1)

                raw_current = self.answers.get(answer_key, [])
                if isinstance(raw_current, str):
                    current_values = [raw_current] if raw_current else []
                else:
                    current_values = [str(v) for v in raw_current if str(v) in options]

                state = {"values": current_values[:]}
                chip_widgets: Dict[str, ctk.CTkButton] = {}

                def update_styles():
                    for opt, btn in chip_widgets.items():
                        if opt in state["values"]:
                            btn.configure(fg_color="#2563eb", text_color="#f8fafc")
                        else:
                            btn.configure(fg_color="#374151", text_color="#e2e8f0")

                def commit_answers():
                    if state["values"]:
                        self.answers[answer_key] = state["values"][:]
                    else:
                        self.answers.pop(answer_key, None)

                def on_click(event, opt):
                    ctrl_pressed = bool(event.state & 0x0004)
                    values = [v for v in state["values"] if v in options]
                    if ctrl_pressed:
                        if opt in values:
                            values = [v for v in values if v != opt]
                        else:
                            values.append(opt)
                    else:
                        values = [opt]

                    state["values"] = values
                    commit_answers()
                    update_styles()
                    self._maybe_rerender(answer_key)
                    return "break"

                for idx, opt in enumerate(options):
                    btn = ctk.CTkButton(
                        chips_frame,
                        text=opt,
                        width=width,
                        height=32,
                        fg_color="#374151",
                        hover_color="#4b5563",
                        text_color="#e2e8f0",
                        corner_radius=8,
                    )
                    btn.grid(row=idx, column=0, sticky="ew", pady=4)
                    btn.bind("<Button-1>", lambda e, o=opt: on_click(e, o))
                    chip_widgets[opt] = btn

                update_styles()

                if instruction:
                    ctk.CTkLabel(
                        container,
                        text="Segure Ctrl para selecionar várias opções.",
                        text_color="#9aa4b2",
                        anchor="w",
                        justify="left",
                    ).pack(fill="x", padx=8, pady=(0, 8))

            for f in visible_fields:
                row = ctk.CTkFrame(self.body, fg_color="transparent"); row.pack(fill="x", pady=8)
                name_text = f.name
                if f.required: name_text += " (*Obrigatório*)"
                if f.readonly: name_text += " [Só leitura]"
                ctk.CTkLabel(row, text=name_text, width=320, anchor="w", fg_color="transparent", justify="left").pack(side="left")

                if not f.readonly:
                    if f.ftype in LIST_FIELD_TYPES:
                        opts = self.master._list_options_for_field(f)

                        if f.ftype in MULTISELECT_FIELD_TYPES:
                            build_multiselect(row, opts, f.id, instruction=True, width=240)
                        else:
                            placeholder = "-- Selecione --"
                            if not opts:
                                om = ctk.CTkOptionMenu(row, values=["(Sem opções definidas)"], state="disabled")
                                om.pack(side="left")
                            else:
                                display_opts = [placeholder] + opts
                                var = tk.StringVar(value=self.answers.get(f.id, placeholder))

                                def on_change(value, fid=f.id):
                                    if value == placeholder:
                                        self.answers.pop(fid, None)
                                    else:
                                        self.answers[fid] = value
                                    self._maybe_rerender(fid)

                                om = ctk.CTkOptionMenu(
                                    row,
                                    values=display_opts,
                                    variable=var,
                                    command=lambda val, fid=f.id: on_change(val, fid),
                                )
                                base = _solid_color()
                                om.configure(fg_color=base, button_color=base, button_hover_color=base)
                                om.pack(side="left")

                    elif f.ftype == "Numérico":
                        prev_value = str(self.answers.get(f.id, ""))
                        num_var = tk.StringVar(value=prev_value)

                        def validate_numeric(new_value: str) -> bool:
                            if new_value == "":
                                return True
                            return bool(re.fullmatch(r"-?\d*(?:[.,]\d*)?", new_value))

                        validate_numeric_cmd = self.register(validate_numeric)

                        ent = ctk.CTkEntry(
                            row,
                            width=200,
                            textvariable=num_var,
                            validate="key",
                            validatecommand=(validate_numeric_cmd, "%P"),
                            justify="right",
                        )
                        ent.pack(side="left")

                        def on_numeric_focus_out(_event=None, fid=f.id, var=num_var):
                            value = var.get().strip()
                            if not value:
                                self.answers.pop(fid, None)
                                return
                            if not re.fullmatch(r"-?\d*(?:[.,]\d+)?", value):
                                messagebox.showwarning(
                                    "Valor inválido",
                                    "Informe apenas números (use vírgula ou ponto para decimais).",
                                    parent=self,
                                )
                                previous = str(self.answers.get(fid, ""))
                                var.set(previous)
                                if previous:
                                    self.answers[fid] = previous
                                else:
                                    self.answers.pop(fid, None)
                                return
                            self.answers[fid] = value

                        ent.bind("<FocusOut>", on_numeric_focus_out)

                    elif f.ftype == "Data":
                        prev_value = str(self.answers.get(f.id, ""))
                        date_var = tk.StringVar(value=prev_value)

                        def validate_date_input(new_value: str) -> bool:
                            return bool(re.fullmatch(r"[0-9/]{0,10}", new_value))

                        validate_date_cmd = self.register(validate_date_input)

                        ent = ctk.CTkEntry(
                            row,
                            width=220,
                            textvariable=date_var,
                            validate="key",
                            validatecommand=(validate_date_cmd, "%P"),
                        )
                        ent.pack(side="left")

                        mask_state = {"updating": False}

                        def apply_date_mask(*_args, var=date_var):
                            if mask_state["updating"]:
                                return
                            mask_state["updating"] = True
                            digits = re.sub(r"\D", "", var.get())[:8]
                            if not digits:
                                formatted = ""
                            elif len(digits) <= 2:
                                formatted = digits
                            elif len(digits) <= 4:
                                formatted = f"{digits[:2]}/{digits[2:]}"
                            else:
                                formatted = f"{digits[:2]}/{digits[2:4]}/{digits[4:]}"
                            var.set(formatted)
                            mask_state["updating"] = False

                        date_var.trace_add("write", apply_date_mask)
                        apply_date_mask()

                        def on_date_focus_out(_event=None, fid=f.id, var=date_var):
                            raw = var.get().strip()
                            if not raw:
                                self.answers.pop(fid, None)
                                return
                            digits = re.sub(r"\D", "", raw)
                            formatted = raw
                            if len(digits) == 8:
                                formatted = f"{digits[0:2]}/{digits[2:4]}/{digits[4:8]}"
                            try:
                                datetime.datetime.strptime(formatted, "%d/%m/%Y")
                            except ValueError:
                                messagebox.showwarning(
                                    "Data inválida",
                                    "Use o formato DD/MM/AAAA.",
                                    parent=self,
                                )
                                previous = str(self.answers.get(fid, ""))
                                var.set(previous)
                                if previous:
                                    self.answers[fid] = previous
                                else:
                                    self.answers.pop(fid, None)
                                return
                            var.set(formatted)
                            self.answers[fid] = formatted
                            self._maybe_rerender(fid)

                        ent.bind("<FocusOut>", on_date_focus_out)

                    elif f.ftype == "Anexo":
                        card = ctk.CTkFrame(row, fg_color="#222222", corner_radius=6); card.pack(side="left", padx=4, fill="x", expand=True)
                        current_value = self.answers.get(f.id, ""); value_var = tk.StringVar(value=current_value or "Nenhum arquivo selecionado")
                        def choose_file(fid=f.id, var=value_var):
                            path = filedialog.askopenfilename()
                            if path: self.answers[fid] = path; var.set(path)
                        ctk.CTkButton(card, text="Selecionar arquivo...", command=choose_file, width=220,).pack(anchor="w", padx=8, pady=(6, 4))
                        ctk.CTkLabel(card, textvariable=value_var, anchor="w", text_color="#c5cdd9", justify="left",).pack(anchor="w", padx=8)
                        ctk.CTkLabel(card, text="Apenas visualização no simulador.", anchor="w", text_color="#9aa4b2",).pack(anchor="w", padx=8, pady=(2, 6))

                    elif f.ftype == "Área de texto":
                        tb = ctk.CTkTextbox(row, width=500, height=90); tb.insert("1.0", self.answers.get(f.id, "")); tb.pack(side="left")
                        tb.bind("<FocusOut>", lambda e, fid=f.id, w=tb: self.answers.__setitem__(fid, w.get("1.0", tk.END).strip()))

                    elif f.ftype == "Componente do sistema":
                        comp_frame = ctk.CTkFrame(row, fg_color="#222222", corner_radius=6); comp_frame.pack(side="left", padx=4, fill="x", expand=True)
                        ctk.CTkLabel(comp_frame, text=f.note or "Simular valor de saída do componente", anchor="w").pack(padx=8, pady=(6, 2), anchor="w")
                        ent = ctk.CTkEntry(comp_frame); ent.insert(0, self.answers.get(f.id, "")); ent.pack(fill="x", padx=8, pady=(0, 6))
                        ent.bind("<FocusOut>", lambda e, fid=f.id, w=ent: (self.answers.__setitem__(fid, w.get()), self._maybe_rerender(fid)))

                    elif f.ftype == "Objeto":
                        obj_box = ctk.CTkFrame(row, fg_color="#222222", corner_radius=6); obj_box.pack(side="left", padx=4, fill="x", expand=True)
                        title = self.project.object_type or "Objeto"
                        ctk.CTkLabel(obj_box, text=title, anchor="center", justify="center").pack(padx=8, pady=(6, 2), fill="x")

                        for ofd in self.project.object_schema:
                            if not ofd.readonly:
                                sub_key = f"{f.id}__{ofd.name}"; sub_row = ctk.CTkFrame(obj_box, fg_color="transparent"); sub_row.pack(fill="x", padx=8, pady=2)
                                sub_name = ofd.name
                                if ofd.required: sub_name += " (*Obrigatório*)"
                                ctk.CTkLabel(sub_row, text=sub_name, width=180, anchor="w", justify="left").pack(side="left")

                                if ofd.ftype in LIST_FIELD_TYPES:
                                    opts = [o.strip() for o in (ofd.options or "").split(";") if o.strip()]

                                    if ofd.ftype in MULTISELECT_FIELD_TYPES:
                                        build_multiselect(sub_row, opts, sub_key, instruction=False, width=200, expand=True)
                                    else:
                                        if not opts:
                                            om = ctk.CTkOptionMenu(sub_row, values=["(Sem opções)"], state="disabled")
                                            om.pack(side="left", fill="x", expand=True)
                                        else:
                                            display_opts = ["-- Selecione --"] + opts
                                            var = tk.StringVar(value=self.answers.get(sub_key, "-- Selecione --"))

                                            def on_sub_change(value, skey=sub_key):
                                                if value == "-- Selecione --":
                                                    self.answers.pop(skey, None)
                                                else:
                                                    self.answers[skey] = value
                                                self._maybe_rerender(skey)

                                            om = ctk.CTkOptionMenu(
                                                sub_row,
                                                values=display_opts,
                                                variable=var,
                                                command=lambda val, skey=sub_key: on_sub_change(val, skey),
                                            )
                                            om.pack(side="left", fill="x", expand=True)
                                else:
                                    ent = ctk.CTkEntry(sub_row); ent.insert(0, self.answers.get(sub_key, "")); ent.pack(side="left", fill="x", expand=True)
                                    ent.bind("<FocusOut>", lambda e, skey=sub_key, w=ent: (self.answers.__setitem__(skey, w.get()), self._maybe_rerender(skey)))

                        if not self.project.object_schema: ctk.CTkLabel(obj_box, text="Conteúdo mapeado externamente (sem esquema).", anchor="w").pack(padx=8, pady=(0, 6))
                        elif all(ofd.readonly for ofd in self.project.object_schema): ctk.CTkLabel(obj_box, text="Todos os campos do Objeto são 'Só Leitura'.", anchor="w").pack(padx=8, pady=(0, 6))

                    else:
                        ent = ctk.CTkEntry(row, width=500); ent.insert(0, self.answers.get(f.id, "")); ent.pack(side="left")
                        ent.bind("<FocusOut>", lambda e, fid=f.id, w=ent: self.answers.__setitem__(fid, w.get()))

                else:
                    if f.ftype == "Informativo":
                        ctk.CTkLabel(row, text=f.options or f.info or "Informativo", anchor="w", justify="left").pack(side="left")
                    elif f.ftype == "Componente do sistema":
                        ctk.CTkLabel(row, text="Componente (Valor é gerado/externo)", anchor="w").pack(side="left")
                        ctk.CTkEntry(row, width=260, state="disabled").pack(side="left")
                    elif f.ftype == "Objeto":
                        obj_box = ctk.CTkFrame(row, fg_color="#222222", corner_radius=6); obj_box.pack(side="left", padx=4, fill="x", expand=True)
                        title = self.project.object_type or "Objeto"; ctk.CTkLabel(obj_box, text=title, anchor="center", justify="center").pack(padx=8, pady=(6, 2), fill="x")
                        for ofd in self.project.object_schema:
                             sub_val = self.answers.get(f"{f.id}__{ofd.name}", "")
                             if isinstance(sub_val, (list, tuple, set)):
                                 display_val = ", ".join(str(v) for v in sub_val) or "(Não preenchido)"
                             else:
                                 display_val = sub_val or "(Não preenchido)"
                             ctk.CTkLabel(obj_box, text=f"- {ofd.name}: {display_val}", anchor="w", text_color="#9aa4b2").pack(anchor="w", padx=8)
                    elif f.ftype == "Anexo":
                        ctk.CTkLabel(row, text="Anexo (Só Leitura)", anchor="w").pack(side="left")
                        ctk.CTkEntry(row, width=260, state="disabled").pack(side="left")
                    elif f.ftype in MULTISELECT_FIELD_TYPES:
                        selected = self.answers.get(f.id, [])
                        if isinstance(selected, str):
                            values = [selected] if selected else []
                        else:
                            values = [str(v) for v in selected]
                        display = ", ".join(values) if values else "(Nenhuma opção selecionada)"
                        ctk.CTkLabel(row, text=display, anchor="w", justify="left").pack(side="left")
                    elif f.ftype in LIST_FIELD_TYPES:
                        display = str(self.answers.get(f.id, "")) or "(Não selecionado)"
                        ctk.CTkLabel(row, text=display, anchor="w", justify="left").pack(side="left")
                    else:
                        ctk.CTkEntry(row, width=260, state="disabled").pack(side="left")

    def open_simulator(self):
        self._commit_active_edits()
        self._commit_all_visible_row_edits()
        if self.sim_window and self.sim_window.winfo_exists():
            try: self.sim_window.focus_set()
            except Exception: pass
            try: self.sim_window.on_model_changed()
            except Exception: pass
            return
        self.sim_window = App.SimWindow(self, self.project, self.current_task_id)
    # ===== Util =====
    def _center_toplevel(self, win, width, height, *, transient=True, fade=False, respect_req_size=True):
        # 1. Configurações iniciais
        if transient and win is not self:
            try: win.transient(self)
            except: pass
        try: win.configure(fg_color=DARK_BG2)
        except: pass

        # 2. CORREÇÃO DUAL MONITOR: Centralizar relativo à janela principal
        try:
            # Pega as coordenadas e tamanho da janela mãe (App)
            root_x = self.winfo_rootx()
            root_y = self.winfo_rooty()
            root_w = self.winfo_width()
            root_h = self.winfo_height()

            # Calcula o centro relativo
            x = root_x + (root_w - width) // 2
            y = root_y + (root_h - height) // 2
        except:
            # Fallback de segurança para o centro da tela principal
            try:
                screen_w = win.winfo_screenwidth()
                screen_h = win.winfo_screenheight()
                x = (screen_w - width) // 2
                y = (screen_h - height) // 2
            except:
                x, y = 0, 0

        # 3. Aplica geometria e estilo
        win.geometry(f"{width}x{height}+{int(x)}+{int(y)}")

        # CORREÇÃO: Aplica a barra escura ENQUANTO a janela ainda está oculta
        # Isso evita o "flash" branco da barra de título do Windows
        try: self._set_dark_title_bar(win)
        except: pass

        # Só agora mostra a janela, já estilizada
        win.deiconify()
        win.attributes("-alpha", 1.0)

    def _finalize_toplevel_position(
        self,
        win: ctk.CTkToplevel,
        width: int,
        height: int,
        fade: bool,
        respect_req_size: bool,
    ) -> None:
        if not win.winfo_exists():
            return

        try:
            self.update_idletasks()
        except Exception:
            pass

        try:
            win.update_idletasks()

            required_w = win.winfo_reqwidth()
            required_h = win.winfo_reqheight()
            if respect_req_size:
                width = max(width, required_w)
                height = max(height, required_h)

            current_w = win.winfo_width()
            current_h = win.winfo_height()
            if current_w != width or current_h != height:
                win.geometry(f"{width}x{height}")
                win.update_idletasks()

            anchor = self if self.winfo_exists() else win
            bounds = _get_monitor_bounds_for_window(anchor)

            centered = _center_within(anchor, width, height)
            if centered is None:
                screen_w = win.winfo_screenwidth()
                screen_h = win.winfo_screenheight()
                x = (screen_w - width) // 2
                y = (screen_h - height) // 2
            else:
                x, y = centered

            x, y = _clamp_to_bounds(int(x), int(y), width, height, bounds)
        except Exception:
            try:
                screen_w = win.winfo_screenwidth()
                screen_h = win.winfo_screenheight()
            except Exception:
                screen_w = screen_h = width
            x = (screen_w - width) // 2
            y = (screen_h - height) // 2

        win.geometry(f"{width}x{height}+{x}+{y}")
        self._set_dark_title_bar(win)
        if fade:
            _animate_fade_in(win)
        elif getattr(win, "_centered_hidden_alpha", False):
            try:
                win.attributes("-alpha", 1.0)
            except Exception:
                pass
        if getattr(win, "_centered_hidden_alpha", False):
            setattr(win, "_centered_hidden_alpha", False)
        try:
            win.focus_force()
        except Exception:
            pass

    def _prompt_text(self, title: str, label: str, initial: str = "") -> Optional[str]:
        win = ctk.CTkToplevel(self); win.title(title)
        win.transient(self) # Fix para múltiplos monitores
        self._center_toplevel(win, 420, 160)
        win.grab_set()

        ctk.CTkLabel(win, text=label).pack(pady=(12,6))
        e = ctk.CTkEntry(win, width=360); e.pack(); e.insert(0, initial)
        out={"v":None}
        def ok():
            val = e.get()
            # Distingue entre 'cancelar' (None) e 'OK com texto vazio' ("")
            out["v"] = val if val is not None else ""
            win.destroy()
        def cancel():
            out["v"] = None
            win.destroy()
        
        btns = ctk.CTkFrame(win, fg_color="transparent")
        btns.pack(pady=10)
        ctk.CTkButton(btns, text="OK", command=ok, width=120).pack(side="left", padx=5)
        btn_cancel = ctk.CTkButton(btns, text="Cancelar", command=cancel, width=120)
        _apply_secondary_style(btn_cancel)
        btn_cancel.pack(side="left", padx=5)

        e.bind("<Return>", lambda event: ok())
        e.focus()
        self.wait_window(win)
        return out["v"]

    def _set_dark_title_bar(self, win: tk.Toplevel):
        """Força a barra de título da janela a usar o modo escuro no Windows."""
        _apply_dark_title_bar(win)

# ---- Main ----
if __name__ == "__main__":
    try:
        app = App()
        # --- SPLASH SCREEN END ---
        # Fecha a tela de carregamento do PyInstaller se ela estiver ativa
        if importlib.util.find_spec("pyi_splash"):
            import pyi_splash
            if pyi_splash.is_alive():
                pyi_splash.close()
        app.mainloop()
    except Exception:
        try:
            log_dir = _best_desktop_dir()
            log_path = os.path.join(log_dir, "designer_campos_error.log")
            with open(log_path, "w", encoding="utf-8") as f:
                f.write(traceback.format_exc())
            print("FALHA AO INICIAR APLICACAO:\n", traceback.format_exc(), file=sys.stderr)
            try:
                root = tk.Tk(); root.withdraw()
                messagebox.showerror("Erro ao iniciar", f"O app fechou inesperadamente.\n\nFoi salvo um log em:\n{log_path}")
            except Exception: pass
        except Exception:
            print(traceback.format_exc(), file=sys.stderr)